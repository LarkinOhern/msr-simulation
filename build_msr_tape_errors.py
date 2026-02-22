#!/usr/bin/env python3
"""
MSR Error Tape Generator
========================
Reads the clean Jan 2026 tape and injects realistic data transmission / entry
errors that a subservicer might submit.  Produces a dirty tape for validation
testing.

Hard Stops (fail immediately):
  1. UPB x 10 (extra zero)               — 3 loans
  2. UPB = $0  (active loan)             — 1 loan
  3. Missing loans (no PIF reported)     — 3 loans
  4. Duplicate loan ID                   — 1 loan
  5. Rate as whole number (6.50 vs .065) — 2 loans
  6. UPB > Original Balance              — 1 loan

Yellow Lights (flag for review):
  1. NSF expressed as percent (FNMA)     — 2 loans
  2. NSF expressed as whole bps (GNMA)  — 2 loans
  3. Status skip (Current -> 90+ DPD)   — 2 loans
  4. P&I inflated ~20%                   — 2 loans
  5. Next Due Date in past (current loan)— 2 loans
  6. Remaining term unchanged            — 1 loan

Usage:
    python build_msr_tape_errors.py
"""

import os
import random
from copy import deepcopy
from datetime import date
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(99)
OUT = os.path.dirname(os.path.abspath(__file__))

DISCLAIMER = (
    "SIMULATED DATA — All loan information is synthetic and generated "
    "for testing purposes only. Not representative of any real portfolio."
)
REPORT_DATE = date(2026, 1, 31)

# ── Column positions (matches build_msr_tape.py 16-col layout) ──────────────
CI = {
    "loan_id":   1,  "loan_type": 2,  "purpose":   3,  "investor":  4,
    "orig_date": 5,  "orig_bal":  6,  "upb":       7,  "rate":      8,
    "nsf":       9,  "rem_term": 10,  "maturity": 11,  "pi":       12,
    "escrow":   13,  "total_pmt":14,  "status":   15,  "ndd":      16,
}
N_COLS = 16

# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex6):  return PatternFill("solid", fgColor=hex6)
def _font(bold=False, color="000000", size=9, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def _side():      return Side(style="thin", color="B8CCE4")
def _border():    s = _side(); return Border(left=s, right=s, top=s, bottom=s)

THIN      = _border()
F_NAVY    = _fill("1F4E79")
F_BLUE    = _fill("2E75B6")
F_LTBLUE  = _fill("D6E4F0")
F_TOTAL   = _fill("BDD7EE")
F_DISC    = _fill("FFF2CC")
F_LOG_HDR = _fill("7B2C2C")   # error log header (dark red)
F_HARD    = _fill("FCE4D6")   # hard stop row tint
F_YELLOW  = _fill("FFEB9C")   # yellow light row tint

WHBOLD  = _font(bold=True,  color="FFFFFF", size=9)
BLKBOLD = _font(bold=True,  size=9)
NORMAL  = _font(size=9)
DISC_FNT= _font(bold=True,  color="7B3F00", size=9, italic=True)

CURR  = '#,##0.00'
CURR0 = '#,##0'
PCT3  = '0.000%'
NUM0  = '#,##0'
DFMT  = 'MM/DD/YYYY'

TAPE_HEADERS = [
    "Loan ID","Loan Type","Purpose","Investor",
    "Orig Date","Original Bal ($)","Current UPB ($)","Rate","Net Serv Fee","Rem Term",
    "Maturity","P&I ($)","Escrow ($)","Total Pmt ($)","Status","Next Due Date"
]

COL_FMTS = {
    "loan_id": None, "loan_type": None, "purpose": None, "investor": None,
    "orig_date": DFMT, "orig_bal": CURR0, "upb": CURR, "rate": PCT3,
    "nsf": PCT3, "rem_term": NUM0, "maturity": DFMT, "pi": CURR,
    "escrow": CURR, "total_pmt": CURR, "status": None, "ndd": DFMT,
}
COL_ALIGN = {
    "loan_id": "center", "loan_type": "center", "purpose": "center",
    "investor": "center", "orig_date": "center", "orig_bal": "right",
    "upb": "right", "rate": "right", "nsf": "right", "rem_term": "center",
    "maturity": "center", "pi": "right", "escrow": "right",
    "total_pmt": "right", "status": "center", "ndd": "center",
}


def set_col_widths(ws):
    widths = {
        "A":13,"B":13,"C":11,"D":10,"E":14,"F":16,"G":14,"H":13,
        "I":12,"J":11,"K":13,"L":13,"M":12,"N":13,"O":13,"P":13,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Read clean Jan 2026 tape ─────────────────────────────────────────────────
def read_jan_tape(filepath):
    """Read Jan 2026 sheet from the clean tape. Returns list of row dicts."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if "jan" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    rows = []
    r = 4  # DATA_START (after disclaimer + title + headers)
    while True:
        loan_id = ws.cell(row=r, column=CI["loan_id"]).value
        if loan_id is None or str(loan_id).upper().startswith("TOTAL"):
            break
        row_dict = {field: ws.cell(row=r, column=col).value
                    for field, col in CI.items()}
        rows.append(row_dict)
        r += 1

    print(f"  Read {len(rows):,} loans from '{ws.title}'")
    return rows


# ── Write tape sheet ──────────────────────────────────────────────────────────
def write_tape_sheet(ws, title_text, title_fill, rows):
    """Disclaimer + title + headers + data rows + totals row."""
    last_col = get_column_letter(N_COLS)

    # Row 1: disclaimer
    ws.merge_cells(f"A1:{last_col}1")
    c = ws.cell(row=1, column=1, value=DISCLAIMER)
    c.fill = F_DISC; c.font = DISC_FNT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN; ws.row_dimensions[1].height = 16

    # Row 2: title
    ws.merge_cells(f"A2:{last_col}2")
    t = ws.cell(row=2, column=1, value=title_text)
    t.fill = title_fill; t.font = _font(bold=True, color="FFFFFF", size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 26

    # Row 3: headers
    for col, h in enumerate(TAPE_HEADERS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill = title_fill; c.font = WHBOLD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "A4"
    set_col_widths(ws)

    # Data rows
    data_start = 4
    for i, row_dict in enumerate(rows):
        r = data_start + i
        alt = F_LTBLUE if r % 2 == 0 else None
        for field, col in CI.items():
            val = row_dict[field]
            c = ws.cell(row=r, column=col, value=val)
            if alt: c.fill = alt
            fmt = COL_FMTS[field]
            if fmt: c.number_format = fmt
            c.font = NORMAL
            c.alignment = Alignment(horizontal=COL_ALIGN[field], vertical="center")
            c.border = THIN

    # Totals row
    data_end = data_start + len(rows) - 1
    tr = data_end + 1
    for col in range(1, N_COLS + 1):
        c = ws.cell(row=tr, column=col)
        c.fill = F_TOTAL; c.border = THIN
    ws.cell(row=tr, column=1).value = "TOTALS / AVERAGES"
    ws.cell(row=tr, column=1).font  = BLKBOLD
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="center")
    for col, cl, fmt in [
        (6,"F",CURR0),(7,"G",CURR),(12,"L",CURR),(13,"M",CURR),(14,"N",CURR)
    ]:
        c = ws.cell(row=tr, column=col)
        c.value = f"=SUM({cl}{data_start}:{cl}{data_end})"
        c.number_format = fmt; c.font = BLKBOLD; c.fill = F_TOTAL
        c.border = THIN; c.alignment = Alignment(horizontal="right")
    for col, cl in [(8,"H"),(9,"I")]:
        c = ws.cell(row=tr, column=col)
        c.value = f"=AVERAGE({cl}{data_start}:{cl}{data_end})"
        c.number_format = PCT3; c.font = BLKBOLD; c.fill = F_TOTAL
        c.border = THIN; c.alignment = Alignment(horizontal="right")
    c = ws.cell(row=tr, column=10)
    c.value = f"=COUNT(G{data_start}:G{data_end})"
    c.number_format = NUM0; c.font = BLKBOLD; c.fill = F_TOTAL
    c.border = THIN; c.alignment = Alignment(horizontal="center")


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    clean_tape = os.path.join(OUT, "MSR_Sample_Tape_Dec2025_Jan2026.xlsx")
    out_path   = os.path.join(OUT, "MSR_Tape_Jan2026_SUBSERVICER.xlsx")

    print("Reading clean Jan 2026 tape...")
    rows = read_jan_tape(clean_tape)
    total_loans = len(rows)

    # ── Select error targets ─────────────────────────────────────────────────
    # Continuing loans: MSR1xxxxx  |  New adds: MSR2xxxxx
    continuing_idx = [i for i, r in enumerate(rows)
                      if r["loan_id"] and str(r["loan_id"]).startswith("MSR1")]
    current_idx    = [i for i in continuing_idx
                      if rows[i].get("status") in (None, "Current")]
    fnma_cont_idx  = [i for i in continuing_idx if rows[i]["investor"] == "FNMA"]
    gnma_cont_idx  = [i for i in continuing_idx if rows[i]["investor"] == "GNMA"]

    used = set()

    def pick(pool, n):
        available = [i for i in pool if i not in used]
        chosen = available[:n]
        used.update(chosen)
        return chosen

    # Hard stop targets (spread across the list)
    upb_x10_idx      = pick(continuing_idx[5:],   3)
    upb_zero_idx     = pick(continuing_idx[50:],  1)
    missing_idx      = pick(continuing_idx[100:], 3)
    rate_whole_idx   = pick(continuing_idx[200:], 2)
    upb_gt_orig_idx  = pick(continuing_idx[300:], 1)
    dup_target_idx   = pick(continuing_idx[150:], 1)

    # Yellow light targets
    fnma_nsf_idx      = pick(fnma_cont_idx, 2)
    gnma_nsf_idx      = pick(gnma_cont_idx, 2)
    status_skip_idx   = pick([i for i in current_idx if i not in used], 2)
    pi_inflate_idx    = pick([i for i in continuing_idx if i not in used], 2)
    ndd_past_idx      = pick([i for i in current_idx if i not in used], 2)
    rem_unch_idx      = pick([i for i in continuing_idx if i not in used], 1)

    # ── Apply errors ─────────────────────────────────────────────────────────
    dirty    = [deepcopy(r) for r in rows]
    missing_set = set(missing_idx)
    error_log   = []

    def log(etype, category, idx, field, orig, submitted, desc):
        error_log.append({
            "error_type": etype,
            "category":   category,
            "loan_id":    dirty[idx]["loan_id"],
            "field":      field,
            "original":   orig,
            "submitted":  submitted,
            "description":desc,
        })

    # Missing loans (logged; rows excluded from output)
    for i in missing_idx:
        lid = dirty[i]["loan_id"]
        error_log.append({
            "error_type":  "HARD STOP",
            "category":    "Missing Loan (no PIF reported)",
            "loan_id":     lid,
            "field":       "—",
            "original":    "Present in prior month tape",
            "submitted":   "Not in submission (no PIF reported either)",
            "description": f"{lid} is absent from the submission with no corresponding PIF entry.",
        })

    # UPB x 10
    for i in upb_x10_idx:
        orig = dirty[i]["upb"]
        dirty[i]["upb"] = round(orig * 10, 2)
        log("HARD STOP", "UPB Extra Zero (x10)", i,
            "Current UPB ($)", f"${orig:,.2f}", f"${dirty[i]['upb']:,.2f}",
            "UPB appears to have an extra zero; submitted value is 10x expected.")

    # UPB = 0
    for i in upb_zero_idx:
        orig = dirty[i]["upb"]
        dirty[i]["upb"] = 0.0
        log("HARD STOP", "UPB = Zero (active loan)", i,
            "Current UPB ($)", f"${orig:,.2f}", "$0.00",
            "Active loan submitted with UPB of zero but not marked Paid in Full.")

    # Rate as whole number (e.g. 0.0650 -> 6.50)
    for i in rate_whole_idx:
        orig = dirty[i]["rate"]
        dirty[i]["rate"] = round(orig * 100, 4)
        log("HARD STOP", "Rate as Whole Number", i,
            "Rate", f"{orig:.4%}", f"{dirty[i]['rate']:.4f}",
            "Rate submitted as whole number (e.g. 6.50) instead of decimal (0.0650).")

    # UPB > Orig Bal
    for i in upb_gt_orig_idx:
        orig_upb = dirty[i]["upb"]
        orig_bal = dirty[i]["orig_bal"]
        dirty[i]["upb"] = round(orig_bal * 1.12, 2)
        log("HARD STOP", "UPB Exceeds Original Balance", i,
            "Current UPB ($)", f"${orig_upb:,.2f}",
            f"${dirty[i]['upb']:,.2f} (Orig Bal: ${orig_bal:,.2f})",
            "Current UPB exceeds original balance — mathematically impossible for amortizing loan.")

    # Duplicate loan ID (appended as extra row at end)
    dup_row = None
    if dup_target_idx:
        i = dup_target_idx[0]
        dup_row = deepcopy(dirty[i])
        lid = dup_row["loan_id"]
        error_log.append({
            "error_type":  "HARD STOP",
            "category":    "Duplicate Loan ID",
            "loan_id":     lid,
            "field":       "Loan ID",
            "original":    "Appears once",
            "submitted":   "Appears twice in submission",
            "description": f"Loan ID {lid} appears twice; duplicate row appended at end of file.",
        })

    # NSF as percent (FNMA: 0.0025 -> 0.25)
    for i in fnma_nsf_idx:
        orig = dirty[i]["nsf"]
        dirty[i]["nsf"] = round(orig * 100, 4)  # 0.0025 -> 0.25
        log("YELLOW LIGHT", "NSF Expressed as Percent (FNMA)", i,
            "Net Serv Fee", f"{orig:.4%}",
            f"{dirty[i]['nsf']:.4f} (expected ~{orig:.4f})",
            "NSF looks like it was entered as a percentage (0.25) instead of decimal (0.0025).")

    # NSF as whole bps (GNMA: 0.0044 -> 44)
    for i in gnma_nsf_idx:
        orig = dirty[i]["nsf"]
        dirty[i]["nsf"] = round(orig * 10000)   # 0.0044 -> 44
        log("YELLOW LIGHT", "NSF Expressed as Whole Basis Points (GNMA)", i,
            "Net Serv Fee", f"{orig:.4%}",
            f"{dirty[i]['nsf']} (expected decimal ~{orig:.4f})",
            "NSF submitted as whole basis points (e.g. 44) instead of decimal (0.0044).")

    # Status skip (Current -> 90+ DPD)
    for i in status_skip_idx:
        dirty[i]["status"] = "90+ DPD"
        dirty[i]["ndd"]    = date(2025, 10, 1)
        log("YELLOW LIGHT", "Status Skip (Current -> 90+ DPD)", i,
            "Status", "Current", "90+ DPD",
            "Loan jumped from Current to 90+ DPD in one month, skipping 30 and 60 DPD buckets.")

    # P&I inflated ~20%
    for i in pi_inflate_idx:
        orig = dirty[i]["pi"]
        dirty[i]["pi"]        = round(orig * 1.20, 2)
        dirty[i]["total_pmt"] = round(dirty[i]["pi"] + (dirty[i]["escrow"] or 0), 2)
        log("YELLOW LIGHT", "P&I Payment Inflated ~20%", i,
            "P&I ($)", f"${orig:,.2f}", f"${dirty[i]['pi']:,.2f}",
            "P&I is ~20% above expected based on UPB, rate, and remaining term.")

    # NDD in past for current loan
    for i in ndd_past_idx:
        orig = dirty[i]["ndd"]
        dirty[i]["ndd"] = date(2025, 6, 1)
        log("YELLOW LIGHT", "Next Due Date in Past (Current Loan)", i,
            "Next Due Date", str(orig), "2025-06-01",
            "Current loan has a Next Due Date in the past, suggesting possible unreported delinquency.")

    # Remaining term unchanged
    for i in rem_unch_idx:
        orig = dirty[i]["rem_term"]
        dirty[i]["rem_term"] = (orig or 0) + 1  # should have decreased by 1
        log("YELLOW LIGHT", "Remaining Term Unchanged", i,
            "Rem Term", str(orig), str(dirty[i]["rem_term"]),
            "Remaining term did not decrease by 1 from prior month, indicating a system update error.")

    # ── Build final output row list ──────────────────────────────────────────
    output_rows = [row for i, row in enumerate(dirty) if i not in missing_set]
    if dup_row:
        output_rows.append(dup_row)   # duplicate appended at end

    # ── Write workbook ───────────────────────────────────────────────────────
    wb = Workbook()

    # Sheet 1: Subservicer submission (no error highlighting — realistic)
    ws_sub = wb.active
    ws_sub.title = "Jan 2026"
    title_txt = (
        f"MSR TAPE — January 2026 (Subservicer Submission)  |  "
        f"As of {REPORT_DATE.strftime('%m/%d/%Y')}  |  {len(output_rows):,} Loans"
    )
    write_tape_sheet(ws_sub, title_txt, F_BLUE, output_rows)

    # Sheet 2: Error Log (reference — not visible to validator)
    ws_log = wb.create_sheet("Error Log - Reference")
    for col, w in zip("ABCDEFG", [14,34,14,18,22,34,55]):
        ws_log.column_dimensions[col].width = w

    LOG_NCOLS = 7
    last_log  = get_column_letter(LOG_NCOLS)
    ws_log.merge_cells(f"A1:{last_log}1")
    t = ws_log.cell(row=1, column=1,
        value="ERROR LOG (Reference Only) — Injected errors for validation testing. Not visible to validator.")
    t.fill = F_LOG_HDR; t.font = _font(bold=True, color="FFFFFF", size=11)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_log.row_dimensions[1].height = 24

    log_hdrs = ["Error Type","Category","Loan ID","Field",
                "Original Value","Submitted Value","Description"]
    for col, h in enumerate(log_hdrs, 1):
        c = ws_log.cell(row=2, column=col, value=h)
        c.fill = F_LOG_HDR; c.font = WHBOLD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
    ws_log.row_dimensions[2].height = 28
    ws_log.freeze_panes = "A3"

    for r, err in enumerate(error_log, 3):
        fill = F_HARD if err["error_type"] == "HARD STOP" else F_YELLOW
        vals = [err["error_type"], err["category"], err["loan_id"], err["field"],
                err["original"], err["submitted"], err["description"]]
        for col, val in enumerate(vals, 1):
            c = ws_log.cell(row=r, column=col, value=val)
            c.fill = fill; c.font = NORMAL
            c.alignment = Alignment(
                horizontal="left" if col in (2, 7) else "center",
                vertical="center", wrap_text=True)
            c.border = THIN
            if col == 1:
                c.font = _font(bold=True, size=9,
                    color="C00000" if err["error_type"] == "HARD STOP" else "7B3F00")
        ws_log.row_dimensions[r].height = 32

    wb.save(out_path)

    hs = sum(1 for e in error_log if e["error_type"] == "HARD STOP")
    yl = sum(1 for e in error_log if e["error_type"] == "YELLOW LIGHT")

    print(f"\n[OK] Subservicer tape saved: {out_path}")
    print(f"     Loans in submission:  {len(output_rows):,}  (clean tape: {total_loans:,})")
    print(f"     Injected errors:      {len(error_log)}  ({hs} hard stops, {yl} yellow lights)")
    print(f"\n  Injected error detail:")
    for err in error_log:
        tag = "HARD" if err["error_type"] == "HARD STOP" else "YLLW"
        print(f"    [{tag}]  {err['loan_id']:12s}  {err['category']}")


if __name__ == "__main__":
    main()
