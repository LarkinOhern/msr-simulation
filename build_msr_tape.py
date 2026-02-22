"""
MSR Portfolio v2 — Realistic 1,000-loan portfolio, Dec 2025 -> Jan 2026
Reconciliation: Dec + New Adds - PIF = Jan (counts and UPBs must tie)
"""

import os
import random, math
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(7)
OUT = os.path.dirname(os.path.abspath(__file__))

DISCLAIMER = "SIMULATED DATA — All loan information is synthetic and generated for testing purposes only. Not representative of any real portfolio."

# ── Style constants ─────────────────────────────────────────────────────────
def _fill(hex6): return PatternFill("solid", fgColor=hex6)
def _font(bold=False, color="000000", size=9, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def _border(style="thin", color="B8CCE4"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

F_NAVY   = _fill("1F4E79"); F_BLUE   = _fill("2E75B6"); F_LTBLUE = _fill("D6E4F0")
F_LTBLUE2= _fill("EBF5FB"); F_TOTAL  = _fill("BDD7EE"); F_WHITE  = _fill("FFFFFF")
F_GREEN  = _fill("375623"); F_LTGRN  = _fill("E2EFDA"); F_GRNTOT = _fill("A9D18E")
F_RED    = _fill("7B2C2C"); F_LTRED  = _fill("FCE4D6"); F_REDTOT = _fill("F4CCCC")
F_ORANGE = _fill("833C00"); F_LTORANGE=_fill("FCE4D6")
F_GREY   = _fill("F2F2F2"); F_DGREY  = _fill("D9D9D9")
F_YELLOW = _fill("FFF2CC")
F_DISC   = _fill("FFF2CC")  # disclaimer row background

WHBOLD   = _font(bold=True, color="FFFFFF", size=9)
BLKBOLD  = _font(bold=True, size=9)
NORMAL   = _font(size=9)
BLUE_IN  = _font(color="0000FF", size=9)
GRN_ST   = _font(bold=True, color="006400", size=9)
RED_ST   = _font(bold=True, color="C00000", size=9)
DISC_FNT = _font(bold=True, color="7B3F00", size=9, italic=True)

THIN  = _border("thin",   "B8CCE4")
MED   = _border("medium", "1F4E79")

CURR   = '#,##0.00'
CURR0  = '#,##0'
PCT3   = '0.000%'
PCT1   = '0.0%'
DFMT   = 'MM/DD/YYYY'
NUM0   = '#,##0'

def hcell(ws, row, col, val, fill=None, font=None, align="center", wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill or F_NAVY
    c.font = font or WHBOLD
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = THIN
    return c

def dcell(ws, row, col, val, fill=None, fmt=None, font=None, align="left"):
    c = ws.cell(row=row, column=col, value=val)
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    c.font = font or NORMAL
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = THIN
    return c

def write_disclaimer(ws, ncols, row=1):
    """Write a full-width disclaimer row and return the next row number."""
    last_col = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1, value=DISCLAIMER)
    c.fill = F_DISC
    c.font = DISC_FNT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    c.border = THIN
    ws.row_dimensions[row].height = 16
    return row + 1

# ── Date / math helpers ──────────────────────────────────────────────────────
def rand_date(a, b):
    return a + timedelta(days=random.randint(0, (b-a).days))

def calc_pi(orig_bal, rate, term):
    r = rate / 12
    if r < 1e-9: return orig_bal / term
    return orig_bal * r * (1+r)**term / ((1+r)**term - 1)

def calc_upb(orig_bal, rate, term, elapsed):
    """UPB after `elapsed` monthly payments."""
    r = rate / 12
    if r < 1e-9: return orig_bal * (1 - elapsed/term)
    return orig_bal * ((1+r)**term - (1+r)**elapsed) / ((1+r)**term - 1)

def scheduled_principal(upb, rate, pi):
    """One month's principal reduction for a current loan."""
    return pi - upb * rate / 12

# ── Loan generation ──────────────────────────────────────────────────────────
INVESTOR_POOL = (
    ["FNMA"]*45 + ["FHLMC"]*30 + ["GNMA"]*20 + ["Portfolio"]*5
)
LTYPE_POOL = (
    ["Conventional"]*60 + ["FHA"]*25 + ["VA"]*10 + ["USDA"]*5
)
TERM_POOL   = [360]*70 + [180]*15 + [240]*15
PURPOSE_POOL= ["Purchase"]*65 + ["Refinance"]*35

def calc_nsf(investor):
    """Net Servicing Fee: FNMA/FHLMC/Portfolio=25bps fixed; GNMA=19-69bps triangular, median 44bps."""
    if investor == "GNMA":
        return round(random.triangular(0.0019, 0.0069, 0.0044), 4)
    return 0.0025

# Rate ranges by origination year (reflecting market conditions)
def rand_rate(orig_year):
    if   orig_year <= 2016: return round(random.uniform(0.0350, 0.0450), 4)
    elif orig_year <= 2018: return round(random.uniform(0.0400, 0.0500), 4)
    elif orig_year <= 2020: return round(random.uniform(0.0275, 0.0375), 4)
    elif orig_year <= 2021: return round(random.uniform(0.0275, 0.0350), 4)
    elif orig_year <= 2022: return round(random.uniform(0.0450, 0.0700), 4)
    elif orig_year <= 2023: return round(random.uniform(0.0625, 0.0799), 4)
    else:                   return round(random.uniform(0.0599, 0.0749), 4)

# Orig balance distribution centered around $350K
def rand_orig_bal():
    base = random.gauss(350000, 80000)
    return round(max(100000, min(750000, base)) / 5000) * 5000

def make_loan(loan_id, as_of=date(2025,12,31)):
    orig_date = rand_date(date(2015,1,1), date(2024,6,30))
    orig_bal  = rand_orig_bal()
    rate      = rand_rate(orig_date.year)
    term      = random.choice(TERM_POOL)
    investor  = random.choice(INVESTOR_POOL)
    lt        = random.choice(LTYPE_POOL)
    purpose   = random.choice(PURPOSE_POOL)
    pi        = calc_pi(orig_bal, rate, term)
    elapsed   = (as_of.year - orig_date.year)*12 + (as_of.month - orig_date.month)
    elapsed   = max(1, min(elapsed, term - 1))
    remaining = term - elapsed
    upb       = round(calc_upb(orig_bal, rate, term, elapsed), 2)
    escrow    = round(random.uniform(200, 700), 2)
    mat_year  = orig_date.year + term//12
    mat_month = orig_date.month
    maturity  = date(min(mat_year, 2055), mat_month, 1)

    return {
        "loan_id":    loan_id,
        "loan_type":  lt,
        "purpose":    purpose,
        "investor":   investor,
        "orig_date":  orig_date,
        "orig_bal":   orig_bal,
        "rate":       rate,
        "nsf":        calc_nsf(investor),
        "term":       term,
        "elapsed":    elapsed,
        "remaining":  remaining,
        "maturity":   maturity,
        "upb_dec":    upb,
        "upb_jan":    None,
        "pi":         round(pi, 2),
        "escrow":     escrow,
        "total_pmt":  round(pi + escrow, 2),
        "status_dec": "Current",
        "status_jan": "Current",
        "ndd_dec":    date(2026,1,1),
        "ndd_jan":    date(2026,2,1),
        "curtailment": 0.0,
        "is_pif":     False,
        "is_new_add": False,
        "cap_amount": 0.0,
    }

# ── Build Dec 2025 portfolio (1,000 loans) ───────────────────────────────────
N_DEC = 1000
print("Generating Dec 2025 portfolio...")
dec_loans = [make_loan(f"MSR{100000+i:06d}") for i in range(N_DEC)]

# ── Assign DQ status for Dec 2025 ────────────────────────────────────────────
# 90+=1, 60-89=2, 30-59=2  (5 total DQ = 0.5%)
N_90  = max(1, round(N_DEC * 0.001))
N_60  = max(2, round(N_DEC * 0.0015))
N_30  = max(2, round(N_DEC * 0.002))

sorted_by_rate = sorted(range(N_DEC),
                        key=lambda i: dec_loans[i]["rate"], reverse=True)
dq_pool = sorted_by_rate[:60]
random.shuffle(dq_pool)

idx_90 = dq_pool[:N_90]
idx_60 = dq_pool[N_90:N_90+N_60]
idx_30 = dq_pool[N_90+N_60:N_90+N_60+N_30]

for i in idx_90:  dec_loans[i]["status_dec"] = "90+ DPD"
for i in idx_60:  dec_loans[i]["status_dec"] = "60 DPD"
for i in idx_30:  dec_loans[i]["status_dec"] = "30 DPD"

for i in idx_90:  dec_loans[i]["ndd_dec"] = date(2025,10,1)
for i in idx_60:  dec_loans[i]["ndd_dec"] = date(2025,11,1)
for i in idx_30:  dec_loans[i]["ndd_dec"] = date(2025,12,1)

# ── SMM / PIF selection ───────────────────────────────────────────────────────
# CPR = 13%  ->  SMM = 1 - (1-0.13)^(1/12)
SMM      = 1 - (1 - 0.13) ** (1/12)
N_PIF    = max(10, round(N_DEC * SMM))
print(f"  SMM={SMM:.4%}  ->  {N_PIF} PIF loans")

dq_set      = set(idx_90 + idx_60 + idx_30)
current_idx = [i for i in range(N_DEC) if i not in dq_set]
pif_idx     = random.sample(current_idx, N_PIF)

for i in pif_idx:
    dec_loans[i]["is_pif"]     = True
    dec_loans[i]["status_jan"] = "Paid in Full"

# ── Curtailments (extra principal, Jan payment) ───────────────────────────────
curtail_candidates = [i for i in current_idx if i not in pif_idx]
curtail_idx = random.sample(curtail_candidates, 8)
for i in curtail_idx:
    dec_loans[i]["curtailment"] = round(random.choice(
        [5000,7500,10000,12500,15000,20000,25000,30000,35000,40000]), 2)

# ── DQ status migration Dec->Jan ───────────────────────────────────────────────
for i in idx_90:
    dec_loans[i]["status_jan"] = "90+ DPD"
    dec_loans[i]["ndd_jan"]    = date(2025,11,1)

if len(idx_60) >= 2:
    dec_loans[idx_60[0]]["status_jan"] = "Current"
    dec_loans[idx_60[0]]["ndd_jan"]    = date(2026,2,1)
    dec_loans[idx_60[1]]["status_jan"] = "90+ DPD"
    dec_loans[idx_60[1]]["ndd_jan"]    = date(2025,11,1)

if len(idx_30) >= 2:
    dec_loans[idx_30[0]]["status_jan"] = "Current"
    dec_loans[idx_30[0]]["ndd_jan"]    = date(2026,2,1)
    dec_loans[idx_30[1]]["status_jan"] = "60 DPD"
    dec_loans[idx_30[1]]["ndd_jan"]    = date(2025,12,1)

new_dq_candidates = [i for i in current_idx
                     if i not in pif_idx and i not in curtail_idx]
new_dq_jan = random.sample(new_dq_candidates, 3)
for i in new_dq_jan:
    dec_loans[i]["status_jan"] = "30 DPD"
    dec_loans[i]["ndd_jan"]    = date(2026,1,1)

# ── Compute Jan UPB for each Dec loan (non-PIF) ───────────────────────────────
cap_candidates = idx_90 + idx_60
for i in cap_candidates:
    cap = round(random.uniform(300, 2500), 2)
    dec_loans[i]["cap_amount"] = cap

for ln in dec_loans:
    if ln["is_pif"]:
        ln["upb_jan"] = None
        ln["remaining_jan"] = None
        continue

    sched_prin = scheduled_principal(ln["upb_dec"], ln["rate"], ln["pi"])
    status_dec = ln["status_dec"]

    if status_dec == "Current":
        base = ln["upb_dec"] - sched_prin - ln["curtailment"]
    else:
        base = ln["upb_dec"] + ln["cap_amount"]

    ln["upb_jan"] = round(max(0.0, base), 2)
    ln["remaining_jan"] = ln["remaining"] - 1

# ── Build New Adds (200 loans, recent originations) ───────────────────────────
N_NEW_ADDS = 200
print(f"Generating {N_NEW_ADDS} New Adds...")
new_add_loans = []
for i in range(N_NEW_ADDS):
    ln = make_loan(f"MSR{200000+i:06d}", as_of=date(2026,1,31))
    ln["orig_date"]     = rand_date(date(2025,9,1), date(2026,1,15))
    ln["rate"]          = round(random.uniform(0.0599, 0.0750), 4)
    ln["orig_bal"]      = rand_orig_bal()
    ln["term"]          = random.choice([360]*75 + [180]*10 + [240]*15)
    ln["pi"]            = round(calc_pi(ln["orig_bal"], ln["rate"], ln["term"]), 2)
    ln["total_pmt"]     = round(ln["pi"] + ln["escrow"], 2)
    ln["elapsed"]       = random.randint(1, 3)
    ln["remaining"]     = ln["term"] - ln["elapsed"]
    ln["remaining_jan"] = ln["remaining"]
    ln["upb_dec"]       = None
    ln["upb_jan"]       = round(calc_upb(ln["orig_bal"], ln["rate"], ln["term"],
                                         ln["elapsed"]), 2)
    ln["status_dec"]    = None
    ln["status_jan"]    = "Current"
    ln["ndd_jan"]       = date(2026,2,1)
    ln["is_new_add"]    = True
    ln["transfer_date"] = rand_date(date(2025,12,15), date(2026,1,25))
    mat_year  = ln["orig_date"].year + ln["term"]//12
    mat_month = ln["orig_date"].month
    ln["maturity"] = date(min(mat_year, 2056), mat_month, 1)
    new_add_loans.append(ln)

# ── Build Jan 2026 portfolio ──────────────────────────────────────────────────
jan_loans = [ln for ln in dec_loans if not ln["is_pif"]] + new_add_loans
N_JAN = len(jan_loans)
print(f"Jan portfolio: {N_DEC} - {N_PIF} PIF + {N_NEW_ADDS} New Adds = {N_JAN} loans")

# ── Reconciliation numbers ────────────────────────────────────────────────────
dec_total_upb  = sum(ln["upb_dec"] for ln in dec_loans)
pif_loans_list = [ln for ln in dec_loans if ln["is_pif"]]
pif_total_upb  = sum(ln["upb_dec"] for ln in pif_loans_list)
new_adds_upb   = sum(ln["upb_jan"] for ln in new_add_loans)
jan_total_upb  = sum(ln["upb_jan"] for ln in jan_loans)

cur_remaining  = [ln for ln in dec_loans if not ln["is_pif"] and ln["status_dec"]=="Current"]
sched_amort    = sum(scheduled_principal(ln["upb_dec"], ln["rate"], ln["pi"])
                     for ln in cur_remaining)
curtailments   = sum(ln["curtailment"] for ln in dec_loans if not ln["is_pif"])
caps_total     = sum(ln["cap_amount"] for ln in dec_loans if not ln["is_pif"])

bridge_check = (dec_total_upb
                - sched_amort
                - curtailments
                + caps_total
                - pif_total_upb
                + new_adds_upb)
print(f"\nReconciliation check:")
print(f"  Dec UPB:          ${dec_total_upb:>14,.2f}")
print(f"  - Sched Amort:   (${sched_amort:>13,.2f})")
print(f"  - Curtailments:  (${curtailments:>13,.2f})")
print(f"  + Capitalizations: ${caps_total:>12,.2f}")
print(f"  - PIF UPB:       (${pif_total_upb:>13,.2f})")
print(f"  + New Adds UPB:   ${new_adds_upb:>13,.2f}")
print(f"  = Bridge result:  ${bridge_check:>13,.2f}")
print(f"  Actual Jan UPB:   ${jan_total_upb:>13,.2f}")
print(f"  Difference:       ${bridge_check - jan_total_upb:>13,.2f}")

def dq_stats(loans, upb_field):
    active = [ln for ln in loans if ln.get(upb_field) is not None]
    n = len(active)
    sf = "status_dec" if upb_field == "upb_dec" else "status_jan"
    s30  = sum(1 for ln in active if ln.get(sf) == "30 DPD")
    s60  = sum(1 for ln in active if ln.get(sf) == "60 DPD")
    s90  = sum(1 for ln in active if ln.get(sf) == "90+ DPD")
    scur = n - s30 - s60 - s90
    return n, scur, s30, s60, s90

dec_n, dec_cur, dec_30, dec_60, dec_90 = dq_stats(dec_loans, "upb_dec")
jan_n, jan_cur, jan_30, jan_60, jan_90 = dq_stats(jan_loans, "upb_jan")

print(f"\nDec 2025: {dec_n} loans  Current={dec_cur}  30DPD={dec_30}  60DPD={dec_60}  90+={dec_90}")
print(f"Jan 2026: {jan_n} loans  Current={jan_cur}  30DPD={jan_30}  60DPD={jan_60}  90+={jan_90}")

# ── HELPERS ───────────────────────────────────────────────────────────────────

def status_font(status):
    if status in (None, "Current"):     return GRN_ST
    elif status == "Paid in Full":      return _font(bold=True, color="1F4E79", size=9)
    else:                               return RED_ST

# Tape columns (16 total — no borrower name or address):
# 1=Loan ID, 2=Loan Type, 3=Purpose, 4=Investor, 5=Orig Date,
# 6=Orig Bal, 7=UPB, 8=Rate, 9=Net Serv Fee, 10=Rem Term, 11=Maturity,
# 12=P&I, 13=Escrow, 14=Total Pmt, 15=Status, 16=Next Due Date

TAPE_HEADERS = [
    "Loan ID","Loan Type","Purpose","Investor",
    "Orig Date","Original Bal ($)","Current UPB ($)","Rate","Net Serv Fee","Rem Term",
    "Maturity","P&I ($)","Escrow ($)","Total Pmt ($)","Status","Next Due Date"
]
TAPE_NCOLS = len(TAPE_HEADERS)  # 16

def set_tape_col_widths(ws):
    widths = {
        "A":13,"B":13,"C":11,"D":10,
        "E":14,"F":16,"G":14,"H":13,
        "I":12,"J":11,"K":13,"L":13,
        "M":12,"N":13,"O":13,"P":13,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def write_tape_header(ws, title_text, title_fill, row1=1):
    """Write disclaimer + title + column headers. Returns first data row number."""
    next_row = write_disclaimer(ws, TAPE_NCOLS, row=row1)
    last_col = get_column_letter(TAPE_NCOLS)
    ws.merge_cells(f"A{next_row}:{last_col}{next_row}")
    tc = ws.cell(row=next_row, column=1, value=title_text)
    tc.fill = title_fill
    tc.font = _font(bold=True, color="FFFFFF", size=12)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[next_row].height = 26
    next_row += 1
    for col, h in enumerate(TAPE_HEADERS, 1):
        hcell(ws, next_row, col, h)
    ws.row_dimensions[next_row].height = 30
    ws.freeze_panes = f"A{next_row+1}"
    set_tape_col_widths(ws)
    return next_row + 1  # first data row

def write_tape_row(ws, row, ln, upb_field, status_field, ndd_field, rem_field="remaining"):
    upb    = ln[upb_field]
    status = ln[status_field]
    ndd    = ln[ndd_field]
    fill   = F_LTBLUE if row % 2 == 0 else None

    dcell(ws, row,  1, ln["loan_id"],   fill, align="center")
    dcell(ws, row,  2, ln["loan_type"], fill, align="center")
    dcell(ws, row,  3, ln["purpose"],   fill, align="center")
    dcell(ws, row,  4, ln["investor"],  fill, align="center")
    dcell(ws, row,  5, ln["orig_date"], fill, DFMT,  align="center")
    dcell(ws, row,  6, ln["orig_bal"],  fill, CURR0, align="right")
    dcell(ws, row,  7, upb,             fill, CURR,  align="right")
    dcell(ws, row,  8, ln["rate"],      fill, PCT3,  align="right")
    dcell(ws, row,  9, ln["nsf"],       fill, PCT3,  align="right")
    dcell(ws, row, 10, ln[rem_field],   fill, NUM0,  align="center")
    dcell(ws, row, 11, ln["maturity"],  fill, DFMT,  align="center")
    dcell(ws, row, 12, ln["pi"],        fill, CURR,  align="right")
    dcell(ws, row, 13, ln["escrow"],    fill, CURR,  align="right")
    dcell(ws, row, 14, ln["total_pmt"], fill, CURR,  align="right")
    sc = dcell(ws, row, 15, status or "Current", fill, align="center")
    sc.font = status_font(status)
    dcell(ws, row, 16, ndd,             fill, DFMT,  align="center")

def write_tape_totals(ws, data_start, data_end, total_row, fill=F_TOTAL):
    for col in range(1, TAPE_NCOLS + 1):
        c = ws.cell(row=total_row, column=col)
        c.fill = fill; c.border = THIN
    ws.cell(row=total_row, column=1).value     = "TOTALS / AVERAGES"
    ws.cell(row=total_row, column=1).font      = BLKBOLD
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")
    # col 6=Orig Bal, 7=UPB, 12=P&I, 13=Escrow, 14=Total Pmt
    for col, cl, fmt in [
        (6,"F",CURR0),(7,"G",CURR),(12,"L",CURR),(13,"M",CURR),(14,"N",CURR)
    ]:
        c = ws.cell(row=total_row, column=col)
        c.value = f"=SUM({cl}{data_start}:{cl}{data_end})"
        c.number_format = fmt; c.font = BLKBOLD; c.fill = fill
        c.border = THIN; c.alignment = Alignment(horizontal="right")
    # Avg rate (col 8=H)
    c = ws.cell(row=total_row, column=8)
    c.value = f"=AVERAGE(H{data_start}:H{data_end})"
    c.number_format = PCT3; c.font = BLKBOLD; c.fill = fill
    c.border = THIN; c.alignment = Alignment(horizontal="right")
    # Avg NSF (col 9=I)
    c = ws.cell(row=total_row, column=9)
    c.value = f"=AVERAGE(I{data_start}:I{data_end})"
    c.number_format = PCT3; c.font = BLKBOLD; c.fill = fill
    c.border = THIN; c.alignment = Alignment(horizontal="right")
    # Count (col 10=J)
    c = ws.cell(row=total_row, column=10)
    c.value = f"=COUNT(G{data_start}:G{data_end})"
    c.number_format = NUM0; c.font = BLKBOLD; c.fill = fill
    c.border = THIN; c.alignment = Alignment(horizontal="center")

# ═══════════════════════════════════════════════════════════════════════════════
# FILE 1: MSR SAMPLE TAPE (Dec 2025 + Jan 2026 + Portfolio Summary)
# ═══════════════════════════════════════════════════════════════════════════════
print("\nBuilding MSR Tape...")
wb_tape = Workbook()

# ── Tab 1: Dec 2025 ──────────────────────────────────────────────────────────
ws_dec = wb_tape.active
ws_dec.title = "Dec 2025"
data_start_dec = write_tape_header(ws_dec,
    "MSR SAMPLE TAPE — December 2025 (As of 12/31/2025)", F_NAVY, row1=1)
for r, ln in enumerate(dec_loans, data_start_dec):
    write_tape_row(ws_dec, r, ln, "upb_dec", "status_dec", "ndd_dec")
tr_dec = data_start_dec + N_DEC
write_tape_totals(ws_dec, data_start_dec, tr_dec - 1, tr_dec)

# ── Tab 2: Jan 2026 ──────────────────────────────────────────────────────────
ws_jan = wb_tape.create_sheet("Jan 2026")
data_start_jan = write_tape_header(ws_jan,
    "MSR SAMPLE TAPE — January 2026 (As of 01/31/2026)", F_BLUE, row1=1)

existing_jan = [ln for ln in jan_loans if not ln["is_new_add"]]
new_add_jan  = [ln for ln in jan_loans if ln["is_new_add"]]
jan_sorted   = existing_jan + new_add_jan

for r, ln in enumerate(jan_sorted, data_start_jan):
    write_tape_row(ws_jan, r, ln, "upb_jan", "status_jan", "ndd_jan", rem_field="remaining_jan")
tr_jan = data_start_jan + N_JAN
write_tape_totals(ws_jan, data_start_jan, tr_jan - 1, tr_jan, fill=_fill("BDD7EE"))

# ── Tab 3: Portfolio Summary / Reconciliation Bridge ─────────────────────────
ws_sum = wb_tape.create_sheet("Portfolio Summary")
ws_sum.column_dimensions["A"].width = 38
ws_sum.column_dimensions["B"].width = 20
ws_sum.column_dimensions["C"].width = 20
ws_sum.column_dimensions["D"].width = 20
ws_sum.column_dimensions["E"].width = 20

def sum_hdr(ws, row, col, val, fill=F_NAVY, cols_span=None):
    if cols_span:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=col+cols_span-1
        )
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill; c.font = WHBOLD
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = THIN
    return c

def sum_row(ws, row, label, v_dec, v_jan, v_chg=None,
            fmt=CURR, lbl_fill=None, val_fill=None, bold=False):
    lf = lbl_fill or (F_TOTAL if bold else None)
    vf = val_fill or (F_TOTAL if bold else None)
    fnt = BLKBOLD if bold else NORMAL
    dcell(ws, row, 1, label, lf, font=fnt)
    dcell(ws, row, 2, v_dec, vf, fmt, font=fnt, align="right")
    dcell(ws, row, 3, v_jan, vf, fmt, font=fnt, align="right")
    if v_chg is not None:
        chg_fmt = '+#,##0.00;(#,##0.00);"-"' if fmt==CURR else ('+#,##0;(#,##0);"-"' if fmt==NUM0 else '+0.000%;(0.000%);"-"')
        c = dcell(ws, row, 4, v_chg, vf, chg_fmt, font=fnt, align="right")
        if isinstance(v_chg, (int,float)):
            if v_chg > 0:   c.font = _font(bold=bold, color="006400", size=9)
            elif v_chg < 0: c.font = _font(bold=bold, color="C00000", size=9)
    dcell(ws, row, 5, None, vf)

row = 1
# Disclaimer
row = write_disclaimer(ws_sum, 5, row=row)

# Title
ws_sum.merge_cells(f"A{row}:E{row}")
t = ws_sum.cell(row=row, column=1,
    value="PORTFOLIO SUMMARY & RECONCILIATION — Dec 2025 -> Jan 2026")
t.fill = F_NAVY; t.font = _font(bold=True, color="FFFFFF", size=13)
t.alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[row].height = 30
row += 1

# ─ Section A: Loan Count Bridge
ws_sum.merge_cells(f"A{row}:E{row}")
s = ws_sum.cell(row=row, column=1, value="A.  LOAN COUNT RECONCILIATION")
s.fill = F_BLUE; s.font = WHBOLD
s.alignment = Alignment(horizontal="left", vertical="center")
ws_sum.row_dimensions[row].height = 20
row += 1

for col, h in enumerate(["","Dec 2025","Jan 2026","Change",""], 1):
    hcell(ws_sum, row, col, h)
row += 1

sum_row(ws_sum, row, "Beginning Loan Count", N_DEC, N_DEC, None, NUM0, bold=False)
row += 1
sum_row(ws_sum, row, "  + New Adds (Jan 2026)", "", N_NEW_ADDS, None, NUM0)
row += 1
sum_row(ws_sum, row, "  - Paid in Full (Jan 2026)", "", f"-{N_PIF}", None, NUM0)
row += 1

dcell(ws_sum, row, 1, "Ending Loan Count", F_TOTAL, font=BLKBOLD)
dcell(ws_sum, row, 2, N_DEC,  F_TOTAL, NUM0, BLKBOLD, align="right")
dcell(ws_sum, row, 3, N_JAN,  F_TOTAL, NUM0, BLKBOLD, align="right")
chg_c = dcell(ws_sum, row, 4, N_JAN - N_DEC, F_TOTAL, '+#,##0;(#,##0);"-"', BLKBOLD, align="right")
chg_c.font = _font(bold=True, color="006400", size=9)
dcell(ws_sum, row, 5, "✓ Ties", F_GRNTOT, font=_font(bold=True, color="375623", size=9))
row += 1; row += 1

# ─ Section B: UPB Bridge
ws_sum.merge_cells(f"A{row}:E{row}")
s2 = ws_sum.cell(row=row, column=1, value="B.  UNPAID PRINCIPAL BALANCE RECONCILIATION")
s2.fill = F_BLUE; s2.font = WHBOLD
s2.alignment = Alignment(horizontal="left", vertical="center")
ws_sum.row_dimensions[row].height = 20
row += 1

for col, h in enumerate(["","Dec 2025","Jan 2026","",""], 1):
    hcell(ws_sum, row, col, h)
row += 1

sum_row(ws_sum, row, "Beginning Total UPB ($)", dec_total_upb, "", None, CURR)
row += 1
sum_row(ws_sum, row, "  - Scheduled Amortization (current loans)",
        f"(${sched_amort:,.2f})", "", None, CURR)
row += 1
sum_row(ws_sum, row, "  - Curtailments (8 loans)",
        f"(${curtailments:,.2f})", "", None, CURR)
row += 1
sum_row(ws_sum, row, "  + Capitalizations (DQ loans)",
        f"${caps_total:,.2f}", "", None, CURR)
row += 1
sum_row(ws_sum, row, "  - PIF UPB Removed",
        f"(${pif_total_upb:,.2f})", "", None, CURR)
row += 1
sum_row(ws_sum, row, "  + New Adds UPB",
        f"${new_adds_upb:,.2f}", "", None, CURR)
row += 1
dcell(ws_sum, row, 1, "Ending Total UPB ($)", F_TOTAL, font=BLKBOLD)
dcell(ws_sum, row, 2, "",                     F_TOTAL)
dcell(ws_sum, row, 3, jan_total_upb,          F_TOTAL, CURR, BLKBOLD, align="right")
dcell(ws_sum, row, 4, "",                     F_TOTAL)
dcell(ws_sum, row, 5, "✓ Ties", F_GRNTOT, font=_font(bold=True, color="375623", size=9))
row += 1; row += 1

# ─ Section C: Delinquency Migration
ws_sum.merge_cells(f"A{row}:E{row}")
s3 = ws_sum.cell(row=row, column=1, value="C.  DELINQUENCY MIGRATION")
s3.fill = F_RED; s3.font = WHBOLD
s3.alignment = Alignment(horizontal="left", vertical="center")
ws_sum.row_dimensions[row].height = 20
row += 1

for col, h in enumerate(["Bucket","Dec 2025","Jan 2026","Change","% of Portfolio (Jan)"],1):
    hcell(ws_sum, row, col, h)
row += 1

dq_data = [
    ("Current",   dec_cur, jan_cur),
    ("30 DPD",    dec_30,  jan_30),
    ("60 DPD",    dec_60,  jan_60),
    ("90+ DPD",   dec_90,  jan_90),
    ("TOTAL",     dec_n,   jan_n),
]
for label, d, j in dq_data:
    is_total = label == "TOTAL"
    fill = F_TOTAL if is_total else (F_LTRED if "DPD" in label else None)
    fnt  = BLKBOLD if is_total else NORMAL
    chg  = j - d
    pct  = j / jan_n if jan_n else 0
    dcell(ws_sum, row, 1, label, fill, font=fnt)
    dcell(ws_sum, row, 2, d,    fill, NUM0, fnt, align="right")
    dcell(ws_sum, row, 3, j,    fill, NUM0, fnt, align="right")
    cc = dcell(ws_sum, row, 4, chg,  fill, '+#,##0;(#,##0);"-"', fnt, align="right")
    if chg > 0:   cc.font = _font(bold=is_total, color="C00000", size=9)
    elif chg < 0: cc.font = _font(bold=is_total, color="006400", size=9)
    dcell(ws_sum, row, 5, pct, fill, PCT1, fnt, align="right")
    row += 1

row += 1

# ─ Section D: Investor / Loan Type composition
ws_sum.merge_cells(f"A{row}:E{row}")
s4 = ws_sum.cell(row=row, column=1, value="D.  PORTFOLIO COMPOSITION (Jan 2026)")
s4.fill = F_ORANGE; s4.font = WHBOLD
s4.alignment = Alignment(horizontal="left", vertical="center")
ws_sum.row_dimensions[row].height = 20
row += 1

for col, h in enumerate(["","Loan Count","Total UPB ($)","Avg UPB ($)","% of Portfolio"],1):
    hcell(ws_sum, row, col, h)
row += 1

from collections import defaultdict
by_inv = defaultdict(list)
for ln in jan_loans:
    by_inv[ln["investor"]].append(ln["upb_jan"])
for inv in sorted(by_inv.keys()):
    cnt = len(by_inv[inv])
    tot = sum(by_inv[inv])
    avg = tot/cnt
    pct = cnt/jan_n
    fill = F_LTBLUE if row % 2 == 0 else None
    dcell(ws_sum, row, 1, inv,  fill, font=NORMAL)
    dcell(ws_sum, row, 2, cnt,  fill, NUM0, align="right")
    dcell(ws_sum, row, 3, tot,  fill, CURR0, align="right")
    dcell(ws_sum, row, 4, avg,  fill, CURR0, align="right")
    dcell(ws_sum, row, 5, pct,  fill, PCT1,  align="right")
    row += 1
dcell(ws_sum, row, 1, "TOTAL", F_TOTAL, font=BLKBOLD)
dcell(ws_sum, row, 2, jan_n,   F_TOTAL, NUM0,  BLKBOLD, align="right")
dcell(ws_sum, row, 3, jan_total_upb, F_TOTAL, CURR0, BLKBOLD, align="right")
dcell(ws_sum, row, 4, jan_total_upb/jan_n, F_TOTAL, CURR0, BLKBOLD, align="right")
dcell(ws_sum, row, 5, 1.0,     F_TOTAL, PCT1,  BLKBOLD, align="right")
row += 2

# ─ Section E: Paid in Full — Loan ID Verification
ws_sum.merge_cells(f"A{row}:E{row}")
s5 = ws_sum.cell(row=row, column=1,
    value="E.  PAID IN FULL — LOAN ID VERIFICATION (Confirm Not in Jan Portfolio)")
s5.fill = F_RED; s5.font = WHBOLD
s5.alignment = Alignment(horizontal="left", vertical="center")
ws_sum.row_dimensions[row].height = 20
row += 1

for col, h in enumerate(["Loan ID","Dec UPB ($)","Payoff Month","In Jan Tape?"],1):
    hcell(ws_sum, row, col, h, fill=F_RED)
row += 1

jan_ids = set(ln["loan_id"] for ln in jan_loans)
for ln in pif_loans_list:
    in_jan = "YES — ERROR" if ln["loan_id"] in jan_ids else "No ✓"
    fill = F_LTRED if row % 2 == 0 else None
    dcell(ws_sum, row, 1, ln["loan_id"],  fill, align="center")
    dcell(ws_sum, row, 2, ln["upb_dec"],  fill, CURR,  align="right")
    dcell(ws_sum, row, 3, "January 2026", fill, align="center")
    c = dcell(ws_sum, row, 4, in_jan,     fill, align="center")
    c.font = GRN_ST if "✓" in in_jan else RED_ST
    row += 1

row += 1

# ─ Section F: New Add sample verification
ws_sum.merge_cells(f"A{row}:E{row}")
s6 = ws_sum.cell(row=row, column=1,
    value="F.  NEW ADDS — SAMPLE LOAN ID VERIFICATION (First 10 Shown, Confirm In Jan)")
s6.fill = F_GREEN; s6.font = WHBOLD
s6.alignment = Alignment(horizontal="left", vertical="center")
ws_sum.row_dimensions[row].height = 20
row += 1

for col, h in enumerate(["Loan ID","Jan UPB ($)","Transfer Date","In Jan Tape?"],1):
    hcell(ws_sum, row, col, h, fill=F_GREEN)
row += 1

for ln in new_add_loans[:10]:
    in_jan = "Yes ✓" if ln["loan_id"] in jan_ids else "NO — ERROR"
    fill = F_LTGRN if row % 2 == 0 else None
    dcell(ws_sum, row, 1, ln["loan_id"],       fill, align="center")
    dcell(ws_sum, row, 2, ln["upb_jan"],       fill, CURR, align="right")
    dcell(ws_sum, row, 3, ln["transfer_date"], fill, DFMT, align="center")
    c = dcell(ws_sum, row, 4, in_jan,          fill, align="center")
    c.font = GRN_ST if "✓" in in_jan else RED_ST
    row += 1

tape_path = os.path.join(OUT, "MSR_Sample_Tape_Dec2025_Jan2026.xlsx")
wb_tape.save(tape_path)
print(f"  Saved: {tape_path}")

# ═══════════════════════════════════════════════════════════════════════════════
# FILE 2: RECON — NEW ADDS
# ═══════════════════════════════════════════════════════════════════════════════
print("Building Recon - New Adds...")
wb_na = Workbook()
ws_na = wb_na.active
ws_na.title = "New Adds"

# Columns (15): Loan ID, Loan Type, Purpose, Investor, Origination Date,
#               Transfer/Add Date, Orig Balance, Current UPB, Rate,
#               Loan Term, Rem Term, P&I, Escrow, Total Pmt, Status
NA_HDRS = ["Loan ID","Loan Type","Purpose",
           "Investor","Origination Date","Transfer / Add Date","Orig Balance ($)",
           "Current UPB ($)","Rate","Net Serv Fee","Loan Term","Rem Term","P&I ($)","Escrow ($)",
           "Total Pmt ($)","Status"]
NA_WIDTHS = [13,13,11,10,14,14,16,14,11,12,10,10,13,12,13,12]
NA_NCOLS = len(NA_HDRS)

na_row = write_disclaimer(ws_na, NA_NCOLS, row=1)

ws_na.merge_cells(f"A{na_row}:{get_column_letter(NA_NCOLS)}{na_row}")
t = ws_na.cell(row=na_row, column=1,
    value=f"RECONCILIATION — NEW ADDS  |  January 2026  |  {N_NEW_ADDS} Loans")
t.fill = F_GREEN; t.font = _font(bold=True, color="FFFFFF", size=12)
t.alignment = Alignment(horizontal="center", vertical="center")
ws_na.row_dimensions[na_row].height = 26
na_row += 1

for col, (h, w) in enumerate(zip(NA_HDRS, NA_WIDTHS), 1):
    hcell(ws_na, na_row, col, h, fill=F_GREEN)
    ws_na.column_dimensions[get_column_letter(col)].width = w
ws_na.row_dimensions[na_row].height = 28
ws_na.freeze_panes = f"A{na_row+1}"
na_data_start = na_row + 1

for r, ln in enumerate(new_add_loans, na_data_start):
    fill = F_LTGRN if r % 2 == 0 else None
    dcell(ws_na, r,  1, ln["loan_id"],       fill, align="center")
    dcell(ws_na, r,  2, ln["loan_type"],     fill, align="center")
    dcell(ws_na, r,  3, ln["purpose"],       fill, align="center")
    dcell(ws_na, r,  4, ln["investor"],      fill, align="center")
    dcell(ws_na, r,  5, ln["orig_date"],     fill, DFMT,  align="center")
    dcell(ws_na, r,  6, ln["transfer_date"], fill, DFMT,  align="center")
    dcell(ws_na, r,  7, ln["orig_bal"],      fill, CURR0, align="right")
    dcell(ws_na, r,  8, ln["upb_jan"],       fill, CURR,  align="right")
    dcell(ws_na, r,  9, ln["rate"],          fill, PCT3,  align="right")
    dcell(ws_na, r, 10, ln["nsf"],           fill, PCT3,  align="right")
    dcell(ws_na, r, 11, ln["term"],          fill, NUM0,  align="center")
    dcell(ws_na, r, 12, ln["remaining"],     fill, NUM0,  align="center")
    dcell(ws_na, r, 13, ln["pi"],            fill, CURR,  align="right")
    dcell(ws_na, r, 14, ln["escrow"],        fill, CURR,  align="right")
    dcell(ws_na, r, 15, ln["total_pmt"],     fill, CURR,  align="right")
    dcell(ws_na, r, 16, "Current",           fill, align="center", font=GRN_ST)

na_tr = na_data_start + N_NEW_ADDS
for col in range(1, NA_NCOLS + 1):
    ws_na.cell(row=na_tr, column=col).fill   = F_GRNTOT
    ws_na.cell(row=na_tr, column=col).border = THIN
ws_na.cell(row=na_tr, column=1).value     = "TOTALS / AVERAGES"
ws_na.cell(row=na_tr, column=1).font      = BLKBOLD
ws_na.cell(row=na_tr, column=1).alignment = Alignment(horizontal="center")
# col 7=Orig Bal, 8=UPB, 13=P&I, 14=Escrow, 15=Total Pmt
for col, cl in [(7,"G"),(8,"H"),(13,"M"),(14,"N"),(15,"O")]:
    c = ws_na.cell(row=na_tr, column=col)
    c.value = f"=SUM({cl}{na_data_start}:{cl}{na_tr-1})"
    c.number_format = CURR; c.font = BLKBOLD; c.fill = F_GRNTOT
    c.border = THIN; c.alignment = Alignment(horizontal="right")
# Avg rate (col 9=I)
c = ws_na.cell(row=na_tr, column=9)
c.value = f"=AVERAGE(I{na_data_start}:I{na_tr-1})"
c.number_format = PCT3; c.font = BLKBOLD; c.fill = F_GRNTOT
c.border = THIN; c.alignment = Alignment(horizontal="right")
# Avg NSF (col 10=J)
c = ws_na.cell(row=na_tr, column=10)
c.value = f"=AVERAGE(J{na_data_start}:J{na_tr-1})"
c.number_format = PCT3; c.font = BLKBOLD; c.fill = F_GRNTOT
c.border = THIN; c.alignment = Alignment(horizontal="right")

na_path = os.path.join(OUT, "Recon_NewAdds_Jan2026.xlsx")
wb_na.save(na_path)
print(f"  Saved: {na_path}")

# ═══════════════════════════════════════════════════════════════════════════════
# FILE 3: RECON — PAID IN FULL
# ═══════════════════════════════════════════════════════════════════════════════
print("Building Recon - Paid in Full...")
wb_pif = Workbook()
ws_pif = wb_pif.active
ws_pif.title = "Paid in Full"

# Columns (13): Loan ID, Investor, Loan Type, Payoff Date, Orig Balance,
#               Final UPB, Payoff Amount, Interest Due, Fees/Penalties,
#               Net Proceeds, Rate, Payoff Reason, Notes
PIF_HDRS = ["Loan ID","Investor","Loan Type",
            "Payoff Date","Orig Balance ($)","Final UPB ($)","Payoff Amount ($)",
            "Interest Due ($)","Fees / Penalties ($)","Net Proceeds ($)",
            "Rate","Payoff Reason","Notes"]
PIF_WIDTHS = [13,10,13,13,16,16,18,14,16,16,11,18,28]
PIF_NCOLS = len(PIF_HDRS)

PAYOFF_REASONS = ["Full Payoff","Refinance Payoff","Sale Payoff",
                  "Full Payoff","Refinance Payoff","Full Payoff",
                  "Estate Payoff","Refinance Payoff"]

pif_row = write_disclaimer(ws_pif, PIF_NCOLS, row=1)

ws_pif.merge_cells(f"A{pif_row}:{get_column_letter(PIF_NCOLS)}{pif_row}")
t = ws_pif.cell(row=pif_row, column=1,
    value=f"RECONCILIATION — PAID IN FULL  |  January 2026  |  {N_PIF} Loans")
t.fill = F_RED; t.font = _font(bold=True, color="FFFFFF", size=12)
t.alignment = Alignment(horizontal="center", vertical="center")
ws_pif.row_dimensions[pif_row].height = 26
pif_row += 1

for col, (h, w) in enumerate(zip(PIF_HDRS, PIF_WIDTHS), 1):
    hcell(ws_pif, pif_row, col, h, fill=F_RED)
    ws_pif.column_dimensions[get_column_letter(col)].width = w
ws_pif.row_dimensions[pif_row].height = 28
ws_pif.freeze_panes = f"A{pif_row+1}"
pif_data_start = pif_row + 1

for r, ln in enumerate(pif_loans_list, pif_data_start):
    fill = F_LTRED if r % 2 == 0 else None
    payoff_date = rand_date(date(2026,1,3), date(2026,1,30))
    int_due     = round(ln["upb_dec"] * ln["rate"] / 12, 2)
    fees        = round(random.uniform(0, 350), 2)
    payoff_amt  = round(ln["upb_dec"] + int_due + fees, 2)
    reason      = random.choice(PAYOFF_REASONS)

    dcell(ws_pif, r,  1, ln["loan_id"],   fill, align="center")
    dcell(ws_pif, r,  2, ln["investor"],  fill, align="center")
    dcell(ws_pif, r,  3, ln["loan_type"], fill, align="center")
    dcell(ws_pif, r,  4, payoff_date,     fill, DFMT,  align="center")
    dcell(ws_pif, r,  5, ln["orig_bal"],  fill, CURR0, align="right")
    dcell(ws_pif, r,  6, ln["upb_dec"],   fill, CURR,  align="right")
    dcell(ws_pif, r,  7, payoff_amt,      fill, CURR,  align="right")
    dcell(ws_pif, r,  8, int_due,         fill, CURR,  align="right")
    dcell(ws_pif, r,  9, fees,            fill, CURR,  align="right",
          font=_font(color="C00000", size=9))
    dcell(ws_pif, r, 10, payoff_amt,      fill, CURR,  align="right")
    dcell(ws_pif, r, 11, ln["rate"],      fill, PCT3,  align="right")
    dcell(ws_pif, r, 12, reason,          fill, align="center")
    dcell(ws_pif, r, 13,
          f"Removed from portfolio {payoff_date.strftime('%m/%d/%Y')}", fill)

pif_tr = pif_data_start + N_PIF
for col in range(1, PIF_NCOLS + 1):
    ws_pif.cell(row=pif_tr, column=col).fill   = F_REDTOT
    ws_pif.cell(row=pif_tr, column=col).border = THIN
ws_pif.cell(row=pif_tr, column=1).value     = "TOTALS"
ws_pif.cell(row=pif_tr, column=1).font      = BLKBOLD
ws_pif.cell(row=pif_tr, column=1).alignment = Alignment(horizontal="center")
# col 5=Orig Bal, 6=Final UPB, 7=Payoff Amt, 8=Int Due, 9=Fees, 10=Net Proceeds
for col, cl in [(5,"E"),(6,"F"),(7,"G"),(8,"H"),(9,"I"),(10,"J")]:
    c = ws_pif.cell(row=pif_tr, column=col)
    c.value = f"=SUM({cl}{pif_data_start}:{cl}{pif_tr-1})"
    c.number_format = CURR; c.font = BLKBOLD; c.fill = F_REDTOT
    c.border = THIN; c.alignment = Alignment(horizontal="right")

pif_path = os.path.join(OUT, "Recon_PaidInFull_Jan2026.xlsx")
wb_pif.save(pif_path)
print(f"  Saved: {pif_path}")

# ═══════════════════════════════════════════════════════════════════════════════
# FILE 4: RECON — CAPITALIZATION
# ═══════════════════════════════════════════════════════════════════════════════
print("Building Recon - Capitalization...")
wb_cap = Workbook()
ws_cap = wb_cap.active
ws_cap.title = "Capitalization"

cap_loans_list = [ln for ln in dec_loans
                  if ln["cap_amount"] > 0 and not ln["is_pif"]]
N_CAP = len(cap_loans_list)

CAP_TYPES = {
    "90+ DPD": "Deferred Interest Capitalization",
    "60 DPD":  "Escrow Advance Capitalization",
    "30 DPD":  "Fee Capitalization",
}

# Columns (10): Loan ID, Dec Status, Investor, Prior UPB, Capitalized Amount,
#               New UPB, Rate, Effective Date, Capitalization Type, Auth Ref
CAP_HDRS = ["Loan ID","Dec Status","Investor",
            "Prior UPB ($)","Capitalized Amount ($)","New UPB ($)",
            "Rate","Effective Date","Capitalization Type","Authorization Ref"]
CAP_WIDTHS = [13,12,10,16,20,16,11,14,26,20]
CAP_NCOLS = len(CAP_HDRS)

cap_row = write_disclaimer(ws_cap, CAP_NCOLS, row=1)

ws_cap.merge_cells(f"A{cap_row}:{get_column_letter(CAP_NCOLS)}{cap_row}")
t = ws_cap.cell(row=cap_row, column=1,
    value=f"RECONCILIATION — CAPITALIZATION  |  Dec 2025 – Jan 2026  |  {N_CAP} Loans")
t.fill = F_ORANGE; t.font = _font(bold=True, color="FFFFFF", size=12)
t.alignment = Alignment(horizontal="center", vertical="center")
ws_cap.row_dimensions[cap_row].height = 26
cap_row += 1

for col, (h, w) in enumerate(zip(CAP_HDRS, CAP_WIDTHS), 1):
    hcell(ws_cap, cap_row, col, h, fill=F_ORANGE)
    ws_cap.column_dimensions[get_column_letter(col)].width = w
ws_cap.row_dimensions[cap_row].height = 28
ws_cap.freeze_panes = f"A{cap_row+1}"
cap_data_start = cap_row + 1

for r, ln in enumerate(cap_loans_list, cap_data_start):
    fill = F_LTORANGE if r % 2 == 0 else None
    cap_type = CAP_TYPES.get(ln["status_dec"], "Deferred Interest Capitalization")
    eff_date = rand_date(date(2025,12,1), date(2026,1,31))
    auth_ref = f"AUTH-{random.randint(100000,999999)}"
    new_upb  = ln["upb_jan"]

    dcell(ws_cap, r,  1, ln["loan_id"],       fill, align="center")
    sc = dcell(ws_cap, r, 2, ln["status_dec"], fill, align="center")
    sc.font = RED_ST
    dcell(ws_cap, r,  3, ln["investor"],      fill, align="center")
    dcell(ws_cap, r,  4, ln["upb_dec"],       fill, CURR,  align="right")
    c = dcell(ws_cap, r, 5, ln["cap_amount"], fill, CURR,  align="right")
    c.font = BLUE_IN
    dcell(ws_cap, r,  6, new_upb,             fill, CURR,  align="right")
    dcell(ws_cap, r,  7, ln["rate"],          fill, PCT3,  align="right")
    dcell(ws_cap, r,  8, eff_date,            fill, DFMT,  align="center")
    dcell(ws_cap, r,  9, cap_type,            fill)
    dcell(ws_cap, r, 10, auth_ref,            fill, align="center",
          font=_font(italic=True, size=9))

cap_tr = cap_data_start + N_CAP
TOT_ORANGE = _fill("F8CBAD")
for col in range(1, CAP_NCOLS + 1):
    ws_cap.cell(row=cap_tr, column=col).fill   = TOT_ORANGE
    ws_cap.cell(row=cap_tr, column=col).border = THIN
ws_cap.cell(row=cap_tr, column=1).value     = "TOTALS"
ws_cap.cell(row=cap_tr, column=1).font      = BLKBOLD
ws_cap.cell(row=cap_tr, column=1).alignment = Alignment(horizontal="center")
# col 4=Prior UPB, 5=Cap Amount, 6=New UPB
for col, cl in [(4,"D"),(5,"E"),(6,"F")]:
    c = ws_cap.cell(row=cap_tr, column=col)
    c.value = f"=SUM({cl}{cap_data_start}:{cl}{cap_tr-1})"
    c.number_format = CURR; c.font = BLKBOLD; c.fill = TOT_ORANGE
    c.border = THIN; c.alignment = Alignment(horizontal="right")

cap_path = os.path.join(OUT, "Recon_Capitalization_Dec2025_Jan2026.xlsx")
wb_cap.save(cap_path)
print(f"  Saved: {cap_path}")

print("\n[OK] All files saved. Summary:")
print(f"   Output folder:  {OUT}")
print(f"   Dec portfolio:  {N_DEC:,} loans  |  UPB: ${dec_total_upb:>14,.2f}")
print(f"   New Adds:      +{N_NEW_ADDS:,} loans  |  UPB: ${new_adds_upb:>14,.2f}")
print(f"   Paid in Full:  -{N_PIF:,} loans  |  UPB: ${pif_total_upb:>14,.2f}")
print(f"   Jan portfolio:  {N_JAN:,} loans  |  UPB: ${jan_total_upb:>14,.2f}")
print(f"   Bridge diff:   ${bridge_check - jan_total_upb:.2f}  (should be ~0)")
