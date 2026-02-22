"""
build_year_simulation.py — MSR Portfolio Year Simulation (Feb–Dec 2026)

Extends the Jan 2026 portfolio month-by-month through December 2026.
Generates: monthly tapes, PIF recon, new add recon, recon reports,
           subservicer dirty tapes (error months), validation reports,
           and a master dashboard Excel file.

Usage:
    python build_year_simulation.py              # full Feb–Dec run
    python build_year_simulation.py --from 10   # re-run Oct–Dec (reads Sep tape from disk)
    python build_year_simulation.py --month 3   # re-run March only (reads Feb tape from disk)
"""

import os, sys, math, random, argparse, subprocess, copy, re
from datetime import date, timedelta
from collections import defaultdict

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DISCLAIMER  = ("SIMULATED DATA — All loan information is synthetic and "
               "generated for testing purposes only.")

# ══════════════════════════════════════════════════════════════════════════════
#  SIMULATION PARAMETERS
# ══════════════════════════════════════════════════════════════════════════════

MONTH_NAME = {
    1:"Jan", 2:"Feb", 3:"Mar", 4:"Apr",  5:"May",  6:"Jun",
    7:"Jul", 8:"Aug", 9:"Sep", 10:"Oct", 11:"Nov", 12:"Dec",
}

def month_label(m):
    return f"{MONTH_NAME[m]}2026"

def tape_date(m):
    """Last calendar day of month m, 2026."""
    if m in (1,3,5,7,8,10,12): return date(2026, m, 31)
    if m in (4,6,9,11):        return date(2026, m, 30)
    return date(2026, 2, 28)

# Annual CPR -> drives monthly PIF count
CPR = {2:0.11, 3:0.12, 4:0.13, 5:0.14, 6:0.15,
       7:0.15, 8:0.14, 9:0.13, 10:0.22, 11:0.16, 12:0.12}

# Seasonal new add volumes
N_NEW_ADDS = {2:155, 3:175, 4:200, 5:220, 6:235,
              7:230, 8:210, 9:185, 10:240, 11:160, 12:130}

# Origination rate range (low, high) for each month's new adds
# Baseline 6.25 -> Oct drop to 5.5 -> drift back to 6.0 by Dec
ORIG_RATE = {
    2:(0.0600,0.0650), 3:(0.0600,0.0650), 4:(0.0610,0.0645),
    5:(0.0615,0.0640), 6:(0.0618,0.0638), 7:(0.0615,0.0640),
    8:(0.0610,0.0645), 9:(0.0610,0.0645),
    10:(0.0525,0.0575),   # Oct rate drop: 5.5% center
    11:(0.0550,0.0600),   # drifting back
    12:(0.0575,0.0625),   # back near 6.0%
}

# Error counts — learning curve: heavy early, taper, spike Oct
N_ERRORS = {2:0, 3:12, 4:10, 5:7, 6:4, 7:2, 8:3, 9:2, 10:9, 11:5, 12:3}

HARD_STOP_TYPES = {"upb_x10","upb_zero","rate_whole","nsf_bps","duplicate","missing"}
YELLOW_TYPES    = {"nsf_pct","status_skip","pi_inflated","ndd_past","rem_unchanged"}

def error_plan(month):
    """Return ordered list of error type strings for a given month."""
    plans = {
        3: ["upb_x10","upb_x10","upb_x10","upb_zero","rate_whole","rate_whole",
            "nsf_bps","nsf_bps","duplicate","missing","nsf_pct","nsf_pct"],
        4: ["upb_x10","upb_x10","rate_whole","nsf_bps","duplicate","missing",
            "nsf_pct","nsf_pct","status_skip","status_skip"],
        5: ["upb_x10","missing","nsf_pct","nsf_pct","status_skip","status_skip","pi_inflated"],
        6: ["missing","nsf_pct","status_skip","ndd_past"],
        7: ["nsf_pct","ndd_past"],
        8: ["upb_x10","nsf_pct","rem_unchanged"],
        9: ["nsf_pct","status_skip"],
        10:["rate_whole","rate_whole","rate_whole","nsf_bps","nsf_bps",
            "nsf_pct","nsf_pct","pi_inflated","missing"],
        11:["rate_whole","nsf_pct","status_skip","pi_inflated","ndd_past"],
        12:["nsf_pct","status_skip","rem_unchanged"],
    }
    return plans.get(month, [])

def investor_pool(m):
    if m in (5,6,7,8):  return ["FNMA"]*40 + ["FHLMC"]*28 + ["GNMA"]*27 + ["Portfolio"]*5
    if m == 10:          return ["FNMA"]*50 + ["FHLMC"]*35 + ["GNMA"]*12 + ["Portfolio"]*3
    return ["FNMA"]*45 + ["FHLMC"]*30 + ["GNMA"]*20 + ["Portfolio"]*5

def purpose_pool(m):
    if m == 10: return ["Purchase"]*40 + ["Refinance"]*60
    if m == 11: return ["Purchase"]*55 + ["Refinance"]*45
    return ["Purchase"]*65 + ["Refinance"]*35

LTYPE_POOL = ["Conventional"]*60 + ["FHA"]*25 + ["VA"]*10 + ["USDA"]*5
TERM_POOL  = [360]*75 + [180]*10 + [240]*15

# ══════════════════════════════════════════════════════════════════════════════
#  MATH / DOMAIN HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def calc_pi(orig_bal, rate, term):
    r = rate / 12
    if r < 1e-9: return orig_bal / term
    return orig_bal * r * (1+r)**term / ((1+r)**term - 1)

def calc_upb_at_elapsed(orig_bal, rate, term, elapsed):
    r = rate / 12
    if r < 1e-9: return orig_bal * (1 - elapsed/term)
    return orig_bal * ((1+r)**term - (1+r)**elapsed) / ((1+r)**term - 1)

def sched_principal(upb, rate, pi):
    return pi - upb * rate / 12

def rand_orig_bal():
    base = random.gauss(360000, 85000)
    return round(max(120000, min(800000, base)) / 5000) * 5000

def calc_nsf(investor):
    if investor == "GNMA":
        return round(random.triangular(0.0019, 0.0069, 0.0044), 4)
    return 0.0025

def ndd_for_status(status, tape_dt):
    """First unpaid installment date given loan status and tape date."""
    y, mo = tape_dt.year, tape_dt.month
    if   status == "Current":  mo += 1
    elif status == "30 DPD":   pass
    elif status == "60 DPD":   mo -= 1
    elif status == "90+ DPD":  mo -= 2
    while mo > 12: mo -= 12; y += 1
    while mo < 1:  mo += 12; y -= 1
    return date(y, mo, 1)

# ══════════════════════════════════════════════════════════════════════════════
#  STYLE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, color="000000", size=9, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def _bdr(style="thin", color="B8CCE4"):
    s = Side(style=style, color=color); return Border(left=s,right=s,top=s,bottom=s)

F_NAVY  = _fill("1F4E79"); F_BLUE  = _fill("2E75B6"); F_LTBLUE= _fill("D6E4F0")
F_TOTAL = _fill("BDD7EE"); F_WHITE = _fill("FFFFFF")
F_GREEN = _fill("375623"); F_LTGRN = _fill("E2EFDA"); F_GRNTOT= _fill("A9D18E")
F_RED   = _fill("7B2C2C"); F_LTRED = _fill("FCE4D6"); F_REDTOT= _fill("F4CCCC")
F_ORNG  = _fill("833C00"); F_LORNG = _fill("FCE4D6")
F_DISC  = _fill("FFF2CC"); F_AMBR  = _fill("FFF2CC")
F_DASH  = _fill("1F4E79")

WHBOLD  = _font(bold=True, color="FFFFFF", size=9)
BLKBOLD = _font(bold=True, size=9)
NORMAL  = _font(size=9)
GRN_ST  = _font(bold=True, color="006400", size=9)
RED_ST  = _font(bold=True, color="C00000", size=9)
DISC_FNT= _font(bold=True, color="7B3F00", size=9, italic=True)
THIN    = _bdr("thin","B8CCE4")

CURR = '#,##0.00'; CURR0 = '#,##0'; PCT3 = '0.000%'
DFMT = 'MM/DD/YYYY'; NUM0 = '#,##0'

def hcell(ws, row, col, val, fill=None, font=None, align="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill or F_NAVY; c.font = font or WHBOLD
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = THIN; return c

def dcell(ws, row, col, val, fill=None, fmt=None, font=None, align="left"):
    c = ws.cell(row=row, column=col, value=val)
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    c.font = font or NORMAL
    c.alignment = Alignment(horizontal=align, vertical="center"); c.border = THIN
    return c

def write_disclaimer(ws, ncols, row=1):
    lc = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{lc}{row}")
    c = ws.cell(row=row, column=1, value=DISCLAIMER)
    c.fill = F_DISC; c.font = DISC_FNT
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = THIN
    ws.row_dimensions[row].height = 16
    return row + 1

def status_font(s):
    if s in (None,"Current"): return GRN_ST
    if s == "Paid in Full":   return _font(bold=True, color="1F4E79", size=9)
    return RED_ST

# ══════════════════════════════════════════════════════════════════════════════
#  TAPE WRITE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

TAPE_HEADERS = [
    "Loan ID","Loan Type","Purpose","Investor",
    "Orig Date","Original Bal ($)","Current UPB ($)","Rate","Net Serv Fee","Rem Term",
    "Maturity","P&I ($)","Escrow ($)","Total Pmt ($)","Status","Next Due Date"
]
TAPE_NCOLS = 16
TAPE_COL_WIDTHS = [13,13,11,10,14,16,14,13,12,11,13,13,12,13,13,13]

def set_tape_col_widths(ws):
    for col, w in zip("ABCDEFGHIJKLMNOP", TAPE_COL_WIDTHS):
        ws.column_dimensions[col].width = w

def write_tape_header(ws, title_text, title_fill, row1=1):
    """Write disclaimer + title + column headers; return first data row."""
    nr = write_disclaimer(ws, TAPE_NCOLS, row=row1)
    lc = get_column_letter(TAPE_NCOLS)
    ws.merge_cells(f"A{nr}:{lc}{nr}")
    tc = ws.cell(row=nr, column=1, value=title_text)
    tc.fill = title_fill; tc.font = _font(bold=True, color="FFFFFF", size=12)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[nr].height = 26; nr += 1
    for col, h in enumerate(TAPE_HEADERS, 1): hcell(ws, nr, col, h)
    ws.row_dimensions[nr].height = 30; ws.freeze_panes = f"A{nr+1}"
    set_tape_col_widths(ws)
    return nr + 1

def write_tape_row(ws, row, ln):
    fill = F_LTBLUE if row % 2 == 0 else None
    dcell(ws, row,  1, ln["loan_id"],   fill, align="center")
    dcell(ws, row,  2, ln["loan_type"], fill, align="center")
    dcell(ws, row,  3, ln["purpose"],   fill, align="center")
    dcell(ws, row,  4, ln["investor"],  fill, align="center")
    dcell(ws, row,  5, ln["orig_date"], fill, DFMT,  align="center")
    dcell(ws, row,  6, ln["orig_bal"],  fill, CURR0, align="right")
    dcell(ws, row,  7, ln["upb"],       fill, CURR,  align="right")
    dcell(ws, row,  8, ln["rate"],      fill, PCT3,  align="right")
    dcell(ws, row,  9, ln["nsf"],       fill, PCT3,  align="right")
    dcell(ws, row, 10, ln["remaining"], fill, NUM0,  align="center")
    dcell(ws, row, 11, ln["maturity"],  fill, DFMT,  align="center")
    dcell(ws, row, 12, ln["pi"],        fill, CURR,  align="right")
    dcell(ws, row, 13, ln["escrow"],    fill, CURR,  align="right")
    dcell(ws, row, 14, ln["total_pmt"],fill, CURR,  align="right")
    sc = dcell(ws, row, 15, ln["status"], fill, align="center")
    sc.font = status_font(ln["status"])
    dcell(ws, row, 16, ln["ndd"],       fill, DFMT,  align="center")

def write_tape_totals(ws, data_start, n_loans, fill=F_TOTAL):
    tr = data_start + n_loans; last = tr - 1
    for col in range(1, TAPE_NCOLS+1):
        c = ws.cell(row=tr, column=col); c.fill = fill; c.border = THIN
    ws.cell(row=tr, column=1).value = "TOTALS / AVERAGES"
    ws.cell(row=tr, column=1).font  = BLKBOLD
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="center")
    for col, cl, fmt in [(6,"F",CURR0),(7,"G",CURR),(12,"L",CURR),(13,"M",CURR),(14,"N",CURR)]:
        c = ws.cell(row=tr, column=col)
        c.value = f"=SUM({cl}{data_start}:{cl}{last})"
        c.number_format = fmt; c.font = BLKBOLD; c.fill = fill
        c.border = THIN; c.alignment = Alignment(horizontal="right")
    for col, cl in [(8,"H"),(9,"I")]:
        c = ws.cell(row=tr, column=col)
        c.value = f"=AVERAGE({cl}{data_start}:{cl}{last})"
        c.number_format = PCT3; c.font = BLKBOLD; c.fill = fill
        c.border = THIN; c.alignment = Alignment(horizontal="right")
    c = ws.cell(row=tr, column=10)
    c.value = f"=COUNT(G{data_start}:G{last})"
    c.number_format = NUM0; c.font = BLKBOLD; c.fill = fill
    c.border = THIN; c.alignment = Alignment(horizontal="center")
    return tr

# ══════════════════════════════════════════════════════════════════════════════
#  LOAD STARTING PORTFOLIO
# ══════════════════════════════════════════════════════════════════════════════

def load_tape_from_xlsx(filepath, sheet_name=None):
    """
    Read a standard MSR tape xlsx and return list of loan dicts.
    Handles both single-sheet monthly files and named sheets (e.g., 'Jan 2026').
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    loans = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        loan_id = row[0]
        if not (loan_id and str(loan_id).startswith("MSR")):
            continue
        status = str(row[14]).strip() if row[14] else "Current"
        if status in ("Paid in Full", "TOTALS / AVERAGES"):
            continue
        try:
            upb = float(row[6]) if row[6] is not None else 0.0
            if upb <= 0: continue
            orig_dt = row[4]  if isinstance(row[4],  date) else date(2020,1,1)
            mat_dt  = row[10] if isinstance(row[10], date) else date(2055,1,1)
            ndd_dt  = row[15] if isinstance(row[15], date) else date(2026,3,1)
            ln = {
                "loan_id":   str(row[0]),
                "loan_type": str(row[1]) if row[1] else "Conventional",
                "purpose":   str(row[2]) if row[2] else "Purchase",
                "investor":  str(row[3]) if row[3] else "FNMA",
                "orig_date": orig_dt,
                "orig_bal":  float(row[5]) if row[5] else 350000.0,
                "upb":       upb,
                "rate":      float(row[7]) if row[7] else 0.065,
                "nsf":       float(row[8]) if row[8] else 0.0025,
                "remaining": int(row[9])   if row[9] else 300,
                "maturity":  mat_dt,
                "pi":        float(row[11]) if row[11] else 1000.0,
                "escrow":    float(row[12]) if row[12] else 400.0,
                "total_pmt": float(row[13]) if row[13] else 1400.0,
                "status":    status,
                "ndd":       ndd_dt,
                "board_month": 1,
                "transfer_date": None,
            }
            loans.append(ln)
        except (TypeError, ValueError, IndexError):
            continue
    return loans

# ══════════════════════════════════════════════════════════════════════════════
#  MONTHLY PORTFOLIO SIMULATION
# ══════════════════════════════════════════════════════════════════════════════

def select_pifs(portfolio, month, n_pif):
    """
    Select n_pif loans to pay off.
    October: weight 65% toward high-rate loans (refi incentive at 5.5%).
    Other months: random from current-status loans.
    """
    current = [ln for ln in portfolio if ln["status"] == "Current"]
    n_pif   = min(n_pif, len(current))
    if n_pif == 0:
        return []
    if month == 10:
        high_rate = [ln for ln in current if ln["rate"] >= 0.062]
        low_rate  = [ln for ln in current if ln["rate"] <  0.062]
        n_high = min(len(high_rate), round(n_pif * 0.65))
        n_low  = min(len(low_rate),  n_pif - n_high)
        if n_high + n_low < n_pif:
            n_high = min(len(high_rate), n_pif - n_low)
        return random.sample(high_rate, n_high) + random.sample(low_rate, n_low)
    return random.sample(current, n_pif)

def migrate_dq(portfolio, pif_ids, month):
    """Apply Markov DQ migration. Modifies loans in place."""
    td = tape_date(month)
    new_dq_prob = 0.005 if month <= 9 else 0.007  # slight drift upward Q4

    for ln in portfolio:
        if ln["loan_id"] in pif_ids:
            continue
        st = ln["status"]
        if st == "Current":
            if random.random() < new_dq_prob:
                ln["status"] = "30 DPD"
                ln["ndd"] = ndd_for_status("30 DPD", td)
        elif st == "30 DPD":
            r = random.random()
            if r < 0.35:
                ln["status"] = "Current"; ln["ndd"] = ndd_for_status("Current", td)
            elif r < 0.60:
                ln["status"] = "60 DPD";  ln["ndd"] = ndd_for_status("60 DPD", td)
            # else stays 30 DPD; ndd stays
        elif st == "60 DPD":
            r = random.random()
            if r < 0.20:
                ln["status"] = "30 DPD";   ln["ndd"] = ndd_for_status("30 DPD", td)
            elif r < 0.50:
                ln["status"] = "90+ DPD";  ln["ndd"] = ndd_for_status("90+ DPD", td)
        elif st == "90+ DPD":
            if random.random() < 0.10:
                ln["status"] = "60 DPD";   ln["ndd"] = ndd_for_status("60 DPD", td)

def amortize(portfolio, pif_ids, month):
    """Amortize UPB and decrement remaining term. Modifies loans in place."""
    td = tape_date(month)
    for ln in portfolio:
        if ln["loan_id"] in pif_ids:
            continue
        if ln["status"] == "Current":
            sched  = sched_principal(ln["upb"], ln["rate"], ln["pi"])
            extra  = 0.0
            if random.random() < 0.015:  # ~1.5% chance of curtailment
                extra = round(random.choice([2000,5000,7500,10000,15000,20000]), 2)
            ln["upb"] = round(max(0.0, ln["upb"] - sched - extra), 2)
        else:
            cap = round(random.uniform(150, 1200), 2)
            # Cap at orig_bal to prevent unrealistic UPB > orig_bal accumulation
            ln["upb"] = round(min(ln["upb"] + cap, ln["orig_bal"]), 2)
        ln["remaining"] = max(0, ln["remaining"] - 1)
        ln["ndd"] = ndd_for_status(ln["status"], td)

def generate_new_adds(month, n_adds):
    """Generate n_adds new loan dicts for the given month."""
    td     = tape_date(month)
    r_lo, r_hi = ORIG_RATE[month]
    inv_pool = investor_pool(month)
    pur_pool = purpose_pool(month)
    loans    = []

    for seq in range(n_adds):
        loan_id  = f"MSR{200000 + month*1000 + seq:06d}"
        investor = random.choice(inv_pool)
        lt       = random.choice(LTYPE_POOL)
        purpose  = random.choice(pur_pool)
        term     = random.choice(TERM_POOL)
        orig_bal = rand_orig_bal()
        rate     = round(random.uniform(r_lo, r_hi), 4)
        nsf      = calc_nsf(investor)
        pi       = round(calc_pi(orig_bal, rate, term), 2)
        escrow   = round(random.uniform(200, 700), 2)
        elapsed  = random.randint(1, 3)
        remaining = term - elapsed

        months_back = random.randint(1, 4)
        orig_m = month - months_back; orig_y = 2026
        while orig_m < 1: orig_m += 12; orig_y -= 1
        orig_date = date(orig_y, orig_m, random.randint(1, 28))

        mat_year = orig_date.year + term // 12
        maturity = date(min(mat_year, 2062), orig_date.month, 1)
        upb      = round(calc_upb_at_elapsed(orig_bal, rate, term, elapsed), 2)

        loans.append({
            "loan_id":       loan_id,
            "loan_type":     lt,
            "purpose":       purpose,
            "investor":      investor,
            "orig_date":     orig_date,
            "orig_bal":      orig_bal,
            "upb":           upb,
            "rate":          rate,
            "nsf":           nsf,
            "remaining":     remaining,
            "maturity":      maturity,
            "pi":            pi,
            "escrow":        escrow,
            "total_pmt":     round(pi + escrow, 2),
            "status":        "Current",
            "ndd":           ndd_for_status("Current", td),
            "board_month":   month,
            "transfer_date": date(2026, month, random.randint(1, 25)),
        })
    return loans

# ══════════════════════════════════════════════════════════════════════════════
#  FILE WRITERS
# ══════════════════════════════════════════════════════════════════════════════

def write_monthly_tape(portfolio, month, filepath):
    title_fills = {
        2:F_BLUE, 3:F_GREEN, 4:F_NAVY, 5:_fill("196F3D"), 6:_fill("7B2C2C"),
        7:_fill("833C00"), 8:_fill("2E75B6"), 9:_fill("375623"),
        10:_fill("1F4E79"), 11:_fill("7B2C2C"), 12:_fill("833C00"),
    }
    td    = tape_date(month)
    fill  = title_fills.get(month, F_NAVY)
    title = f"MSR TAPE — {MONTH_NAME[month].upper()} 2026 (As of {td.strftime('%m/%d/%Y')})"

    wb = Workbook()
    ws = wb.active
    ws.title = f"{MONTH_NAME[month]} 2026"
    ws.sheet_view.showGridLines = False

    data_start = write_tape_header(ws, title, fill)
    for r, ln in enumerate(portfolio, data_start):
        write_tape_row(ws, r, ln)
    write_tape_totals(ws, data_start, len(portfolio))
    wb.save(filepath)
    print(f"    Tape:       {os.path.basename(filepath)}")

def write_pif_recon(pif_loans, month, filepath):
    PIF_HDRS   = ["Loan ID","Investor","Loan Type","Payoff Date","Orig Balance ($)",
                  "Final UPB ($)","Payoff Amount ($)","Interest Due ($)",
                  "Fees / Penalties ($)","Rate","Payoff Reason","Notes"]
    PIF_WIDTHS = [13,10,13,13,16,16,18,14,16,11,18,30]
    nc = len(PIF_HDRS)
    REASONS = ["Full Payoff","Refinance Payoff","Sale Payoff","Refinance Payoff",
               "Full Payoff","Estate Payoff","Full Payoff","Refinance Payoff"]

    wb = Workbook(); ws = wb.active; ws.title = "Paid in Full"
    ws.sheet_view.showGridLines = False
    r = write_disclaimer(ws, nc)
    ws.merge_cells(f"A{r}:{get_column_letter(nc)}{r}")
    t = ws.cell(row=r, column=1,
        value=f"RECONCILIATION — PAID IN FULL  |  {MONTH_NAME[month]} 2026  |  {len(pif_loans)} Loans")
    t.fill = F_RED; t.font = _font(bold=True, color="FFFFFF", size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 26; r += 1
    for col, (h, w) in enumerate(zip(PIF_HDRS, PIF_WIDTHS), 1):
        hcell(ws, r, col, h, fill=F_RED)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[r].height = 28; ws.freeze_panes = f"A{r+1}"
    data_start = r + 1
    for row, ln in enumerate(pif_loans, data_start):
        fill = F_LTRED if row % 2 == 0 else None
        payoff_date = date(2026, month, random.randint(2, 27))
        int_due  = round(ln["upb"] * ln["rate"] / 12, 2)
        fees     = round(random.uniform(0, 350), 2)
        pay_amt  = round(ln["upb"] + int_due + fees, 2)
        reason   = random.choice(REASONS)
        dcell(ws, row,  1, ln["loan_id"],   fill, align="center")
        dcell(ws, row,  2, ln["investor"],  fill, align="center")
        dcell(ws, row,  3, ln["loan_type"], fill, align="center")
        dcell(ws, row,  4, payoff_date,     fill, DFMT,  align="center")
        dcell(ws, row,  5, ln["orig_bal"],  fill, CURR0, align="right")
        dcell(ws, row,  6, ln["upb"],       fill, CURR,  align="right")
        dcell(ws, row,  7, pay_amt,         fill, CURR,  align="right")
        dcell(ws, row,  8, int_due,         fill, CURR,  align="right")
        dcell(ws, row,  9, fees,            fill, CURR,  align="right",
              font=_font(color="C00000",size=9))
        dcell(ws, row, 10, ln["rate"],      fill, PCT3,  align="right")
        dcell(ws, row, 11, reason,          fill, align="center")
        dcell(ws, row, 12,
              f"Removed from portfolio {payoff_date.strftime('%m/%d/%Y')}", fill)
    # totals
    tr = data_start + len(pif_loans)
    for col in range(1, nc+1):
        ws.cell(row=tr, column=col).fill   = F_REDTOT
        ws.cell(row=tr, column=col).border = THIN
    ws.cell(row=tr, column=1).value = "TOTALS"; ws.cell(row=tr, column=1).font = BLKBOLD
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="center")
    for col, cl in [(5,"E"),(6,"F"),(7,"G"),(8,"H"),(9,"I")]:
        c = ws.cell(row=tr, column=col)
        c.value = f"=SUM({cl}{data_start}:{cl}{tr-1})"
        c.number_format = CURR; c.font = BLKBOLD; c.fill = F_REDTOT
        c.border = THIN; c.alignment = Alignment(horizontal="right")
    wb.save(filepath)
    print(f"    PIF recon:  {os.path.basename(filepath)}")

def write_new_add_recon(new_adds, month, filepath):
    NA_HDRS   = ["Loan ID","Loan Type","Purpose","Investor","Origination Date",
                 "Transfer / Add Date","Orig Balance ($)","Current UPB ($)","Rate",
                 "Net Serv Fee","Loan Term","Rem Term","P&I ($)","Escrow ($)",
                 "Total Pmt ($)","Status"]
    NA_WIDTHS = [13,13,11,10,14,14,16,14,11,12,10,10,13,12,13,12]
    nc = len(NA_HDRS)

    wb = Workbook(); ws = wb.active; ws.title = "New Adds"
    ws.sheet_view.showGridLines = False
    r = write_disclaimer(ws, nc)
    ws.merge_cells(f"A{r}:{get_column_letter(nc)}{r}")
    t = ws.cell(row=r, column=1,
        value=f"RECONCILIATION — NEW ADDS  |  {MONTH_NAME[month]} 2026  |  {len(new_adds)} Loans")
    t.fill = F_GREEN; t.font = _font(bold=True, color="FFFFFF", size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 26; r += 1
    for col, (h, w) in enumerate(zip(NA_HDRS, NA_WIDTHS), 1):
        hcell(ws, r, col, h, fill=F_GREEN)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[r].height = 28; ws.freeze_panes = f"A{r+1}"
    data_start = r + 1
    for row, ln in enumerate(new_adds, data_start):
        fill = F_LTGRN if row % 2 == 0 else None
        dcell(ws, row,  1, ln["loan_id"],       fill, align="center")
        dcell(ws, row,  2, ln["loan_type"],     fill, align="center")
        dcell(ws, row,  3, ln["purpose"],       fill, align="center")
        dcell(ws, row,  4, ln["investor"],      fill, align="center")
        dcell(ws, row,  5, ln["orig_date"],     fill, DFMT,  align="center")
        dcell(ws, row,  6, ln["transfer_date"], fill, DFMT,  align="center")
        dcell(ws, row,  7, ln["orig_bal"],      fill, CURR0, align="right")
        dcell(ws, row,  8, ln["upb"],           fill, CURR,  align="right")
        dcell(ws, row,  9, ln["rate"],          fill, PCT3,  align="right")
        dcell(ws, row, 10, ln["nsf"],           fill, PCT3,  align="right")
        dcell(ws, row, 11, ln["remaining"] + 1, fill, NUM0,  align="center")  # approx term
        dcell(ws, row, 12, ln["remaining"],     fill, NUM0,  align="center")
        dcell(ws, row, 13, ln["pi"],            fill, CURR,  align="right")
        dcell(ws, row, 14, ln["escrow"],        fill, CURR,  align="right")
        dcell(ws, row, 15, ln["total_pmt"],     fill, CURR,  align="right")
        dcell(ws, row, 16, "Current",           fill, align="center", font=GRN_ST)
    # totals
    tr = data_start + len(new_adds)
    for col in range(1, nc+1):
        ws.cell(row=tr, column=col).fill   = F_GRNTOT
        ws.cell(row=tr, column=col).border = THIN
    ws.cell(row=tr, column=1).value = "TOTALS / AVERAGES"
    ws.cell(row=tr, column=1).font  = BLKBOLD
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="center")
    for col, cl, fmt in [(7,"G",CURR),(8,"H",CURR),(13,"M",CURR),(14,"N",CURR),(15,"O",CURR)]:
        c = ws.cell(row=tr, column=col)
        c.value = f"=SUM({cl}{data_start}:{cl}{tr-1})"
        c.number_format = fmt; c.font = BLKBOLD; c.fill = F_GRNTOT
        c.border = THIN; c.alignment = Alignment(horizontal="right")
    for col, cl in [(9,"I"),(10,"J")]:
        c = ws.cell(row=tr, column=col)
        c.value = f"=AVERAGE({cl}{data_start}:{cl}{tr-1})"
        c.number_format = PCT3; c.font = BLKBOLD; c.fill = F_GRNTOT
        c.border = THIN; c.alignment = Alignment(horizontal="right")
    wb.save(filepath)
    print(f"    New adds:   {os.path.basename(filepath)}")

# ══════════════════════════════════════════════════════════════════════════════
#  ERROR INJECTION
# ══════════════════════════════════════════════════════════════════════════════

def inject_errors(portfolio, month, pif_ids):
    """
    Deep-copy portfolio, inject errors per error_plan(month).
    Returns (dirty_loans, error_log).
    """
    plan   = error_plan(month)
    dirty  = [copy.copy(ln) for ln in portfolio]
    errors = []
    used   = set()   # indices already modified

    def pick(status="Current", investor=None):
        cands = [i for i, ln in enumerate(dirty)
                 if i not in used
                 and (status is None or ln["status"] == status)
                 and (investor is None or ln["investor"] == investor)]
        if not cands:
            cands = [i for i in range(len(dirty)) if i not in used]
        return random.choice(cands) if cands else None

    for err in plan:
        if err == "upb_x10":
            i = pick()
            if i is None: continue
            used.add(i); orig = dirty[i]["upb"]
            dirty[i]["upb"] = round(orig * 10, 2)
            errors.append({"loan_id":dirty[i]["loan_id"],"type":"UPB × 10 (extra zero)",
                "field":"Current UPB ($)","submitted":f"${dirty[i]['upb']:,.2f}",
                "expected":f"~${orig:,.2f}"})

        elif err == "upb_zero":
            i = pick()
            if i is None: continue
            used.add(i); dirty[i]["upb"] = 0.0
            errors.append({"loan_id":dirty[i]["loan_id"],"type":"UPB = Zero (active loan)",
                "field":"Current UPB ($)","submitted":"$0.00","expected":">$0 (active)"})

        elif err == "rate_whole":
            i = pick()
            if i is None: continue
            used.add(i); orig = dirty[i]["rate"]
            dirty[i]["rate"] = round(orig * 100, 4)
            errors.append({"loan_id":dirty[i]["loan_id"],"type":"Rate as Whole Number",
                "field":"Rate","submitted":f"{dirty[i]['rate']:.4f}",
                "expected":f"~{orig:.4f} (decimal)"})

        elif err == "nsf_bps":
            gnma = [i for i,ln in enumerate(dirty)
                    if i not in used and ln["investor"]=="GNMA"]
            i = random.choice(gnma) if gnma else pick()
            if i is None: continue
            used.add(i); orig = dirty[i]["nsf"]
            dirty[i]["nsf"] = round(orig * 10000, 1)
            errors.append({"loan_id":dirty[i]["loan_id"],"type":"NSF as Whole Basis Points",
                "field":"Net Serv Fee","submitted":f"{dirty[i]['nsf']}",
                "expected":f"~{orig:.4f} (decimal)"})

        elif err == "nsf_pct":
            i = pick()
            if i is None: continue
            used.add(i); orig = dirty[i]["nsf"]
            dirty[i]["nsf"] = round(orig * 100, 4)
            errors.append({"loan_id":dirty[i]["loan_id"],"type":"NSF as Percent",
                "field":"Net Serv Fee","submitted":f"{dirty[i]['nsf']:.4f}",
                "expected":f"~{orig:.4f} (decimal)"})

        elif err == "duplicate":
            i = pick()
            if i is None: continue
            used.add(i)
            dirty.append(copy.copy(dirty[i]))
            errors.append({"loan_id":dirty[i]["loan_id"],"type":"Duplicate Loan ID",
                "field":"Loan ID","submitted":"Appears 2+ times","expected":"Unique"})

        elif err == "missing":
            cands = [i for i,ln in enumerate(dirty)
                     if i not in used and ln["loan_id"] not in pif_ids]
            if not cands: continue
            i = random.choice(cands); used.add(i)
            removed_id = dirty[i]["loan_id"]
            dirty[i]["_remove"] = True
            errors.append({"loan_id":removed_id,"type":"Missing Loan (no PIF)",
                "field":"—","submitted":"Not present","expected":"Present (no PIF)"})

        elif err == "status_skip":
            cur = [i for i,ln in enumerate(dirty)
                   if i not in used and ln["status"]=="Current"]
            if not cur: continue
            i = random.choice(cur); used.add(i)
            dirty[i]["status"] = "90+ DPD"
            errors.append({"loan_id":dirty[i]["loan_id"],"type":"Status Bucket Skip",
                "field":"Status","submitted":"Current -> 90+ DPD",
                "expected":"Max 1-bucket change per month"})

        elif err == "pi_inflated":
            i = pick()
            if i is None: continue
            used.add(i); orig = dirty[i]["pi"]
            dirty[i]["pi"] = round(orig * 1.20, 2)
            dirty[i]["total_pmt"] = round(dirty[i]["pi"] + dirty[i]["escrow"], 2)
            errors.append({"loan_id":dirty[i]["loan_id"],"type":"P&I Inflated ~20%",
                "field":"P&I ($)","submitted":f"${dirty[i]['pi']:,.2f}",
                "expected":f"~${orig:,.2f}"})

        elif err == "ndd_past":
            i = pick()
            if i is None: continue
            used.add(i)
            bad = date(2025, random.randint(3, 8), 1)
            dirty[i]["ndd"] = bad
            errors.append({"loan_id":dirty[i]["loan_id"],"type":"Next Due Date in Past",
                "field":"Next Due Date","submitted":str(bad),
                "expected":"Future date for Current loan"})

        elif err == "rem_unchanged":
            i = pick()
            if i is None: continue
            used.add(i)
            dirty[i]["remaining"] += 1  # undo month decrement
            errors.append({"loan_id":dirty[i]["loan_id"],"type":"Remaining Term Unchanged",
                "field":"Rem Term","submitted":str(dirty[i]["remaining"]),
                "expected":str(dirty[i]["remaining"]-1)})

    dirty = [ln for ln in dirty if not ln.get("_remove")]
    return dirty, errors

def write_subservicer_tape(dirty_loans, error_log, month, filepath):
    """Write dirty subservicer tape with an Error Log reference sheet."""
    td    = tape_date(month)
    title = (f"MSR TAPE — {MONTH_NAME[month].upper()} 2026 — SUBSERVICER SUBMISSION "
             f"(As of {td.strftime('%m/%d/%Y')})")

    wb = Workbook(); ws = wb.active
    ws.title = f"{MONTH_NAME[month]} 2026"
    ws.sheet_view.showGridLines = False

    data_start = write_tape_header(ws, title, F_NAVY)
    for r, ln in enumerate(dirty_loans, data_start):
        write_tape_row(ws, r, ln)
    write_tape_totals(ws, data_start, len(dirty_loans))

    # Error Log reference sheet
    ws2 = wb.create_sheet("Error Log - Reference")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 22
    ws2.column_dimensions["E"].width = 22
    r2 = write_disclaimer(ws2, 5)
    ws2.merge_cells(f"A{r2}:E{r2}")
    t = ws2.cell(row=r2, column=1,
        value=f"ERROR LOG — {MONTH_NAME[month]} 2026 SUBSERVICER SUBMISSION ({len(error_log)} injected errors)")
    t.fill = F_RED; t.font = _font(bold=True, color="FFFFFF", size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[r2].height = 26; r2 += 1
    for col, h in enumerate(["Loan ID","Error Type","Field","Submitted Value","Expected Value"],1):
        hcell(ws2, r2, col, h, fill=F_RED)
    ws2.row_dimensions[r2].height = 22; r2 += 1
    for i, e in enumerate(error_log):
        fill = F_LTRED if i % 2 == 0 else None
        dcell(ws2, r2, 1, e["loan_id"], fill, align="center")
        dcell(ws2, r2, 2, e["type"],    fill)
        dcell(ws2, r2, 3, e["field"],   fill, align="center")
        dcell(ws2, r2, 4, e["submitted"],fill, align="right")
        dcell(ws2, r2, 5, e["expected"], fill)
        r2 += 1

    wb.save(filepath)
    print(f"    Subserv:    {os.path.basename(filepath)}")

# ══════════════════════════════════════════════════════════════════════════════
#  SUBPROCESS RUNNERS
# ══════════════════════════════════════════════════════════════════════════════

def run_recon(prior_path, prior_sheet, curr_path):
    recon_script = os.path.join(SCRIPT_DIR, "recon_automation.py")
    cmd = [sys.executable, recon_script, prior_path, curr_path]
    if prior_sheet:
        cmd += ["--sheet-m1", prior_sheet]
    try:
        subprocess.run(cmd, check=True, capture_output=True)
        print(f"    Recon:      generated")
    except subprocess.CalledProcessError as e:
        print(f"    Recon:      WARNING — recon_automation failed: {e.stderr[-200:]}")

def run_validation(prior_path, prior_sheet, subserv_path, pif_path, na_path):
    val_script = os.path.join(SCRIPT_DIR, "validate_msr_tape.py")
    cmd = [sys.executable, val_script,
           "--tape",           prior_path,
           "--submission",     subserv_path,
           "--pif-report",     pif_path,
           "--new-add-report", na_path]
    if prior_sheet:
        cmd += ["--prior-sheet", prior_sheet]
    try:
        result = subprocess.run(cmd, check=False, capture_output=True, text=True)
        print(f"    Validation: generated")
        return result.stdout
    except Exception as e:
        print(f"    Validation: WARNING — {e}")
        return ""

def parse_validation_counts(md_path):
    """Return (hard_stops, yellow_lights) parsed from a validation .md file."""
    if not os.path.exists(md_path):
        return None, None
    hs = yl = None
    with open(md_path, encoding="utf-8") as f:
        for line in f:
            if "HARD STOPS" in line and "|" in line:
                m = re.search(r"\*\*(\d+)\*\*", line)
                if m: hs = int(m.group(1))
            if "YELLOW LIGHTS" in line and "|" in line:
                m = re.search(r"\*\*(\d+)\*\*", line)
                if m: yl = int(m.group(1))
    return hs, yl

# ══════════════════════════════════════════════════════════════════════════════
#  MASTER DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

def write_master_dashboard(master_rows, jan_baseline):
    """Write MSR_Dashboard_2026.xlsx with year-over-year summary."""
    path = os.path.join(SCRIPT_DIR, "MSR_Dashboard_2026.xlsx")
    wb   = Workbook()

    # -- Tab 1: Year Overview --------------------------------------------------
    ws = wb.active; ws.title = "Year Overview"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    ncols = 11
    r = write_disclaimer(ws, ncols)
    ws.merge_cells(f"A{r}:{get_column_letter(ncols)}{r}")
    t = ws.cell(row=r, column=1,
        value="MSR PORTFOLIO — 2026 YEAR-TO-DATE DASHBOARD")
    t.fill = F_NAVY; t.font = _font(bold=True, color="FFFFFF", size=14)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 30; r += 1

    ws.merge_cells(f"A{r}:{get_column_letter(ncols)}{r}")
    sub = ws.cell(row=r, column=1,
        value="Portfolio Growth · PIF & New Add Activity · DQ Trends · Validation Results")
    sub.fill = _fill("2E75B6"); sub.font = _font(italic=True, color="FFFFFF", size=10)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18; r += 1

    hdrs = ["Month","Loan Count","Total UPB ($)","New Adds","PIFs",
            "Net Change","DQ Count","DQ %","Error Month","Hard Stops","Yellow Lights"]
    col_widths = [10,12,18,10,8,10,10,8,13,12,14]
    for col, (h, w) in enumerate(zip(hdrs, col_widths), 1):
        hcell(ws, r, col, h); ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[r].height = 26; r += 1

    # Jan baseline row
    jan_upb = sum(ln["upb"] for ln in jan_baseline)
    jan_dq  = sum(1 for ln in jan_baseline if ln["status"] != "Current")
    jan_row = {
        "month":1, "label":"Jan2026",
        "n_loans":len(jan_baseline), "total_upb":jan_upb,
        "n_new_adds":200, "n_pif":12,
        "dq_count":jan_dq, "n_errors":0, "hard_stops":0, "yellow_lights":0,
    }
    all_rows = [jan_row] + master_rows
    prev_count = None

    for i, row_data in enumerate(all_rows):
        fill = F_LTBLUE if i % 2 == 0 else None
        net  = (row_data["n_loans"] - prev_count) if prev_count is not None else None
        dq_pct = row_data["dq_count"] / row_data["n_loans"] if row_data["n_loans"] else 0
        err_month = "YES" if row_data["n_errors"] > 0 else "—"

        dcell(ws, r, 1,  MONTH_NAME[row_data["month"]]+" 2026", fill, align="center",
              font=_font(bold=True))
        dcell(ws, r, 2,  row_data["n_loans"],    fill, NUM0,  align="right")
        dcell(ws, r, 3,  row_data["total_upb"],  fill, CURR,  align="right")
        dcell(ws, r, 4,  row_data["n_new_adds"], fill, NUM0,  align="right",
              font=GRN_ST if row_data["n_new_adds"] else NORMAL)
        dcell(ws, r, 5,  row_data["n_pif"],      fill, NUM0,  align="right",
              font=RED_ST if row_data["n_pif"] else NORMAL)
        nc = dcell(ws, r, 6, net if net is not None else "—",
                   fill, '+#,##0;(#,##0);"-"' if net is not None else None, align="right")
        if net and net > 0:   nc.font = _font(bold=False, color="006400", size=9)
        elif net and net < 0: nc.font = _font(bold=False, color="C00000", size=9)
        dcell(ws, r, 7, row_data["dq_count"],    fill, NUM0,  align="right")
        dq_c = dcell(ws, r, 8, dq_pct, fill, '0.00%', align="right")
        if dq_pct > 0.015: dq_c.font = _font(bold=True, color="C00000", size=9)
        err_c = dcell(ws, r, 9, err_month, fill, align="center")
        if err_month == "YES": err_c.font = _font(bold=True, color="7B3F00", size=9)
        hs = row_data.get("hard_stops")
        yl = row_data.get("yellow_lights")
        hs_c = dcell(ws, r, 10, hs if hs is not None else "—", fill, align="right",
                     font=RED_ST if hs else NORMAL)
        yl_c = dcell(ws, r, 11, yl if yl is not None else "—", fill, align="right",
                     font=_font(bold=True, color="7B3F00", size=9) if yl else NORMAL)

        ws.row_dimensions[r].height = 18
        prev_count = row_data["n_loans"]
        r += 1

    # Totals
    total_na  = sum(rd["n_new_adds"] for rd in all_rows)
    total_pif = sum(rd["n_pif"]      for rd in all_rows)
    final_upb = all_rows[-1]["total_upb"] if all_rows else 0
    for col in range(1, ncols+1):
        c = ws.cell(row=r, column=col); c.fill = F_TOTAL; c.border = THIN
    ws.cell(row=r,column=1).value = "YTD TOTALS"; ws.cell(row=r,column=1).font = BLKBOLD
    ws.cell(row=r,column=1).alignment = Alignment(horizontal="center")
    dcell(ws,r,2,all_rows[-1]["n_loans"], F_TOTAL,NUM0, BLKBOLD,"right")
    dcell(ws,r,3,final_upb,               F_TOTAL,CURR, BLKBOLD,"right")
    dcell(ws,r,4,total_na,                F_TOTAL,NUM0, BLKBOLD,"right")
    dcell(ws,r,5,total_pif,               F_TOTAL,NUM0, BLKBOLD,"right")
    ws.row_dimensions[r].height = 20

    # -- Tab 2: Portfolio Growth (chart-ready data) ----------------------------
    ws2 = wb.create_sheet("Portfolio Growth")
    ws2.sheet_view.showGridLines = False
    r2  = write_disclaimer(ws2, 4)
    ws2.merge_cells(f"A{r2}:D{r2}")
    t2 = ws2.cell(row=r2, column=1,
        value="PORTFOLIO GROWTH DATA — Chart this in Excel (Insert -> Chart)")
    t2.fill = F_BLUE; t2.font = _font(bold=True, color="FFFFFF", size=11)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[r2].height = 24; r2 += 1
    for col, h in enumerate(["Month","Loan Count","Total UPB ($)","DQ Count"],1):
        hcell(ws2, r2, col, h)
        ws2.column_dimensions[get_column_letter(col)].width = [12,12,20,12][col-1]
    ws2.row_dimensions[r2].height = 22; r2 += 1
    for i, rd in enumerate(all_rows):
        fill = F_LTBLUE if i % 2 == 0 else None
        dcell(ws2, r2, 1, f"{MONTH_NAME[rd['month']]} 2026", fill, align="center",
              font=_font(bold=True))
        dcell(ws2, r2, 2, rd["n_loans"],   fill, NUM0, align="right")
        dcell(ws2, r2, 3, rd["total_upb"], fill, CURR, align="right")
        dcell(ws2, r2, 4, rd["dq_count"],  fill, NUM0, align="right")
        ws2.row_dimensions[r2].height = 16; r2 += 1

    wb.save(path)
    print(f"\n  Dashboard:  {os.path.basename(path)}")

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN SIMULATION LOOP
# ══════════════════════════════════════════════════════════════════════════════

def run_month(month, portfolio, master_data, prior_tape_path, prior_sheet):
    print(f"\n  {'-'*56}")
    print(f"  {MONTH_NAME[month].upper()} 2026"
          f"   |  CPR={CPR[month]:.0%}  "
          f"  New Adds={N_NEW_ADDS[month]}  "
          f"  Errors={N_ERRORS[month]}")
    print(f"  {'-'*56}")

    td    = tape_date(month)
    label = month_label(month)

    # 1 — Select PIFs
    cpr   = CPR[month]
    smm   = 1 - (1 - cpr) ** (1/12)
    n_pif = max(5, round(len(portfolio) * smm))
    pifs  = select_pifs(portfolio, month, n_pif)
    pif_ids = {ln["loan_id"] for ln in pifs}
    print(f"    SMM={smm:.4%}  ->  {len(pifs)} PIFs from {len(portfolio):,} loans")

    # 2 — DQ migration (pre-amortization)
    migrate_dq(portfolio, pif_ids, month)

    # 3 — Amortize survivors
    amortize(portfolio, pif_ids, month)

    # 4 — New adds
    new_adds = generate_new_adds(month, N_NEW_ADDS[month])

    # 5 — Build updated portfolio
    survivors = [ln for ln in portfolio if ln["loan_id"] not in pif_ids]
    portfolio  = survivors + new_adds
    total_upb  = sum(ln["upb"] for ln in portfolio)
    dq_count   = sum(1 for ln in portfolio if ln["status"] != "Current")
    print(f"    Portfolio:  {len(portfolio):,} loans  |  UPB: ${total_upb:>14,.0f}")

    # 6 — Write files
    tape_path = os.path.join(SCRIPT_DIR, f"MSR_Tape_{label}.xlsx")
    pif_path  = os.path.join(SCRIPT_DIR, f"Recon_PaidInFull_{label}.xlsx")
    na_path   = os.path.join(SCRIPT_DIR, f"Recon_NewAdds_{label}.xlsx")

    write_monthly_tape(portfolio, month, tape_path)
    write_pif_recon(pifs, month, pif_path)
    write_new_add_recon(new_adds, month, na_path)

    # 7 — Recon report (prior -> current)
    run_recon(prior_tape_path, prior_sheet, tape_path)

    # 8 — Error injection + validation
    n_err = N_ERRORS[month]
    if n_err > 0:
        dirty, err_log = inject_errors(portfolio, month, pif_ids)
        subserv_path = os.path.join(SCRIPT_DIR, f"MSR_Tape_{label}_SUBSERVICER.xlsx")
        write_subservicer_tape(dirty, err_log, month, subserv_path)
        run_validation(prior_tape_path, prior_sheet, subserv_path, pif_path, na_path)
        # Parse validation result counts for dashboard
        val_md = os.path.join(SCRIPT_DIR, f"Validation_{label}_SUBSERVICER.md")
        hs, yl = parse_validation_counts(val_md)
        print(f"    Errors: {n_err} injected  |  Caught: {hs or '?'} hard stops, {yl or '?'} yellow lights")
    else:
        # Clean month — run validator against clean tape
        run_validation(prior_tape_path, prior_sheet, tape_path, pif_path, na_path)
        hs, yl = 0, 0
        print(f"    Clean submission — validation run as baseline")

    # 9 — Track dashboard metrics
    master_data.append({
        "month":         month,
        "label":         label,
        "n_loans":       len(portfolio),
        "total_upb":     total_upb,
        "n_pif":         len(pifs),
        "n_new_adds":    N_NEW_ADDS[month],
        "dq_count":      dq_count,
        "n_errors":      n_err,
        "hard_stops":    hs,
        "yellow_lights": yl,
    })

    return portfolio

def main():
    parser = argparse.ArgumentParser(description="MSR Year Simulation — Feb–Dec 2026")
    grp = parser.add_mutually_exclusive_group()
    grp.add_argument("--from", dest="from_month", type=int, metavar="M",
                     help="Re-run from month M through Dec (reads M-1 tape from disk)")
    grp.add_argument("--month", dest="only_month", type=int, metavar="M",
                     help="Re-run only month M (reads M-1 tape from disk)")
    args = parser.parse_args()

    # Determine which months to run and where to start
    if args.only_month:
        months_to_run = [args.only_month]
    elif args.from_month:
        months_to_run = list(range(args.from_month, 13))
    else:
        months_to_run = list(range(2, 13))

    start_month = months_to_run[0]

    print("=" * 60)
    print(f"  MSR YEAR SIMULATION — 2026")
    print(f"  Months: {', '.join(MONTH_NAME[m] for m in months_to_run)}")
    print("=" * 60)

    # Load starting portfolio
    if start_month == 2:
        combined = os.path.join(SCRIPT_DIR, "MSR_Sample_Tape_Dec2025_Jan2026.xlsx")
        print(f"\nLoading Jan 2026 portfolio from {os.path.basename(combined)}...")
        portfolio      = load_tape_from_xlsx(combined, sheet_name="Jan 2026")
        prior_tape_path = combined
        prior_sheet     = "Jan 2026"
    else:
        prior_m = start_month - 1
        prior_file = os.path.join(SCRIPT_DIR, f"MSR_Tape_{month_label(prior_m)}.xlsx")
        if not os.path.exists(prior_file):
            print(f"ERROR: Cannot find prior tape: {prior_file}")
            print(f"  Run months sequentially from Feb, or run --from 2 to start fresh.")
            sys.exit(1)
        print(f"\nLoading {MONTH_NAME[prior_m]} 2026 portfolio from {os.path.basename(prior_file)}...")
        portfolio       = load_tape_from_xlsx(prior_file)
        prior_tape_path = prior_file
        prior_sheet     = None

    print(f"  Starting portfolio: {len(portfolio):,} loans")

    # Load Jan portfolio separately for dashboard baseline (only if running from Feb)
    if start_month == 2:
        jan_baseline = portfolio
    else:
        combined = os.path.join(SCRIPT_DIR, "MSR_Sample_Tape_Dec2025_Jan2026.xlsx")
        jan_baseline = (load_tape_from_xlsx(combined, sheet_name="Jan 2026")
                        if os.path.exists(combined) else portfolio)

    master_data = []

    for month in months_to_run:
        portfolio = run_month(
            month, portfolio, master_data,
            prior_tape_path, prior_sheet
        )
        prior_tape_path = os.path.join(SCRIPT_DIR, f"MSR_Tape_{month_label(month)}.xlsx")
        prior_sheet     = None   # standalone files from here on

    # Write master dashboard
    write_master_dashboard(master_data, jan_baseline)

    print("\n" + "=" * 60)
    print("  SIMULATION COMPLETE")
    print(f"  Final portfolio: {len(portfolio):,} loans")
    print(f"  Final UPB:       ${sum(ln['upb'] for ln in portfolio):>14,.0f}")
    print("=" * 60)

if __name__ == "__main__":
    main()
