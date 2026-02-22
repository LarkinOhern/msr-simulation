#!/usr/bin/env python3
"""
MSR Monthly Reconciliation Automation
======================================
Reads two monthly MSR tape Excel files and automatically produces a
reconciliation report documenting all changes between periods.

Usage:
    python recon_automation.py <tape_month1.xlsx> <tape_month2.xlsx> [--output-dir DIR]

    OR (auto-discover tapes in a folder):
    python recon_automation.py --folder ./msr_simulation

Outputs:
    - Recon_Report_<M1>_to_<M2>.md     (human-readable markdown report)
    - Recon_Summary_<M1>_to_<M2>.xlsx  (structured Excel summary)
"""

import sys, os, re, argparse
_LOAN_ID_RE = re.compile(r"^MSR\d+$")
from datetime import date, datetime
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Expected column positions in the MSR tape (1-indexed) ───────────────────
# Header row is row 3 (row 1=disclaimer, row 2=title), data starts row 4
# Columns: Loan ID, Loan Type, Purpose, Investor, Orig Date, Orig Bal,
#          UPB, Rate, Net Serv Fee, Rem Term, Maturity, P&I, Escrow, Total Pmt, Status, NDD
COL = {
    "loan_id":    1,
    "loan_type":  2,
    "purpose":    3,
    "investor":   4,
    "orig_date":  5,
    "orig_bal":   6,
    "upb":        7,
    "rate":       8,
    "nsf":        9,
    "rem_term":   10,
    "maturity":   11,
    "pi":         12,
    "escrow":     13,
    "total_pmt":  14,
    "status":     15,
    "ndd":        16,
}

HEADER_ROW = 3   # column header row (after disclaimer + title)
DATA_START  = 4  # first data row

# ── Helpers ──────────────────────────────────────────────────────────────────
def _fill(hex6): return PatternFill("solid", fgColor=hex6)
def _font(bold=False, color="000000", size=9):
    return Font(name="Arial", bold=bold, color=color, size=size)
def _border():
    s = Side(style="thin", color="B8CCE4")
    return Border(left=s, right=s, top=s, bottom=s)

THIN = _border()

def _cell(ws, row, col, val, fill=None, fmt=None, bold=False,
          color="000000", align="left"):
    c = ws.cell(row=row, column=col, value=val)
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    c.font = _font(bold=bold, color=color, size=9)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = THIN
    return c

def _hcell(ws, row, col, val, fill=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill or _fill("1F4E79")
    c.font = _font(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = THIN
    return c

CURR = '#,##0.00'
NUM0 = '#,##0'
PCT3 = '0.000%'
PCT1 = '0.0%'
DFMT = 'MM/DD/YYYY'

# ── Load tape ─────────────────────────────────────────────────────────────────
def load_tape(filepath, sheet_hint=None):
    """
    Load an MSR tape from an Excel file.
    Returns dict: {loan_id -> loan_dict}
    Auto-detects the correct sheet (first sheet with MSR loan IDs).
    Handles tapes with disclaimer row (data starts at row 4).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Pick sheet
    target_sheet = None
    if sheet_hint and sheet_hint in wb.sheetnames:
        target_sheet = wb[sheet_hint]
    else:
        for name in wb.sheetnames:
            ws = wb[name]
            if any(kw in name.lower() for kw in ["summary","recon","portfolio"]):
                continue
            # Check rows 3-5 for a valid loan ID (MSR followed by digits)
            for try_row in range(3, 8):
                val = ws.cell(row=try_row, column=COL["loan_id"]).value
                if val and _LOAN_ID_RE.match(str(val).strip()):
                    target_sheet = ws
                    break
            if target_sheet:
                break
        if target_sheet is None:
            target_sheet = wb.active

    ws = target_sheet

    # Detect actual data start row (first row with a valid MSR\d+ loan ID)
    actual_data_start = DATA_START
    for try_row in range(1, 12):
        val = ws.cell(row=try_row, column=COL["loan_id"]).value
        if val and _LOAN_ID_RE.match(str(val).strip()):
            actual_data_start = try_row
            break

    loans = {}
    row = actual_data_start
    while True:
        loan_id = ws.cell(row=row, column=COL["loan_id"]).value
        if loan_id is None:
            break
        loan_id = str(loan_id).strip()
        if not loan_id or loan_id.upper() in ("TOTALS","TOTALS / AVERAGES"):
            break

        def v(col_name):
            return ws.cell(row=row, column=COL[col_name]).value

        upb = v("upb")
        try:    upb = float(upb) if upb is not None else None
        except: upb = None

        rate = v("rate")
        try:    rate = float(rate) if rate is not None else None
        except: rate = None

        nsf = v("nsf")
        try:    nsf = float(nsf) if nsf is not None else None
        except: nsf = None

        pi = v("pi")
        try:    pi = float(pi) if pi is not None else None
        except: pi = None

        orig_bal = v("orig_bal")
        try:    orig_bal = float(orig_bal) if orig_bal is not None else None
        except: orig_bal = None

        loans[loan_id] = {
            "loan_id":   loan_id,
            "loan_type": v("loan_type"),
            "investor":  v("investor"),
            "upb":       upb,
            "rate":      rate,
            "nsf":       nsf,
            "pi":        pi,
            "orig_bal":  orig_bal,
            "status":    v("status"),
            "rem_term":  v("rem_term"),
            "ndd":       v("ndd"),
        }
        row += 1

    print(f"    Loaded {len(loans):,} loans from '{ws.title}' in {os.path.basename(filepath)}")
    return loans

# ── Core reconciliation ───────────────────────────────────────────────────────
def reconcile(loans_m1: dict, loans_m2: dict, label_m1: str, label_m2: str) -> dict:
    """
    Compare two monthly loan dicts and return a structured reconciliation.
    """
    ids_m1 = set(loans_m1.keys())
    ids_m2 = set(loans_m2.keys())

    new_add_ids  = ids_m2 - ids_m1
    pif_ids      = ids_m1 - ids_m2
    continuing   = ids_m1 & ids_m2

    # UPB totals
    upb_m1   = sum(ln["upb"] or 0 for ln in loans_m1.values())
    upb_m2   = sum(ln["upb"] or 0 for ln in loans_m2.values())
    upb_new  = sum(loans_m2[i]["upb"] or 0 for i in new_add_ids)
    upb_pif  = sum(loans_m1[i]["upb"] or 0 for i in pif_ids)

    # UPB changes for continuing loans
    upb_changes = {}
    for lid in continuing:
        upb1 = loans_m1[lid]["upb"] or 0
        upb2 = loans_m2[lid]["upb"] or 0
        upb_changes[lid] = upb2 - upb1

    upb_delta_continuing = sum(upb_changes.values())
    sched_amort_proxy    = -sum(v for v in upb_changes.values() if v < 0)
    cap_increases        = sum(v for v in upb_changes.values() if v > 0)

    # Bridge check
    bridge   = upb_m1 + upb_new - upb_pif + upb_delta_continuing
    variance = bridge - upb_m2

    # Status changes for continuing loans
    status_changes = []
    for lid in continuing:
        s1 = loans_m1[lid]["status"] or "Current"
        s2 = loans_m2[lid]["status"] or "Current"
        if s1 != s2:
            status_changes.append({
                "loan_id":     lid,
                "from_status": s1,
                "to_status":   s2,
                "upb":         loans_m2[lid]["upb"],
            })

    # DQ bucket counts
    def dq_buckets(loans):
        buckets = {"Current": 0, "30 DPD": 0, "60 DPD": 0, "90+ DPD": 0, "Other": 0}
        for ln in loans.values():
            s = ln["status"] or "Current"
            if s in buckets:
                buckets[s] += 1
            else:
                buckets["Other"] += 1
        return buckets

    b1 = dq_buckets(loans_m1)
    b2 = dq_buckets(loans_m2)

    # Investor mix
    def inv_summary(loans):
        d = defaultdict(lambda: {"count": 0, "upb": 0.0})
        for ln in loans.values():
            inv = ln["investor"] or "Unknown"
            d[inv]["count"] += 1
            d[inv]["upb"]   += ln["upb"] or 0
        return dict(d)

    inv_m1 = inv_summary(loans_m1)
    inv_m2 = inv_summary(loans_m2)

    # Curtailment candidates
    avg_drop = sched_amort_proxy / max(len(continuing), 1)
    curtailments = []
    for lid in continuing:
        delta = upb_changes[lid]
        if delta < -2.5 * avg_drop and avg_drop > 0:
            curtailments.append({
                "loan_id":   lid,
                "upb_m1":    loans_m1[lid]["upb"],
                "upb_m2":    loans_m2[lid]["upb"],
                "extra_pay": abs(delta) - avg_drop,
            })

    return {
        "label_m1": label_m1,
        "label_m2": label_m2,
        "count_m1": len(loans_m1),
        "count_m2": len(loans_m2),
        "new_add_ids":  sorted(new_add_ids),
        "pif_ids":      sorted(pif_ids),
        "continuing":   sorted(continuing),
        "upb_m1":       upb_m1,
        "upb_m2":       upb_m2,
        "upb_new":      upb_new,
        "upb_pif":      upb_pif,
        "upb_delta_continuing": upb_delta_continuing,
        "sched_amort_proxy":    sched_amort_proxy,
        "cap_increases":        cap_increases,
        "bridge":       bridge,
        "variance":     variance,
        "status_changes": status_changes,
        "dq_m1": b1,
        "dq_m2": b2,
        "upb_changes": upb_changes,
        "curtailments": curtailments,
        "inv_m1": inv_m1,
        "inv_m2": inv_m2,
        "loans_m1": loans_m1,
        "loans_m2": loans_m2,
    }

# ── Markdown report ───────────────────────────────────────────────────────────
def write_markdown(r: dict, out_path: str):
    m1, m2 = r["label_m1"], r["label_m2"]
    now = datetime.now().strftime("%B %d, %Y")
    ties = "[OK] TIES" if abs(r["variance"]) < 1.0 else f"[!]️  VARIANCE ${r['variance']:,.2f}"

    lines = [
        f"# MSR Monthly Reconciliation Report",
        f"**Period:** {m1} -> {m2}",
        f"**Generated:** {now}",
        f"",
        "> **SIMULATED DATA** — All loan information is synthetic and generated for testing purposes only.",
        "",
        "---",
        "",
        "## A. Loan Count Reconciliation",
        "",
        f"| | Count |",
        f"|---|---:|",
        f"| **Beginning Count ({m1})** | {r['count_m1']:,} |",
        f"| + New Adds | +{len(r['new_add_ids']):,} |",
        f"| - Paid in Full / Removed | -{len(r['pif_ids']):,} |",
        f"| **Ending Count ({m2})** | **{r['count_m2']:,}** |",
        f"| Bridge Result | {r['count_m1'] + len(r['new_add_ids']) - len(r['pif_ids']):,} |",
        f"| **Count Variance** | **{'[OK] 0' if r['count_m1'] + len(r['new_add_ids']) - len(r['pif_ids']) == r['count_m2'] else '[!]️  MISMATCH'}** |",
        "",
        "---",
        "",
        "## B. UPB Reconciliation",
        "",
        f"| Component | Amount |",
        f"|---|---:|",
        f"| Beginning UPB ({m1}) | ${r['upb_m1']:>16,.2f} |",
        f"| + New Adds UPB | ${r['upb_new']:>16,.2f} |",
        f"| - PIF UPB Removed | $({r['upb_pif']:>15,.2f}) |",
        f"| +/- Continuing Loan UPB Change | ${r['upb_delta_continuing']:>+16,.2f} |",
        f"|   (of which: scheduled amort) | $({r['sched_amort_proxy']:>15,.2f}) |",
        f"|   (of which: capitalizations) | ${r['cap_increases']:>16,.2f} |",
        f"| **Bridge Result** | **${r['bridge']:>16,.2f}** |",
        f"| Actual Ending UPB ({m2}) | ${r['upb_m2']:>16,.2f} |",
        f"| **UPB Variance** | **{ties}  ${r['variance']:>+,.2f}** |",
        "",
        "---",
        "",
        "## C. Delinquency Migration",
        "",
        f"| Bucket | {m1} | {m2} | Change | % of Portfolio ({m2}) |",
        f"|---|---:|---:|---:|---:|",
    ]

    for bucket in ["Current", "30 DPD", "60 DPD", "90+ DPD"]:
        d1 = r["dq_m1"].get(bucket, 0)
        d2 = r["dq_m2"].get(bucket, 0)
        chg = d2 - d1
        pct = d2 / r["count_m2"] if r["count_m2"] else 0
        chg_str = f"+{chg}" if chg > 0 else str(chg)
        lines.append(
            f"| {bucket} | {d1:,} | {d2:,} | {chg_str} | {pct:.2%} |"
        )

    lines += [
        "",
        "### DQ Status Changes (Loan-Level)",
        "",
    ]
    if r["status_changes"]:
        lines.append("| Loan ID | From | To | UPB ($) |")
        lines.append("|---|---|---|---:|")
        for sc in r["status_changes"]:
            lines.append(
                f"| {sc['loan_id']} | "
                f"{sc['from_status']} | {sc['to_status']} | "
                f"${sc['upb']:,.2f} |"
            )
    else:
        lines.append("_No delinquency status changes detected._")

    lines += [
        "",
        "---",
        "",
        "## D. New Adds Summary",
        "",
        f"**Total New Adds: {len(r['new_add_ids']):,} loans  |  Total UPB: ${r['upb_new']:,.2f}**",
        "",
        "| Loan ID | Investor | UPB ($) |",
        "|---|---|---:|",
    ]
    for lid in r["new_add_ids"][:20]:
        ln = r["loans_m2"][lid]
        lines.append(f"| {lid} | {ln['investor']} | ${ln['upb']:,.2f} |")
    if len(r["new_add_ids"]) > 20:
        lines.append(f"| _... and {len(r['new_add_ids'])-20:,} more_ | | |")

    lines += [
        "",
        "---",
        "",
        "## E. Paid in Full Summary",
        "",
        f"**Total PIF: {len(r['pif_ids']):,} loans  |  Total UPB Removed: ${r['upb_pif']:,.2f}**",
        "",
        "| Loan ID | Investor | Loan Type | Final UPB ($) | Rate |",
        "|---|---|---|---:|---:|",
    ]
    for lid in r["pif_ids"]:
        ln = r["loans_m1"][lid]
        lines.append(
            f"| {lid} | {ln['investor']} | {ln['loan_type']} | "
            f"${ln['upb']:,.2f} | {ln['rate']:.3%} |"
        )

    lines += [
        "",
        "---",
        "",
        "## F. Notable UPB Changes (Curtailments / Large Paydowns)",
        "",
    ]
    if r["curtailments"]:
        lines.append("| Loan ID | Prior UPB ($) | Current UPB ($) | Estimated Extra Principal ($) |")
        lines.append("|---|---:|---:|---:|")
        for c in sorted(r["curtailments"], key=lambda x: x["extra_pay"], reverse=True):
            lines.append(
                f"| {c['loan_id']} | "
                f"${c['upb_m1']:,.2f} | ${c['upb_m2']:,.2f} | "
                f"${c['extra_pay']:,.2f} |"
            )
    else:
        lines.append("_No significant curtailments detected._")

    lines += [
        "",
        "---",
        "",
        "## G. Investor Mix Comparison",
        "",
        f"| Investor | {m1} Count | {m1} UPB ($) | {m2} Count | {m2} UPB ($) | Count Delta |",
        "|---|---:|---:|---:|---:|---:|",
    ]
    all_investors = sorted(set(list(r["inv_m1"].keys()) + list(r["inv_m2"].keys())))
    for inv in all_investors:
        d1 = r["inv_m1"].get(inv, {"count":0,"upb":0})
        d2 = r["inv_m2"].get(inv, {"count":0,"upb":0})
        chg = d2["count"] - d1["count"]
        chg_str = f"+{chg}" if chg >= 0 else str(chg)
        lines.append(
            f"| {inv} | {d1['count']:,} | ${d1['upb']:,.0f} | "
            f"{d2['count']:,} | ${d2['upb']:,.0f} | {chg_str} |"
        )

    lines += [
        "",
        "---",
        "",
        f"_Report generated by MSR Recon Automation — {now}_",
        "",
    ]

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Markdown report: {out_path}")

# ── Excel summary ─────────────────────────────────────────────────────────────
def write_excel_summary(r: dict, out_path: str):
    wb = Workbook()
    m1, m2 = r["label_m1"], r["label_m2"]
    now = datetime.now().strftime("%Y-%m-%d")

    DISCLAIMER = "SIMULATED DATA — All loan information is synthetic and generated for testing purposes only."

    # ── Sheet 1: Summary Dashboard ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Recon Summary"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    NAVY = _fill("1F4E79"); BLUE = _fill("2E75B6"); TOTAL = _fill("BDD7EE")
    GRNH = _fill("375623"); REDH = _fill("7B2C2C"); LTGRN = _fill("E2EFDA")
    LTRED = _fill("FCE4D6"); GREY = _fill("F2F2F2")
    VGRN = _fill("A9D18E"); VRED = _fill("F4CCCC")
    TIES_GREEN = _fill("C6EFCE")
    DISC_FILL  = _fill("FFF2CC")

    row = 1
    # Disclaimer row
    ws.merge_cells(f"A{row}:D{row}")
    dc = ws.cell(row=row, column=1, value=DISCLAIMER)
    dc.fill = DISC_FILL
    dc.font = _font(bold=True, color="7B3F00", size=9)
    dc.alignment = Alignment(horizontal="center", vertical="center")
    dc.border = THIN
    ws.row_dimensions[row].height = 16
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    t = ws.cell(row=row, column=1,
        value=f"MSR RECONCILIATION REPORT  |  {m1} -> {m2}  |  Generated {now}")
    t.fill = NAVY; t.font = _font(bold=True, color="FFFFFF", size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28; row += 1; row += 1

    def section(ws, row, title, fill):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.fill = fill; c.font = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        return row + 1

    # Count bridge
    row = section(ws, row, "A.  LOAN COUNT RECONCILIATION", BLUE)
    for col, h in enumerate(["","Beginning","Ending","Change"],1):
        _hcell(ws, row, col, h)
    row += 1
    data = [
        ("Loan Count", r["count_m1"], r["count_m2"], r["count_m2"]-r["count_m1"], TOTAL),
        ("  + New Adds", "", len(r["new_add_ids"]),  "", None),
        ("  - Paid in Full", "", -len(r["pif_ids"]), "", None),
        ("Bridge (should = Ending)", r["count_m1"],
         r["count_m1"]+len(r["new_add_ids"])-len(r["pif_ids"]),
         None, TOTAL),
    ]
    for label, v1, v2, chg, fill in data:
        _cell(ws, row, 1, label, fill, bold=(fill==TOTAL))
        _cell(ws, row, 2, v1, fill, NUM0 if isinstance(v1,int) else None,
              bold=(fill==TOTAL), align="right")
        _cell(ws, row, 3, v2, fill, NUM0 if isinstance(v2,int) else None,
              bold=(fill==TOTAL), align="right")
        if chg is not None:
            cc = _cell(ws, row, 4, chg, fill, "+#,##0;(#,##0);-",
                       bold=(fill==TOTAL), align="right")
            if isinstance(chg,int) and chg > 0:  cc.font = _font(bold=(fill==TOTAL), color="006400")
            if isinstance(chg,int) and chg < 0:  cc.font = _font(bold=(fill==TOTAL), color="C00000")
        else:
            ties_val = "[OK] TIES" if r["count_m1"]+len(r["new_add_ids"])-len(r["pif_ids"])==r["count_m2"] else "[!]️ ERROR"
            _cell(ws, row, 4, ties_val, TIES_GREEN if "[OK]" in ties_val else VRED,
                  bold=True, align="center")
        row += 1
    row += 1

    # UPB bridge
    row = section(ws, row, "B.  UPB RECONCILIATION", BLUE)
    for col, h in enumerate(["Component","Amount ($)","",""],1):
        _hcell(ws, row, col, h)
    row += 1
    upb_rows = [
        (f"Beginning UPB ({m1})",                 r["upb_m1"],    None),
        ("  + New Adds UPB",                      r["upb_new"],   None),
        (f"  - PIF UPB Removed",                 -r["upb_pif"],   None),
        ("  +/- Continuing Loan Changes",         r["upb_delta_continuing"], None),
        ("      (Scheduled Amortization)",        -r["sched_amort_proxy"],   None),
        ("      (Capitalizations / Increases)",    r["cap_increases"],        None),
        ("Bridge Result",                          r["bridge"],    "bridge"),
        (f"Actual Ending UPB ({m2})",             r["upb_m2"],    "actual"),
        ("Variance (Bridge - Actual)",            r["variance"],   "variance"),
    ]
    for label, amt, tag in upb_rows:
        is_bold = tag in ("bridge","actual","variance")
        fill = TOTAL if tag in ("bridge","actual") else \
               (TIES_GREEN if (tag=="variance" and abs(r["variance"])<1) else \
                (VRED if tag=="variance" else None))
        _cell(ws, row, 1, label, fill, bold=is_bold)
        c = _cell(ws, row, 2, amt, fill, CURR, bold=is_bold, align="right")
        if tag == "variance":
            c.font = _font(bold=True, color="006400" if abs(amt)<1 else "C00000")
        ws.cell(row=row, column=3).border = THIN
        ws.cell(row=row, column=4).border = THIN
        if tag == "variance":
            ws.cell(row=row, column=3).value = "[OK] TIES" if abs(amt)<1 else "[!]️ CHECK"
            ws.cell(row=row, column=3).font  = _font(bold=True,
                color="006400" if abs(amt)<1 else "C00000")
        row += 1
    row += 1

    # DQ migration
    row = section(ws, row, "C.  DELINQUENCY MIGRATION", REDH)
    for col, h in enumerate(["Bucket", m1, m2, "Change"],1):
        _hcell(ws, row, col, h, fill=REDH)
    row += 1
    for bucket in ["Current","30 DPD","60 DPD","90+ DPD","TOTAL"]:
        if bucket == "TOTAL":
            d1 = sum(r["dq_m1"].values())
            d2 = sum(r["dq_m2"].values())
            fill = TOTAL
        else:
            d1 = r["dq_m1"].get(bucket,0)
            d2 = r["dq_m2"].get(bucket,0)
            fill = LTRED if "DPD" in bucket else None
        chg = d2 - d1
        _cell(ws, row, 1, bucket, fill, bold=(fill==TOTAL))
        _cell(ws, row, 2, d1, fill, NUM0, bold=(fill==TOTAL), align="right")
        _cell(ws, row, 3, d2, fill, NUM0, bold=(fill==TOTAL), align="right")
        cc = _cell(ws, row, 4, chg, fill, "+#,##0;(#,##0);-",
                   bold=(fill==TOTAL), align="right")
        if chg > 0 and "DPD" in bucket: cc.font = _font(bold=False, color="C00000")
        if chg < 0 and "DPD" in bucket: cc.font = _font(bold=False, color="006400")
        row += 1
    row += 1

    # ── Sheet 2: New Adds Detail ───────────────────────────────────────────────
    ws2 = wb.create_sheet("New Adds")
    ws2.column_dimensions["A"].width = 13
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 13
    ws2.column_dimensions["D"].width = 16
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:D1")
    t = ws2.cell(row=1, column=1,
        value=f"NEW ADDS DETAIL  |  {m2}  |  {len(r['new_add_ids']):,} Loans")
    t.fill = GRNH; t.font = _font(bold=True, color="FFFFFF", size=11)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24

    for col, h in enumerate(["Loan ID","Investor","Loan Type","UPB ($)"],1):
        _hcell(ws2, 2, col, h, fill=GRNH)

    for r2, lid in enumerate(r["new_add_ids"], 3):
        ln = r["loans_m2"][lid]
        fill = LTGRN if r2 % 2 == 0 else None
        _cell(ws2, r2, 1, lid,             fill, align="center")
        _cell(ws2, r2, 2, ln["investor"],  fill, align="center")
        _cell(ws2, r2, 3, ln["loan_type"], fill, align="center")
        _cell(ws2, r2, 4, ln["upb"],       fill, CURR, align="right")
    tr2 = len(r["new_add_ids"]) + 3
    _cell(ws2, tr2, 1, "TOTAL", VGRN, bold=True)
    c = ws2.cell(row=tr2, column=4)
    c.value = f"=SUM(D3:D{tr2-1})"
    c.number_format = CURR; c.font = _font(bold=True); c.fill = VGRN
    c.border = THIN; c.alignment = Alignment(horizontal="right")
    for col in [2, 3]:
        ws2.cell(row=tr2, column=col).fill = VGRN
        ws2.cell(row=tr2, column=col).border = THIN

    # ── Sheet 3: PIF Detail ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Paid in Full")
    ws3.column_dimensions["A"].width = 13
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 13
    ws3.column_dimensions["D"].width = 16
    ws3.column_dimensions["E"].width = 10
    ws3.freeze_panes = "A2"

    ws3.merge_cells("A1:E1")
    t = ws3.cell(row=1, column=1,
        value=f"PAID IN FULL DETAIL  |  {m2}  |  {len(r['pif_ids']):,} Loans")
    t.fill = REDH; t.font = _font(bold=True, color="FFFFFF", size=11)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 24

    for col, h in enumerate(["Loan ID","Investor","Loan Type","Final UPB ($)","Rate"],1):
        _hcell(ws3, 2, col, h, fill=REDH)

    for r3, lid in enumerate(r["pif_ids"], 3):
        ln = r["loans_m1"][lid]
        fill = LTRED if r3 % 2 == 0 else None
        _cell(ws3, r3, 1, lid,             fill, align="center")
        _cell(ws3, r3, 2, ln["investor"],  fill, align="center")
        _cell(ws3, r3, 3, ln["loan_type"], fill, align="center")
        _cell(ws3, r3, 4, ln["upb"],       fill, CURR, align="right")
        _cell(ws3, r3, 5, ln["rate"],      fill, PCT3, align="right")
    tr3 = len(r["pif_ids"]) + 3
    _cell(ws3, tr3, 1, "TOTAL", VRED, bold=True)
    c = ws3.cell(row=tr3, column=4)
    c.value = f"=SUM(D3:D{tr3-1})"
    c.number_format = CURR; c.font = _font(bold=True); c.fill = VRED
    c.border = THIN; c.alignment = Alignment(horizontal="right")
    for col in [2, 3, 5]:
        ws3.cell(row=tr3, column=col).fill = VRED
        ws3.cell(row=tr3, column=col).border = THIN

    # ── Sheet 4: DQ Status Changes ────────────────────────────────────────────
    ws4 = wb.create_sheet("DQ Changes")
    ws4.column_dimensions["A"].width = 13
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 14
    ws4.column_dimensions["D"].width = 16
    ws4.freeze_panes = "A2"

    ws4.merge_cells("A1:D1")
    t = ws4.cell(row=1, column=1,
        value=f"DELINQUENCY STATUS CHANGES  |  {m1} -> {m2}")
    t.fill = REDH; t.font = _font(bold=True, color="FFFFFF", size=11)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 24

    for col, h in enumerate(["Loan ID",f"Status ({m1})",f"Status ({m2})","UPB ($)"],1):
        _hcell(ws4, 2, col, h, fill=REDH)

    if r["status_changes"]:
        for r4, sc in enumerate(r["status_changes"], 3):
            fill = LTRED if r4 % 2 == 0 else None
            _cell(ws4, r4, 1, sc["loan_id"],      fill, align="center")
            c1 = _cell(ws4, r4, 2, sc["from_status"], fill, align="center")
            c2 = _cell(ws4, r4, 3, sc["to_status"],   fill, align="center")
            is_worsening = ("DPD" in (sc["to_status"] or "") and
                            sc["from_status"] in ("Current","30 DPD","60 DPD"))
            c2.font = _font(bold=True, color="C00000" if is_worsening else "006400")
            _cell(ws4, r4, 4, sc["upb"], fill, CURR, align="right")
    else:
        ws4.cell(row=3, column=1).value = "No delinquency status changes detected."
        ws4.cell(row=3, column=1).font  = _font(italic=True)

    wb.save(out_path)
    print(f"  Excel summary:   {out_path}")

# ── Main ──────────────────────────────────────────────────────────────────────
def detect_tape_files(folder):
    """Auto-detect monthly tape files in a folder (looks for MSR_Sample_Tape_*.xlsx)."""
    files = [f for f in os.listdir(folder)
             if f.startswith("MSR_Sample_Tape") and f.endswith(".xlsx")]
    files.sort()
    return [os.path.join(folder, f) for f in files]

def parse_label(filepath):
    """Extract a short month label from filename or return basename."""
    base = os.path.basename(filepath).replace("MSR_Sample_Tape_","").replace(".xlsx","")
    parts = base.replace("_"," ").strip()
    return parts if parts else base

def main():
    parser = argparse.ArgumentParser(
        description="MSR Monthly Reconciliation Automation"
    )
    parser.add_argument("files", nargs="*",
        help="Two MSR tape Excel files to reconcile (month1.xlsx month2.xlsx)")
    parser.add_argument("--folder", "-f",
        help="Folder to auto-discover tape files (picks first two alphabetically)")
    parser.add_argument("--output-dir", "-o", default=None,
        help="Output directory for reports (default: same as first input file)")
    parser.add_argument("--sheet-m1", default=None,
        help="Sheet name to use in month1 file")
    parser.add_argument("--sheet-m2", default=None,
        help="Sheet name to use in month2 file")
    args = parser.parse_args()

    # Resolve input files
    if args.folder:
        tapes = detect_tape_files(args.folder)
        if len(tapes) < 2:
            print(f"ERROR: Need at least 2 tape files in {args.folder}, found {len(tapes)}")
            sys.exit(1)
        file_m1, file_m2 = tapes[0], tapes[1]
        print(f"Auto-detected tapes:\n  M1: {file_m1}\n  M2: {file_m2}")
    elif len(args.files) == 2:
        file_m1, file_m2 = args.files
    elif len(args.files) == 1:
        file_m1 = file_m2 = args.files[0]
    else:
        print("ERROR: Provide two tape files, or use --folder to auto-detect.")
        print("Usage: python recon_automation.py file1.xlsx file2.xlsx")
        sys.exit(1)

    out_dir = args.output_dir or os.path.dirname(os.path.abspath(file_m1))

    # For the combined tape (same file, two sheets)
    if file_m1 == file_m2:
        wb_tmp = openpyxl.load_workbook(file_m1, data_only=True)
        data_sheets = [s for s in wb_tmp.sheetnames
                       if not any(kw in s.lower() for kw in ["summary","recon","portfolio"])]
        if len(data_sheets) < 2:
            print("ERROR: Need at least 2 data sheets in the combined tape.")
            sys.exit(1)
        sheet_m1 = args.sheet_m1 or data_sheets[0]
        sheet_m2 = args.sheet_m2 or data_sheets[1]
        label_m1 = sheet_m1
        label_m2 = sheet_m2
    else:
        sheet_m1 = args.sheet_m1
        sheet_m2 = args.sheet_m2
        label_m1 = parse_label(file_m1)
        label_m2 = parse_label(file_m2)

    print(f"\nLoading tapes...")
    loans_m1 = load_tape(file_m1, sheet_hint=sheet_m1)
    loans_m2 = load_tape(file_m2, sheet_hint=sheet_m2)

    print(f"\nReconciling {label_m1} -> {label_m2}...")
    result = reconcile(loans_m1, loans_m2, label_m1, label_m2)

    slug = f"{label_m1.replace(' ','_')}_to_{label_m2.replace(' ','_')}"
    slug = re.sub(r"[^\w_-]", "", slug)

    md_path   = os.path.join(out_dir, f"Recon_Report_{slug}.md")
    xlsx_path = os.path.join(out_dir, f"Recon_Summary_{slug}.xlsx")

    print(f"\nWriting reports...")
    write_markdown(result, md_path)
    write_excel_summary(result, xlsx_path)

    print(f"\n{'='*60}")
    print(f"RECONCILIATION COMPLETE")
    print(f"{'='*60}")
    print(f"  {label_m1}: {result['count_m1']:,} loans  |  UPB: ${result['upb_m1']:,.2f}")
    print(f"  New Adds:    +{len(result['new_add_ids']):,} loans  |  UPB: +${result['upb_new']:,.2f}")
    print(f"  PIF:         -{len(result['pif_ids']):,} loans  |  UPB: -${result['upb_pif']:,.2f}")
    print(f"  {label_m2}: {result['count_m2']:,} loans  |  UPB: ${result['upb_m2']:,.2f}")
    print(f"  UPB Bridge Variance: ${result['variance']:,.2f}  "
          f"{'[OK] TIES' if abs(result['variance'])<1 else '[!]️  CHECK'}")
    print(f"\nOutputs:")
    print(f"  {md_path}")
    print(f"  {xlsx_path}")

if __name__ == "__main__":
    main()
