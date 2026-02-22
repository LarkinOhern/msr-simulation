"""
build_content_artifacts.py
Generates two content Excel workbooks for sharing / LinkedIn:
  Content_SubservicerChecklist.xlsx  — error taxonomy and detection framework
  Content_BlindTestSummary.xlsx      — annotated blind test results
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Colours ──────────────────────────────────────────────────────────────────
NAVY        = PatternFill("solid", fgColor="1F4E79")
DARK_RED    = PatternFill("solid", fgColor="833C00")   # hard stop header
AMBER       = PatternFill("solid", fgColor="7B3F00")   # yellow-light header
DARK_GREEN  = PatternFill("solid", fgColor="375623")   # pass / cleared header
MID_BLUE    = PatternFill("solid", fgColor="2E75B6")   # section sub-header
LIGHT_RED   = PatternFill("solid", fgColor="FCE4D6")   # hard stop row
LIGHT_AMBER = PatternFill("solid", fgColor="FFF2CC")   # yellow row
LIGHT_GREEN = PatternFill("solid", fgColor="E2EFDA")   # cleared/pass row
LIGHT_BLUE  = PatternFill("solid", fgColor="DEEAF1")   # info row
TEAL_TOTAL  = PatternFill("solid", fgColor="BDD7EE")   # totals / summary
WHITE       = PatternFill("solid", fgColor="FFFFFF")
SIMDATA_YEL = PatternFill("solid", fgColor="FFFF00")   # sim-data banner

W  = Font(bold=True, color="FFFFFF")
WN = Font(color="FFFFFF")
B  = Font(bold=True, color="000000")
N  = Font(color="000000")
RED_FONT    = Font(bold=True, color="C0392B")
AMBER_FONT  = Font(bold=True, color="7B3F00")
GREEN_FONT  = Font(bold=True, color="375623")
NAVY_FONT   = Font(bold=True, color="1F4E79")
ITALIC      = Font(italic=True, color="555555")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)

THIN = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)
THICK_BOTTOM = Border(bottom=Side(style="medium"))

def hdr(ws, row, col, val, fill=NAVY, font=W, align=CENTER, border=THIN):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill; c.font = font; c.alignment = align; c.border = border

def cell(ws, row, col, val, fill=WHITE, font=N, align=LEFT, border=THIN, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill; c.font = font; c.alignment = align; c.border = border
    if fmt: c.number_format = fmt

def merge_hdr(ws, row, c1, c2, val, fill=NAVY, font=W, align=CENTER):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=val)
    c.fill = fill; c.font = font; c.alignment = align

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════════════════
#  WORKBOOK 1 — Subservicer Checklist
# ══════════════════════════════════════════════════════════════════════════════

def build_checklist(path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Error Taxonomy ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Error Taxonomy"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # Title
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20
    merge_hdr(ws, 1, 1, 7, "WHAT YOUR SUBSERVICER TAPE MIGHT BE HIDING", NAVY, W)
    ws.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=14)
    merge_hdr(ws, 2, 1, 7,
              "MSR Tape Error Taxonomy — Field-Level & Cross-Period Detection Framework",
              MID_BLUE, WN)
    ws.cell(2, 1).font = Font(italic=True, color="FFFFFF", size=10)

    # Sim-data banner
    ws.row_dimensions[3].height = 16
    merge_hdr(ws, 3, 1, 7,
              "SIMULATED DATA — All loan information is synthetic and generated for testing purposes only.",
              SIMDATA_YEL, Font(bold=True, color="000000", size=9), CENTER)

    # Column headers
    COLS = ["#", "Error Type", "Severity", "Detection Layer",
            "Field(s) Checked", "What It Looks Like", "Action Required"]
    for i, h in enumerate(COLS, 1):
        hdr(ws, 4, i, h, NAVY, W, CENTER)
    ws.row_dimensions[4].height = 22

    # ── Hard Stop rows ────────────────────────────────────────────────────────
    hs_data = [
        (1, "UPB with extra zero (×10 error)",    "HARD STOP", "Layer 1",
         "Current UPB ($)",
         "UPB submitted as $1,577,419 — original balance was $325,000",
         "Return tape; correct all UPB values"),
        (2, "UPB = $0 for an active loan",         "HARD STOP", "Layer 1",
         "Current UPB ($)",
         "Loan status = Current, UPB = $0.00; loan not marked Paid in Full",
         "Confirm loan is active; resubmit correct UPB"),
        (3, "UPB exceeds original balance",        "HARD STOP", "Layer 1",
         "Current UPB ($), Original Bal ($)",
         "Current UPB of $347,200 vs original balance of $310,000",
         "Validate amortization; return tape with corrected UPB"),
        (4, "Rate expressed as whole number",      "HARD STOP", "Layer 1",
         "Rate",
         "Rate = 6.50 instead of 0.0650 — decimal format required",
         "Reformat all rates as decimals; resubmit"),
        (5, "NSF expressed as whole basis points", "HARD STOP", "Layer 1",
         "Net Serv Fee",
         "NSF = 46.0 instead of 0.0046 — 10,000× off",
         "Reformat NSF fields; resubmit"),
        (6, "Duplicate loan ID",                   "HARD STOP", "Layer 1",
         "Loan ID",
         "Loan MSR100152 appears twice in the submission",
         "De-duplicate; verify which record is correct"),
        (7, "Loan missing — no PIF explanation",   "HARD STOP", "Layer 2",
         "Loan ID (cross-period)",
         "Prior month loan MSR100102 not in submission; not in PIF report",
         "Locate loan; confirm status; resubmit or provide PIF documentation"),
    ]

    r = 5
    for row_data in hs_data:
        ws.row_dimensions[r].height = 44
        num, etype, sev, layer, fields, example, action = row_data
        cell(ws, r, 1, num,    LIGHT_RED, Font(bold=True, color="C0392B"), CENTER)
        cell(ws, r, 2, etype,  LIGHT_RED, Font(bold=True, color="C0392B"), LEFT)
        cell(ws, r, 3, sev,    LIGHT_RED, Font(bold=True, color="C0392B"), CENTER)
        cell(ws, r, 4, layer,  LIGHT_RED, N, CENTER)
        cell(ws, r, 5, fields, LIGHT_RED, N, LEFT)
        cell(ws, r, 6, example,LIGHT_RED, N, LEFT)
        cell(ws, r, 7, action, LIGHT_RED, N, LEFT)
        r += 1

    # Section break
    ws.row_dimensions[r].height = 14
    merge_hdr(ws, r, 1, 7, "", AMBER, W)
    r += 1

    # ── Yellow Light rows ─────────────────────────────────────────────────────
    yl_data = [
        (8,  "NSF may be expressed as percent",   "YELLOW LIGHT", "Layer 1",
         "Net Serv Fee, Investor",
         "FNMA NSF = 0.2500 — looks like 25% instead of 0.25bps",
         "Confirm with subservicer; correct if misformatted"),
        (9,  "NSF outside investor-expected range","YELLOW LIGHT", "Layer 1",
         "Net Serv Fee, Investor",
         "GNMA NSF should be 19–69bps; submitted value is outside that band",
         "Review against investor guidelines; request correction if wrong"),
        (10, "Next due date in the past",          "YELLOW LIGHT", "Layer 1",
         "Next Due Date, Status",
         "Current-status loan; next due date = June 2025 (7 months past)",
         "Verify loan is not delinquent; confirm correct date"),
        (11, "Status bucket skip",                 "YELLOW LIGHT", "Layer 2",
         "Status (cross-period)",
         "Prior month: Current → This month: 90+ DPD (skipped 30 and 60 DPD)",
         "Confirm DQ history; validate status transition is real"),
        (12, "P&I inflated vs. prior month",       "YELLOW LIGHT", "Layer 2",
         "P&I ($) (cross-period)",
         "P&I increased from $1,432 to $1,719 (+20%) with no rate change",
         "Verify no modification or rate reset; request corrected payment"),
        (13, "Remaining term did not decrease",    "YELLOW LIGHT", "Layer 2",
         "Rem Term (cross-period)",
         "Rem term = 111 this month vs 111 prior month (should be 110)",
         "Confirm loan is paying as scheduled; verify no forbearance"),
        (14, "New add not in recon report",        "YELLOW LIGHT", "Layer 2",
         "Loan ID (cross-period)",
         "MSR300198 appears in submission but not in New Add recon report",
         "Confirm loan was legitimately boarded; provide boarding documentation"),
    ]

    for row_data in yl_data:
        ws.row_dimensions[r].height = 44
        num, etype, sev, layer, fields, example, action = row_data
        cell(ws, r, 1, num,    LIGHT_AMBER, Font(bold=True, color="7B3F00"), CENTER)
        cell(ws, r, 2, etype,  LIGHT_AMBER, Font(bold=True, color="7B3F00"), LEFT)
        cell(ws, r, 3, sev,    LIGHT_AMBER, Font(bold=True, color="7B3F00"), CENTER)
        cell(ws, r, 4, layer,  LIGHT_AMBER, N, CENTER)
        cell(ws, r, 5, fields, LIGHT_AMBER, N, LEFT)
        cell(ws, r, 6, example,LIGHT_AMBER, N, LEFT)
        cell(ws, r, 7, action, LIGHT_AMBER, N, LEFT)
        r += 1

    # Footer note
    r += 1
    ws.row_dimensions[r].height = 16
    merge_hdr(ws, r, 1, 7,
              "Validator auto-discovers PIF and New Add recon files — confirmed PIFs are cleared (informational). "
              "Only genuinely unexplained absences become hard stops.",
              LIGHT_BLUE, Font(italic=True, color="1F4E79", size=9), LEFT)

    set_col_widths(ws, [4, 32, 14, 16, 22, 46, 36])

    # ── Sheet 2: Quick Reference Checklist ───────────────────────────────────
    ws2 = wb.create_sheet("Quick Reference")
    ws2.sheet_view.showGridLines = False

    ws2.row_dimensions[1].height = 30
    ws2.row_dimensions[2].height = 18
    ws2.row_dimensions[3].height = 16
    merge_hdr(ws2, 1, 1, 4, "MSR TAPE SUBMISSION — QUICK REFERENCE CHECKLIST", NAVY, W)
    ws2.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=13)
    merge_hdr(ws2, 2, 1, 4,
              "Review before accepting any subservicer tape submission", MID_BLUE, WN)
    merge_hdr(ws2, 3, 1, 4,
              "SIMULATED DATA — Synthetic portfolio for illustration only",
              SIMDATA_YEL, Font(bold=True, color="000000", size=9))

    # Headers
    for i, h in enumerate(["Check", "What to Verify", "Flag If...", "Severity"], 1):
        hdr(ws2, 4, i, h, NAVY, W, CENTER)
    ws2.row_dimensions[4].height = 22

    checks = [
        # (category_header, None) or (check#, what, flag_if, severity, fill, font)
        ("LAYER 1 — STANDALONE FIELD CHECKS", None, None, None, MID_BLUE, W),
        (1,  "UPB — magnitude",
             "Current UPB > Original Balance (loan can't grow without mod)",
             "HARD STOP", LIGHT_RED, Font(bold=True, color="C0392B")),
        (2,  "UPB — active loan",
             "UPB = $0 and status ≠ Paid in Full",
             "HARD STOP", LIGHT_RED, Font(bold=True, color="C0392B")),
        (3,  "Rate — decimal format",
             "Rate > 1.0 (likely submitted as whole number: 6.50 vs 0.0650)",
             "HARD STOP", LIGHT_RED, Font(bold=True, color="C0392B")),
        (4,  "NSF — decimal format (whole bps)",
             "NSF > 1.0 (e.g., 46.0 instead of 0.0046)",
             "HARD STOP", LIGHT_RED, Font(bold=True, color="C0392B")),
        (5,  "Loan IDs — uniqueness",
             "Any Loan ID appears more than once",
             "HARD STOP", LIGHT_RED, Font(bold=True, color="C0392B")),
        (6,  "NSF — percent format",
             "NSF between 0.05–1.0 (may be 5%–100% rather than bps)",
             "YELLOW LIGHT", LIGHT_AMBER, Font(bold=True, color="7B3F00")),
        (7,  "NSF — investor range",
             "NSF outside expected band for that investor type",
             "YELLOW LIGHT", LIGHT_AMBER, Font(bold=True, color="7B3F00")),
        (8,  "Next Due Date",
             "Current-status loan with a next due date in the past",
             "YELLOW LIGHT", LIGHT_AMBER, Font(bold=True, color="7B3F00")),
        ("LAYER 2 — CROSS-PERIOD CHECKS (vs. PRIOR MONTH)", None, None, None, MID_BLUE, W),
        (9,  "Missing loans",
             "Prior-month loan absent with no PIF documentation",
             "HARD STOP", LIGHT_RED, Font(bold=True, color="C0392B")),
        (10, "Missing loans — PIF-confirmed",
             "Prior-month loan absent AND confirmed in PIF report",
             "CLEARED", LIGHT_GREEN, Font(bold=True, color="375623")),
        (11, "DQ status progression",
             "Status jumps more than one bucket (e.g., Current → 90+ DPD)",
             "YELLOW LIGHT", LIGHT_AMBER, Font(bold=True, color="7B3F00")),
        (12, "P&I payment change",
             "P&I increased > 10% month-over-month with no rate change",
             "YELLOW LIGHT", LIGHT_AMBER, Font(bold=True, color="7B3F00")),
        (13, "Remaining term",
             "Remaining term same as or greater than prior month",
             "YELLOW LIGHT", LIGHT_AMBER, Font(bold=True, color="7B3F00")),
        (14, "New add boarding",
             "Loan appears in submission but not in New Add recon report",
             "YELLOW LIGHT", LIGHT_AMBER, Font(bold=True, color="7B3F00")),
    ]

    r = 5
    for item in checks:
        ws2.row_dimensions[r].height = 36
        check_num, what, flag, sev, fill, font = item
        if what is None:
            # Category header
            merge_hdr(ws2, r, 1, 4, check_num, fill, font)
            ws2.row_dimensions[r].height = 20
        else:
            cell(ws2, r, 1, check_num, fill, font, CENTER)
            cell(ws2, r, 2, what,       fill, N,    LEFT)
            cell(ws2, r, 3, flag,       fill, N,    LEFT)
            cell(ws2, r, 4, sev,        fill, font, CENTER)
        r += 1

    set_col_widths(ws2, [6, 34, 52, 14])

    wb.save(path)
    print(f"  Saved: {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  WORKBOOK 2 — Blind Test Summary
# ══════════════════════════════════════════════════════════════════════════════

def build_blind_test(path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Scorecard ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Scorecard"
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 16
    merge_hdr(ws, 1, 1, 5,
              "MSR TAPE VALIDATOR — BLIND TEST RESULTS", NAVY, W)
    ws.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=14)
    merge_hdr(ws, 2, 1, 5,
              "Validator run against undisclosed injections — no prior knowledge of error types or locations",
              MID_BLUE, WN)
    merge_hdr(ws, 3, 1, 5,
              "Result:  8 for 8  — every manually injected error caught",
              PatternFill("solid", fgColor="196F3D"), Font(bold=True, color="FFFFFF", size=12))
    merge_hdr(ws, 4, 1, 5,
              "SIMULATED DATA — All loan information is synthetic and generated for testing purposes only.",
              SIMDATA_YEL, Font(bold=True, color="000000", size=9))

    # Summary table
    r = 6
    hdr(ws, r, 1, "Metric", NAVY, W, CENTER)
    hdr(ws, r, 2, "Known Injection Test", NAVY, W, CENTER)
    hdr(ws, r, 3, "Blind Test (LO Jitter)", NAVY, W, CENTER)
    hdr(ws, r, 4, "Delta", NAVY, W, CENTER)
    hdr(ws, r, 5, "Notes", NAVY, W, CENTER)
    ws.row_dimensions[r].height = 22
    r += 1

    scorecard = [
        ("Prior Month Loans",          "1,000",   "1,000",    "—",   "Same clean Dec 2025 tape"),
        ("Submission Loans (raw)",      "1,186",   "1,187",    "+1",  "One extra in blind test"),
        ("Duplicate Loan IDs",          "1",       "1",        "—",   "Same injected duplicate"),
        ("Missing Loans (total)",       "15",      "15",       "—",   "12 PIF + 3 unexplained in both"),
        ("  PIF-Explained (cleared)",   "12",      "12",       "—",   "Correctly cleared in both"),
        ("  Unexplained → Hard Stop",   "3",       "3",        "—",   "Same 3 removed without PIF"),
        ("New Adds (submitted)",        "200",     "201",      "+1",  "Blind: 1 phantom add (MSR300198)"),
        ("  Confirmed by New Add Recon","200",     "200",      "—",   ""),
        ("  Unconfirmed → Yellow Light","0",       "1",        "+1",  "Phantom loan caught"),
        ("HARD STOPS",                  "13",      "13",       "—",   "Same base hard stops"),
        ("YELLOW LIGHTS",               "9",       "17",       "+8",  "8 blind injections → yellow lights"),
        ("Loans Passing All Checks",    "1,166",   "1,161",    "−5",  ""),
        ("Injected Errors (undisclosed)","n/a",    "8",        "—",   "None known to validator in advance"),
        ("Errors Caught",               "n/a",     "8",        "—",   "8 for 8 — 100% catch rate"),
    ]

    fills_sc = [LIGHT_BLUE, LIGHT_BLUE, LIGHT_BLUE, LIGHT_BLUE, WHITE, WHITE,
                LIGHT_BLUE, WHITE, WHITE,
                LIGHT_RED, LIGHT_AMBER, LIGHT_GREEN, LIGHT_BLUE, LIGHT_GREEN]
    fonts_sc = [N]*10 + [Font(bold=True, color="C0392B"), Font(bold=True, color="7B3F00"),
                          Font(bold=True, color="375623"), N,
                          Font(bold=True, color="196F3D")]

    for i, (metric, known, blind, delta, note) in enumerate(scorecard):
        ws.row_dimensions[r].height = 18
        fill = fills_sc[i] if i < len(fills_sc) else WHITE
        font = fonts_sc[i] if i < len(fonts_sc) else N
        cell(ws, r, 1, metric, fill, font,                       LEFT)
        cell(ws, r, 2, known,  fill, Font(bold=True) if "STOP" in metric or "LIGHT" in metric else N, CENTER)
        cell(ws, r, 3, blind,  fill, Font(bold=True) if "STOP" in metric or "LIGHT" in metric else N, CENTER)
        cell(ws, r, 4, delta,  fill, N, CENTER)
        cell(ws, r, 5, note,   fill, Font(italic=True, color="555555"), LEFT)
        r += 1

    set_col_widths(ws, [30, 22, 22, 10, 42])

    # ── Sheet 2: Error Detail ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Blind Injections Detail")
    ws2.sheet_view.showGridLines = False

    ws2.row_dimensions[1].height = 30
    ws2.row_dimensions[2].height = 18
    ws2.row_dimensions[3].height = 16
    merge_hdr(ws2, 1, 1, 6,
              "BLIND TEST — INJECTION DETAIL", NAVY, W)
    ws2.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=13)
    merge_hdr(ws2, 2, 1, 6,
              "Errors manually constructed by Larkin O'Hern — validator had no prior knowledge",
              MID_BLUE, WN)
    merge_hdr(ws2, 3, 1, 6,
              "SIMULATED DATA — All loan information is synthetic and generated for testing purposes only.",
              SIMDATA_YEL, Font(bold=True, color="000000", size=9))

    for i, h in enumerate(["#", "Loan ID", "Injection Type", "What Was Changed",
                             "Rule Triggered", "Severity"], 1):
        hdr(ws2, 4, i, h, NAVY, W, CENTER)
    ws2.row_dimensions[4].height = 22

    injections = [
        (1, "MSR100005", "NSF Percent Format (stacked)",
         "NSF changed from 0.0046 to 0.5500\n(stacked on existing UPB hard stop)",
         "NSF May Be Expressed as Percent", "YELLOW LIGHT"),
        (2, "MSR100006", "NSF Percent Format (stacked)",
         "NSF changed from 0.0025 to 0.2500\n(stacked on existing UPB hard stop)",
         "NSF May Be Expressed as Percent", "YELLOW LIGHT"),
        (3, "MSR100015", "NSF Percent Format (clean loan)",
         "NSF changed from 0.0025 to 0.2500\non an otherwise clean FHLMC loan",
         "NSF May Be Expressed as Percent", "YELLOW LIGHT"),
        (4, "MSR100028", "Invalid Status Value",
         'Status changed from "60 DPD" to "60+ DPD"\n(realistic field encoding error)',
         "Invalid Status Value", "YELLOW LIGHT"),
        (5, "MSR300198", "Phantom Loan (unboarded)",
         "New loan ID not present in any recon file\n(appeared in submission with no boarding trail)",
         "Unboarded Loan — not in New Add report", "YELLOW LIGHT"),
        (6, "MSR100017", "Status Bucket Skip",
         "Status changed from Current → 90+ DPD\n(skipped 30 DPD and 60 DPD)",
         "Status Bucket Skip", "YELLOW LIGHT"),
        (7, "MSR100024", "Status Bucket Skip",
         "Status changed from Current → 90+ DPD\n(skipped 30 DPD and 60 DPD)",
         "Status Bucket Skip", "YELLOW LIGHT"),
        (8, "MSR100025", "Status Bucket Skip",
         "Status changed from Current → 90+ DPD\n(skipped 30 DPD and 60 DPD)",
         "Status Bucket Skip", "YELLOW LIGHT"),
    ]

    notable = {4, 5}   # rows to highlight as especially interesting
    r = 5
    for inj in injections:
        ws2.row_dimensions[r].height = 48
        num, loan, itype, what, rule, sev = inj
        fill = LIGHT_AMBER
        note_font = Font(bold=True, color="7B3F00") if num in notable else Font(bold=True, color="7B3F00")
        cell(ws2, r, 1, num,   fill, note_font, CENTER)
        cell(ws2, r, 2, loan,  fill, Font(bold=True), CENTER)
        cell(ws2, r, 3, itype, fill, note_font, LEFT)
        cell(ws2, r, 4, what,  fill, N, LEFT)
        cell(ws2, r, 5, rule,  fill, N, LEFT)
        cell(ws2, r, 6, sev,   fill, Font(bold=True, color="7B3F00"), CENTER)
        r += 1

    # Notable catches callout
    r += 1
    ws2.row_dimensions[r].height = 16
    merge_hdr(ws2, r, 1, 6,
              "Notable catches:", DARK_GREEN, W)
    r += 1
    ws2.row_dimensions[r].height = 28
    merge_hdr(ws2, r, 1, 6,
              '★  MSR100028 — "60+ DPD" typo caught as invalid status value '
              '(realistic encoding error; not a predefined injection type in the known test)',
              LIGHT_GREEN, Font(italic=True, color="375623"))
    r += 1
    ws2.row_dimensions[r].height = 28
    merge_hdr(ws2, r, 1, 6,
              "★  MSR300198 — Phantom loan flagged as unboarded "
              "(appeared in submission with no New Add or PIF trail; zero prior knowledge needed)",
              LIGHT_GREEN, Font(italic=True, color="375623"))

    set_col_widths(ws2, [4, 14, 28, 44, 38, 14])

    # ── Sheet 3: Validator Logic ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Validator Logic")
    ws3.sheet_view.showGridLines = False

    ws3.row_dimensions[1].height = 30
    ws3.row_dimensions[2].height = 18
    ws3.row_dimensions[3].height = 16
    merge_hdr(ws3, 1, 1, 4, "VALIDATOR ARCHITECTURE — TWO-LAYER DETECTION", NAVY, W)
    ws3.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=13)
    merge_hdr(ws3, 2, 1, 4,
              "validate_msr_tape.py — Python + openpyxl · One command · No manual steps",
              MID_BLUE, WN)
    merge_hdr(ws3, 3, 1, 4,
              "SIMULATED DATA — All loan information is synthetic and generated for testing purposes only.",
              SIMDATA_YEL, Font(bold=True, color="000000", size=9))

    # Layer 1
    r = 5
    ws3.row_dimensions[r].height = 22
    merge_hdr(ws3, r, 1, 4,
              "LAYER 1 — Standalone Field-Level Rules (applied to every loan in the submission)",
              PatternFill("solid", fgColor="833C00"), W)
    r += 1

    l1_rows = [
        ("UPB = $0 for active loan",        "Hard Stop",    "Loan marked active but has zero balance"),
        ("UPB > original balance",          "Hard Stop",    "Loan cannot grow without a modification"),
        ("Rate > 1.0 or < 0.005",           "Hard Stop",    "Whole-number or near-zero rate detected"),
        ("NSF > 1.0",                       "Hard Stop",    "Whole basis-point format detected"),
        ("Duplicate Loan IDs",              "Hard Stop",    "Same loan ID appears twice in submission"),
        ("NSF between 0.05–1.0",            "Yellow Light", "Possible percent format (5%–100%)"),
        ("NSF outside investor range",      "Yellow Light", "Value inconsistent with investor type"),
        ("Next Due Date in past (Current)", "Yellow Light", "Current loan with stale due date"),
        ("Invalid Status Value",            "Yellow Light", "Status not in allowed set"),
    ]
    for rule, sev, why in l1_rows:
        ws3.row_dimensions[r].height = 22
        fill = LIGHT_RED if sev == "Hard Stop" else LIGHT_AMBER
        font = Font(bold=True, color="C0392B") if sev == "Hard Stop" else Font(bold=True, color="7B3F00")
        cell(ws3, r, 1, rule, fill, font, LEFT)
        cell(ws3, r, 2, sev,  fill, font, CENTER)
        cell(ws3, r, 3, why,  fill, N, LEFT)
        ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        r += 1

    r += 1
    # Layer 2
    ws3.row_dimensions[r].height = 22
    merge_hdr(ws3, r, 1, 4,
              "LAYER 2 — Cross-Period Checks vs. Prior Month (continuing loans only)",
              PatternFill("solid", fgColor="375623"), W)
    r += 1

    l2_rows = [
        ("Missing loan — no PIF",           "Hard Stop",    "Prior-month loan absent; not in PIF report"),
        ("Missing loan — PIF confirmed",    "Cleared",      "Absence explained; auto-cleared from hard stops"),
        ("Status bucket skip",              "Yellow Light", "DQ status jumped more than one bucket"),
        ("P&I increased > 10%",             "Yellow Light", "Payment inflated vs prior month with no rate change"),
        ("Remaining term unchanged",        "Yellow Light", "Loan did not amortize as scheduled"),
        ("Unboarded new add",               "Yellow Light", "New loan not found in New Add recon report"),
    ]
    for rule, sev, why in l2_rows:
        ws3.row_dimensions[r].height = 22
        if sev == "Hard Stop":
            fill, font = LIGHT_RED, Font(bold=True, color="C0392B")
        elif sev == "Cleared":
            fill, font = LIGHT_GREEN, Font(bold=True, color="375623")
        else:
            fill, font = LIGHT_AMBER, Font(bold=True, color="7B3F00")
        cell(ws3, r, 1, rule, fill, font, LEFT)
        cell(ws3, r, 2, sev,  fill, font, CENTER)
        cell(ws3, r, 3, why,  fill, N, LEFT)
        ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        r += 1

    r += 1
    ws3.row_dimensions[r].height = 28
    merge_hdr(ws3, r, 1, 4,
              "Auto-discovery: validator scans its directory for Recon_PaidInFull_*.xlsx and Recon_NewAdds_*.xlsx. "
              "No manual file path configuration required.",
              LIGHT_BLUE, Font(italic=True, color="1F4E79", size=9), LEFT)

    set_col_widths(ws3, [36, 14, 44, 10])

    wb.save(path)
    print(f"  Saved: {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))

    print("Building content artifacts…")

    checklist_path   = os.path.join(script_dir, "Content_SubservicerChecklist.xlsx")
    blind_test_path  = os.path.join(script_dir, "Content_BlindTestSummary.xlsx")

    build_checklist(checklist_path)
    build_blind_test(blind_test_path)

    print("\nDone. Files created:")
    for p in [checklist_path, blind_test_path]:
        print(f"  {os.path.basename(p)}")
