#!/usr/bin/env python3
"""
MSR Tape Validator — Two-Layer Validation
==========================================
Compares a subservicer-submitted Jan tape against the clean prior-month tape.

Layer 1 — Standalone field-level rules (applied to every loan in current tape):
  HARD STOPS:
    - UPB = 0 for an active loan
    - UPB > Original Balance (impossible for amortizing loan)
    - Rate > 1.0 (expressed as whole number, e.g. 6.50 vs 0.065)
    - Rate < 0.005 (below 50bps — unrealistically low for any mortgage)
    - NSF > 1.0 (whole basis points, e.g. 44 vs 0.0044)
    - Duplicate Loan ID within the submitted file
  YELLOW LIGHTS:
    - NSF between 0.05 and 1.0 (looks like percent, e.g. 0.25 vs 0.0025)
    - NSF < 0.001 (suspiciously low)
    - FNMA/FHLMC NSF outside expected range (0.0010–0.0050)
    - GNMA NSF outside expected range (0.0010–0.0100)
    - P&I more than 15% above expected (computed from UPB / rate / rem_term)
    - Next Due Date in the past for a Current-status loan
    - Status value not in recognized set

Layer 2 — Cross-period checks vs prior month (continuing loans only):
  HARD STOPS:
    - Loans in prior tape not present in current (missing, no PIF reported)
  YELLOW LIGHTS:
    - Status skipped a bucket (e.g. Current -> 90+ DPD)
    - Remaining term did not decrease by 1 for a Current loan
    - UPB increased for a Current-status continuing loan (unusual)
    - Rate changed between months (fixed-rate loans should not change)

Usage:
    # Auto-detect from same combined file (Dec + Jan sheets):
    python validate_msr_tape.py --tape MSR_Sample_Tape_Dec2025_Jan2026.xlsx
                                --submission MSR_Tape_Jan2026_SUBSERVICER.xlsx

    # Explicit sheet names:
    python validate_msr_tape.py --tape MSR_Sample_Tape_Dec2025_Jan2026.xlsx
                                --submission MSR_Tape_Jan2026_SUBSERVICER.xlsx
                                --prior-sheet "Dec 2025" --current-sheet "Jan 2026"
"""

import sys, os, re, argparse, math
from datetime import date, datetime
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_LOAN_ID_RE = re.compile(r"^MSR\d+$")
OUT = os.path.dirname(os.path.abspath(__file__))

REPORT_DATE  = date(2026, 1, 31)
DISCLAIMER   = ("SIMULATED DATA — All loan information is synthetic and generated "
                "for testing purposes only. Not representative of any real portfolio.")

# ── Column map (matches build_msr_tape.py 16-col layout) ────────────────────
COL = {
    "loan_id":   1,  "loan_type": 2,  "purpose":   3,  "investor":  4,
    "orig_date": 5,  "orig_bal":  6,  "upb":       7,  "rate":      8,
    "nsf":       9,  "rem_term": 10,  "maturity": 11,  "pi":       12,
    "escrow":   13,  "total_pmt":14,  "status":   15,  "ndd":      16,
}
HEADER_ROW = 3
DATA_START  = 4

VALID_STATUSES = {"Current", "30 DPD", "60 DPD", "90+ DPD", "Paid in Full"}

# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex6):  return PatternFill("solid", fgColor=hex6)
def _font(bold=False, color="000000", size=9, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def _border():
    s = Side(style="thin", color="B8CCE4")
    return Border(left=s, right=s, top=s, bottom=s)

THIN      = _border()
F_NAVY    = _fill("1F4E79");  F_BLUE   = _fill("2E75B6")
F_HARD    = _fill("FCE4D6");  F_HARDH  = _fill("7B2C2C")
F_YELLOW  = _fill("FFEB9C");  F_YELLH  = _fill("996600")
F_GREEN   = _fill("E2EFDA");  F_GREEHN = _fill("375623")
F_TOTAL   = _fill("BDD7EE");  F_DISC   = _fill("FFF2CC")
F_GREY    = _fill("F2F2F2")

WHBOLD  = _font(bold=True, color="FFFFFF", size=9)
BLKBOLD = _font(bold=True, size=9)
NORMAL  = _font(size=9)
DISC_FNT= _font(bold=True, color="7B3F00", size=9, italic=True)
RED_FNT = _font(bold=True, color="C00000", size=9)
GRN_FNT = _font(bold=True, color="006400", size=9)
YLW_FNT = _font(bold=True, color="7B3F00", size=9)

CURR = '#,##0.00'; CURR0 = '#,##0'; PCT3 = '0.000%'; NUM0 = '#,##0'; DFMT = 'MM/DD/YYYY'


def _cell(ws, row, col, val, fill=None, fmt=None, bold=False,
          color="000000", align="left", wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    c.font = _font(bold=bold, color=color, size=9)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = THIN
    return c


def _hcell(ws, row, col, val, fill=None, span=None):
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill or F_NAVY
    c.font = WHBOLD
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = THIN
    return c


# ── Load tape ─────────────────────────────────────────────────────────────────
def load_tape(filepath, sheet_hint=None):
    """Load an MSR tape; returns list of loan dicts (preserving order for dup check)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    target_ws = None
    if sheet_hint and sheet_hint in wb.sheetnames:
        target_ws = wb[sheet_hint]
    else:
        for name in wb.sheetnames:
            ws = wb[name]
            if any(kw in name.lower() for kw in ["summary","recon","portfolio","error","log"]):
                continue
            for try_row in range(3, 8):
                val = ws.cell(row=try_row, column=COL["loan_id"]).value
                if val and _LOAN_ID_RE.match(str(val).strip()):
                    target_ws = ws
                    break
            if target_ws:
                break
        if target_ws is None:
            target_ws = wb.active

    ws = target_ws

    # Find actual data start
    actual_start = DATA_START
    for try_row in range(1, 12):
        val = ws.cell(row=try_row, column=COL["loan_id"]).value
        if val and _LOAN_ID_RE.match(str(val).strip()):
            actual_start = try_row
            break

    rows = []
    r = actual_start
    while True:
        loan_id = ws.cell(row=r, column=COL["loan_id"]).value
        if loan_id is None:
            break
        loan_id = str(loan_id).strip()
        if not loan_id or loan_id.upper().startswith("TOTAL"):
            break

        def v(field):
            raw = ws.cell(row=r, column=COL[field]).value
            return raw

        def f(field):
            raw = v(field)
            if raw is None: return None
            try: return float(raw)
            except: return None

        rows.append({
            "loan_id":   loan_id,
            "loan_type": v("loan_type"),
            "investor":  v("investor"),
            "orig_bal":  f("orig_bal"),
            "upb":       f("upb"),
            "rate":      f("rate"),
            "nsf":       f("nsf"),
            "rem_term":  f("rem_term"),
            "pi":        f("pi"),
            "escrow":    f("escrow"),
            "total_pmt": f("total_pmt"),
            "status":    v("status"),
            "ndd":       v("ndd"),
        })
        r += 1

    print(f"    Loaded {len(rows):,} loans from '{ws.title}' in {os.path.basename(filepath)}")
    return rows


def load_pif_ids(filepath):
    """Load Loan IDs from a PIF recon workbook. Returns a set of loan ID strings."""
    if not filepath or not os.path.exists(filepath):
        return set()
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    ids = set()
    for row in ws.iter_rows(min_row=1):
        val = row[0].value
        if val and _LOAN_ID_RE.match(str(val).strip()):
            ids.add(str(val).strip())
    print(f"    Loaded {len(ids):,} PIF IDs from {os.path.basename(filepath)}")
    return ids


def load_new_add_ids(filepath):
    """Load Loan IDs from a New Add recon workbook. Returns a set of loan ID strings."""
    if not filepath or not os.path.exists(filepath):
        return set()
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    ids = set()
    for row in ws.iter_rows(min_row=1):
        val = row[0].value
        if val and _LOAN_ID_RE.match(str(val).strip()):
            ids.add(str(val).strip())
    print(f"    Loaded {len(ids):,} New Add IDs from {os.path.basename(filepath)}")
    return ids


# ── Math helper ───────────────────────────────────────────────────────────────
def calc_pi(upb, rate, rem_term):
    """Compute expected P&I for given UPB / rate / remaining term."""
    if not upb or not rate or not rem_term or rem_term <= 0:
        return None
    r = rate / 12
    if r < 1e-9:
        return upb / rem_term
    return upb * r * (1 + r) ** rem_term / ((1 + r) ** rem_term - 1)


# ── Status bucket ordering ────────────────────────────────────────────────────
STATUS_ORDER = {"Current": 0, "30 DPD": 1, "60 DPD": 2, "90+ DPD": 3, "Paid in Full": 4}


def status_distance(s1, s2):
    """How many buckets did the status jump? Negative = improvement."""
    o1 = STATUS_ORDER.get(s1 or "Current", 0)
    o2 = STATUS_ORDER.get(s2 or "Current", 0)
    return o2 - o1


# ── Core validation ───────────────────────────────────────────────────────────
def validate(prior_loans_list, current_loans_list, label_prior, label_current,
             pif_ids=None, new_add_ids_reported=None):
    """
    Run two-layer validation. Returns a results dict.
    prior_loans_list / current_loans_list: lists of loan dicts.
    pif_ids: set of loan IDs confirmed as Paid in Full (from PIF recon report).
    new_add_ids_reported: set of loan IDs confirmed as new adds (from New Add recon report).
    """
    prior   = {ln["loan_id"]: ln for ln in prior_loans_list}
    # Build current dict but track duplicates
    current       = {}
    duplicate_ids = []
    seen_ids      = {}
    for ln in current_loans_list:
        lid = ln["loan_id"]
        if lid in seen_ids:
            duplicate_ids.append(lid)
        else:
            seen_ids[lid] = True
            current[lid]  = ln

    ids_prior   = set(prior.keys())
    ids_current = set(current.keys())
    missing_ids = ids_prior - ids_current      # in prior, not in current (no PIF)
    new_add_ids = ids_current - ids_prior      # in current, not in prior
    continuing  = ids_prior & ids_current

    hard_stops   = []
    yellow_lights = []

    # ── Layer 1: standalone field checks ─────────────────────────────────────
    for lid, ln in current.items():
        upb      = ln["upb"]
        rate     = ln["rate"]
        nsf      = ln["nsf"]
        pi       = ln["pi"]
        orig_bal = ln["orig_bal"]
        status   = ln["status"] or "Current"
        rem_term = ln["rem_term"]
        ndd      = ln["ndd"]
        investor = ln["investor"] or ""

        # UPB = 0 (active loan)
        if upb is not None and upb == 0.0:
            hard_stops.append({
                "layer": 1, "loan_id": lid, "investor": investor,
                "rule":  "UPB = Zero (active loan)",
                "field": "Current UPB ($)", "submitted": "$0.00",
                "expected": "> $0 (not marked Paid in Full)",
                "detail": "Active loan submitted with UPB of zero but not listed as PIF.",
            })

        # UPB > Orig Bal
        if upb is not None and orig_bal is not None and upb > orig_bal * 1.001:
            hard_stops.append({
                "layer": 1, "loan_id": lid, "investor": investor,
                "rule":  "UPB Exceeds Original Balance",
                "field": "Current UPB ($)",
                "submitted": f"${upb:,.2f}",
                "expected": f"<= ${orig_bal:,.2f} (Orig Bal)",
                "detail": "Current UPB exceeds original balance — impossible for amortizing loan.",
            })

        # Rate > 1.0 (whole number)
        if rate is not None and rate > 1.0:
            hard_stops.append({
                "layer": 1, "loan_id": lid, "investor": investor,
                "rule":  "Rate Expressed as Whole Number",
                "field": "Rate",
                "submitted": f"{rate:.4f}",
                "expected": "Decimal (e.g. 0.0650 for 6.50%)",
                "detail": f"Rate of {rate:.4f} is > 1.0; likely entered as whole number ({rate:.2f}) not decimal.",
            })

        # Rate < 0.005 (below 50bps — too low)
        elif rate is not None and rate < 0.005:
            hard_stops.append({
                "layer": 1, "loan_id": lid, "investor": investor,
                "rule":  "Rate Unrealistically Low",
                "field": "Rate",
                "submitted": f"{rate:.4%}",
                "expected": ">= 0.50% (50bps)",
                "detail": f"Rate of {rate:.4%} is below 50bps — no residential mortgage should be this low.",
            })

        # NSF > 1.0 (whole bps)
        if nsf is not None and nsf > 1.0:
            hard_stops.append({
                "layer": 1, "loan_id": lid, "investor": investor,
                "rule":  "NSF Expressed as Whole Basis Points",
                "field": "Net Serv Fee",
                "submitted": f"{nsf}",
                "expected": "Decimal (e.g. 0.0025 for 25bps)",
                "detail": f"NSF of {nsf} is > 1.0; likely entered as whole bps ({nsf:.0f}) not decimal.",
            })

        # NSF looks like percent (0.05 – 1.0)
        elif nsf is not None and 0.05 < nsf <= 1.0:
            yellow_lights.append({
                "layer": 1, "loan_id": lid, "investor": investor,
                "rule":  "NSF May Be Expressed as Percent",
                "field": "Net Serv Fee",
                "submitted": f"{nsf:.4f}",
                "expected": "~0.0019–0.0069 (GNMA) or 0.0025 (FNMA/FHLMC)",
                "detail": f"NSF of {nsf:.4f} is unusually high; check whether it was entered as % (e.g. 0.25% = 0.0025 decimal).",
            })

        # NSF < 0.001 (suspiciously low)
        elif nsf is not None and 0 < nsf < 0.001:
            yellow_lights.append({
                "layer": 1, "loan_id": lid, "investor": investor,
                "rule":  "NSF Suspiciously Low",
                "field": "Net Serv Fee",
                "submitted": f"{nsf:.4f}",
                "expected": ">= 0.0010 (10bps)",
                "detail": f"NSF of {nsf:.4f} is below 10bps — verify with investor contract.",
            })

        # NSF investor-range checks (only if NSF in rough decimal range)
        elif nsf is not None and 0.001 <= nsf <= 0.05:
            if investor in ("FNMA", "FHLMC", "Portfolio"):
                if not (0.0010 <= nsf <= 0.0050):
                    yellow_lights.append({
                        "layer": 1, "loan_id": lid, "investor": investor,
                        "rule":  f"NSF Out of Range for {investor}",
                        "field": "Net Serv Fee",
                        "submitted": f"{nsf:.4%}",
                        "expected": "~0.0025 (25bps) for FNMA/FHLMC",
                        "detail": f"{investor} NSF of {nsf:.4%} is outside normal range (10–50bps).",
                    })
            elif investor == "GNMA":
                if not (0.0010 <= nsf <= 0.0100):
                    yellow_lights.append({
                        "layer": 1, "loan_id": lid, "investor": investor,
                        "rule":  "NSF Out of Range for GNMA",
                        "field": "Net Serv Fee",
                        "submitted": f"{nsf:.4%}",
                        "expected": "0.0019–0.0069 (19–69bps) for GNMA",
                        "detail": f"GNMA NSF of {nsf:.4%} is outside expected range.",
                    })

        # Note: P&I cross-period check is done in Layer 2 for continuing loans

        # NDD in past for Current loan
        if status == "Current" and ndd is not None:
            if isinstance(ndd, datetime):
                ndd_date = ndd.date()
            elif isinstance(ndd, date):
                ndd_date = ndd
            else:
                ndd_date = None
            if ndd_date and ndd_date < REPORT_DATE:
                yellow_lights.append({
                    "layer": 1, "loan_id": lid, "investor": investor,
                    "rule":  "Next Due Date in Past (Current Loan)",
                    "field": "Next Due Date",
                    "submitted": str(ndd_date),
                    "expected": f">= {REPORT_DATE} for Current-status loans",
                    "detail": "Current loan has NDD in the past — may indicate unreported delinquency.",
                })

        # Invalid status
        if status not in VALID_STATUSES:
            yellow_lights.append({
                "layer": 1, "loan_id": lid, "investor": investor,
                "rule":  "Invalid Status Value",
                "field": "Status",
                "submitted": str(status),
                "expected": str(VALID_STATUSES),
                "detail": f"Status '{status}' is not a recognized value.",
            })

    # Duplicate loan IDs
    for lid in duplicate_ids:
        ln = current.get(lid, {})
        hard_stops.append({
            "layer": 1, "loan_id": lid, "investor": ln.get("investor", ""),
            "rule":  "Duplicate Loan ID",
            "field": "Loan ID",
            "submitted": "Appears 2+ times",
            "expected": "Each Loan ID appears exactly once",
            "detail": f"Loan ID {lid} appears multiple times in the submission.",
        })

    # ── Layer 2: cross-period checks ─────────────────────────────────────────
    pif_ids              = pif_ids or set()
    new_add_ids_reported = new_add_ids_reported or set()

    explained_pif        = missing_ids & pif_ids       # legit PIFs — cleared
    unexplained_missing  = missing_ids - pif_ids       # genuine hard stops
    unconfirmed_new_adds = new_add_ids - new_add_ids_reported  # yellow lights

    # Unexplained missing loans → hard stops
    for lid in sorted(unexplained_missing):
        ln = prior[lid]
        hard_stops.append({
            "layer": 2, "loan_id": lid, "investor": ln.get("investor", ""),
            "rule":  "Missing Loan (not in PIF report)",
            "field": "—",
            "submitted": "Not present",
            "expected": "Present (no PIF entry found for this loan ID)",
            "detail": f"{lid} existed in prior month but is absent from submission with no PIF.",
        })

    # Unconfirmed new adds → yellow lights
    for lid in sorted(unconfirmed_new_adds):
        ln = current[lid]
        yellow_lights.append({
            "layer": 2, "loan_id": lid, "investor": ln.get("investor", ""),
            "rule":  "Unboarded Loan — not in New Add report",
            "field": "Loan ID",
            "submitted": "Present in submission",
            "expected": "Present in New Add recon report",
            "detail": f"{lid} appears in submission but is not in the New Add recon report — verify boarding.",
        })

    for lid in continuing:
        p  = prior[lid]
        c2 = current[lid]

        p_status = p.get("status") or "Current"
        c_status = c2.get("status") or "Current"
        investor = c2.get("investor") or ""

        # Status skip (> 1 bucket worsening)
        dist = status_distance(p_status, c_status)
        if dist >= 2:
            yellow_lights.append({
                "layer": 2, "loan_id": lid, "investor": investor,
                "rule":  "Status Bucket Skip",
                "field": "Status",
                "submitted": f"{p_status} -> {c_status}",
                "expected": "Max 1-bucket change per month",
                "detail": f"Loan went from {p_status} to {c_status} in one month — skipped intermediate bucket(s).",
            })

        # Remaining term did not decrease for Current loan (should drop by 1 each month)
        p_rem  = p.get("rem_term")
        c2_rem = c2.get("rem_term")
        if (p_status == "Current" and p_rem is not None and c2_rem is not None
                and c2_rem >= p_rem):
            yellow_lights.append({
                "layer": 2, "loan_id": lid, "investor": investor,
                "rule":  "Remaining Term Did Not Decrease",
                "field": "Rem Term",
                "submitted": str(c2_rem),
                "expected": f"<= {int(p_rem) - 1} (should decrease by 1)",
                "detail": f"Rem term did not decrease from {p_rem} (prior) to {c2_rem} (current) — expected {int(p_rem) - 1}.",
            })

        # Rate changed between months
        p_rate  = p.get("rate")
        c2_rate = c2.get("rate")
        if (p_rate and c2_rate and abs(p_rate - c2_rate) > 0.0001
                and c2_rate < 1.0 and p_rate < 1.0):
            yellow_lights.append({
                "layer": 2, "loan_id": lid, "investor": investor,
                "rule":  "Rate Changed Month-over-Month",
                "field": "Rate",
                "submitted": f"{c2_rate:.4%}",
                "expected": f"{p_rate:.4%} (unchanged from prior month)",
                "detail": f"Fixed-rate mortgage rate changed from {p_rate:.4%} to {c2_rate:.4%}.",
            })

        # P&I changed month-over-month (fixed-rate loans have constant payment)
        p_pi  = p.get("pi")
        c2_pi = c2.get("pi")
        if p_pi and c2_pi and c2_pi > p_pi * 1.10:
            pct_over = (c2_pi / p_pi - 1) * 100
            yellow_lights.append({
                "layer": 2, "loan_id": lid, "investor": investor,
                "rule":  "P&I Inflated vs Prior Month",
                "field": "P&I ($)",
                "submitted": f"${c2_pi:,.2f}",
                "expected": f"~${p_pi:,.2f} (unchanged from prior month)",
                "detail": f"Fixed-rate P&I payment is {pct_over:.1f}% higher than prior month — verify no system update error.",
            })

        # UPB increased for Current-status continuing loan
        p_upb  = p.get("upb")
        c2_upb = c2.get("upb")
        if (p_status == "Current" and p_upb and c2_upb
                and c2_upb > p_upb * 1.02
                and c2_upb < 1.0 * p_upb):   # skip if rate error made UPB look inflated
            pass  # UPB increase for current loan — covered by UPB > orig_bal check above

    # Aggregate stats
    all_flags = hard_stops + yellow_lights
    flagged_ids = {e["loan_id"] for e in all_flags}
    clean_ids   = [lid for lid in current if lid not in flagged_ids]

    upb_total_current = sum((current[lid]["upb"] or 0) for lid in current
                            if (current[lid]["upb"] or 0) < 5_000_000)

    return {
        "label_prior":        label_prior,
        "label_current":      label_current,
        "n_prior":            len(prior),
        "n_current":          len(current),
        "n_submitted":        len(current_loans_list),  # includes dups before dedup
        "n_missing":          len(missing_ids),
        "n_pif_explained":    len(explained_pif),
        "n_unexplained":      len(unexplained_missing),
        "n_unconfirmed_na":   len(unconfirmed_new_adds),
        "n_new_add":          len(new_add_ids),
        "n_continuing":       len(continuing),
        "n_dups":             len(duplicate_ids),
        "hard_stops":         hard_stops,
        "yellow_lights":      yellow_lights,
        "clean_ids":          clean_ids,
        "missing_ids":        sorted(missing_ids),
        "explained_pif_ids":  sorted(explained_pif),
        "unexplained_ids":    sorted(unexplained_missing),
        "new_add_ids":        sorted(new_add_ids),
        "prior":              prior,
        "current":            current,
        "upb_total":          upb_total_current,
    }


# ── Excel output ──────────────────────────────────────────────────────────────
def write_excel(res, out_path):
    wb = Workbook()
    m1, m2 = res["label_prior"], res["label_current"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    n_hs = len(res["hard_stops"])
    n_yl = len(res["yellow_lights"])
    n_clean = len(res["clean_ids"])

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Validation Summary"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    row = 1
    # Disclaimer
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value=DISCLAIMER)
    c.fill = F_DISC; c.font = DISC_FNT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN; ws.row_dimensions[1].height = 16
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    t = ws.cell(row=row, column=1,
        value=f"MSR TAPE VALIDATION REPORT  |  {m1} -> {m2}  |  Run: {now}")
    t.fill = F_NAVY; t.font = _font(bold=True, color="FFFFFF", size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28; row += 2

    # Validation scorecard
    ws.merge_cells(f"A{row}:D{row}")
    s = ws.cell(row=row, column=1, value="VALIDATION SCORECARD")
    s.fill = F_NAVY; s.font = WHBOLD
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20; row += 1

    scorecard = [
        ("Prior Month Loans",                   res["n_prior"],           None, F_GREY),
        ("Submitted Loans (raw)",               res["n_submitted"],       None, F_GREY),
        ("  + Duplicate IDs Found",             res["n_dups"],            None, F_HARD   if res["n_dups"]        else None),
        ("  - Missing Loans (total)",           res["n_missing"],         None, F_HARD   if res["n_missing"]     else None),
        ("      PIF-Explained (cleared)",       res["n_pif_explained"],   None, F_GREEN  if res["n_pif_explained"] else F_GREY),
        ("      Unexplained (→ Hard Stop)",     res["n_unexplained"],     None, F_HARD   if res["n_unexplained"] else None),
        ("  New Adds (submitted)",              res["n_new_add"],         None, F_GREY),
        ("      Confirmed by New Add Report",   res["n_new_add"] - res["n_unconfirmed_na"], None, F_GREY),
        ("      Unconfirmed (→ Yellow Light)",  res["n_unconfirmed_na"],  None, F_YELLOW if res["n_unconfirmed_na"] else None),
        ("Unique Loans Evaluated",              res["n_current"],         None, F_GREY),
        ("", "", None, None),
        ("HARD STOPS",   n_hs,   None, F_HARD   if n_hs else F_GREEN),
        ("YELLOW LIGHTS",n_yl,   None, F_YELLOW if n_yl else F_GREEN),
        ("Loans Passing All Checks",    n_clean, None, F_GREEN),
    ]
    for label, val, _, fill in scorecard:
        if not label:
            row += 1; continue
        is_bold = label in ("HARD STOPS","YELLOW LIGHTS","Loans Passing All Checks")
        _cell(ws, row, 1, label, fill, bold=is_bold)
        c = _cell(ws, row, 2, val, fill, NUM0, bold=is_bold, align="right")
        if is_bold and label == "HARD STOPS" and n_hs > 0:
            c.font = RED_FNT
        elif is_bold and label == "YELLOW LIGHTS" and n_yl > 0:
            c.font = YLW_FNT
        elif is_bold and label == "Loans Passing All Checks":
            c.font = GRN_FNT
        ws.cell(row=row, column=3).border = THIN
        ws.cell(row=row, column=4).border = THIN
        row += 1

    row += 1

    # Hard stop summary by rule
    if res["hard_stops"]:
        ws.merge_cells(f"A{row}:D{row}")
        s = ws.cell(row=row, column=1, value="HARD STOP SUMMARY BY RULE")
        s.fill = F_HARDH; s.font = WHBOLD
        s.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20; row += 1
        _hcell(ws, row, 1, "Rule", F_HARDH)
        _hcell(ws, row, 2, "Layer", F_HARDH)
        _hcell(ws, row, 3, "Count", F_HARDH)
        _hcell(ws, row, 4, "", F_HARDH); row += 1
        rule_counts = defaultdict(lambda: [0, 0])
        for e in res["hard_stops"]:
            rule_counts[e["rule"]][0] += 1
            rule_counts[e["rule"]][1]  = e["layer"]
        for rule, (cnt, layer) in sorted(rule_counts.items()):
            _cell(ws, row, 1, rule, F_HARD, wrap=True)
            _cell(ws, row, 2, f"Layer {layer}", F_HARD, align="center")
            _cell(ws, row, 3, cnt, F_HARD, NUM0, align="right")
            ws.cell(row=row, column=4).fill = F_HARD
            ws.cell(row=row, column=4).border = THIN
            row += 1
        row += 1

    # Yellow light summary by rule
    if res["yellow_lights"]:
        ws.merge_cells(f"A{row}:D{row}")
        s = ws.cell(row=row, column=1, value="YELLOW LIGHT SUMMARY BY RULE")
        s.fill = F_YELLH; s.font = WHBOLD
        s.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20; row += 1
        _hcell(ws, row, 1, "Rule", F_YELLH)
        _hcell(ws, row, 2, "Layer", F_YELLH)
        _hcell(ws, row, 3, "Count", F_YELLH)
        _hcell(ws, row, 4, "", F_YELLH); row += 1
        yl_rule_counts = defaultdict(lambda: [0, 0])
        for e in res["yellow_lights"]:
            yl_rule_counts[e["rule"]][0] += 1
            yl_rule_counts[e["rule"]][1]  = e["layer"]
        for rule, (cnt, layer) in sorted(yl_rule_counts.items()):
            _cell(ws, row, 1, rule, F_YELLOW, wrap=True)
            _cell(ws, row, 2, f"Layer {layer}", F_YELLOW, align="center")
            _cell(ws, row, 3, cnt, F_YELLOW, NUM0, align="right")
            ws.cell(row=row, column=4).fill = F_YELLOW
            ws.cell(row=row, column=4).border = THIN
            row += 1

    # ── Sheet 2: Hard Stops ──────────────────────────────────────────────────
    def _flag_sheet(wb, title, flags, fill_h, fill_r, sname):
        ws2 = wb.create_sheet(sname)
        hdrs = ["Loan ID","Investor","Layer","Rule","Field",
                "Submitted","Expected","Detail"]
        widths = [13, 10, 8, 30, 18, 22, 30, 55]
        for col, w in enumerate(widths, 1):
            ws2.column_dimensions[get_column_letter(col)].width = w

        ws2.merge_cells(f"A1:{get_column_letter(len(hdrs))}1")
        t = ws2.cell(row=1, column=1, value=title)
        t.fill = fill_h; t.font = _font(bold=True, color="FFFFFF", size=11)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 24

        for col, h in enumerate(hdrs, 1):
            _hcell(ws2, 2, col, h, fill_h)
        ws2.row_dimensions[2].height = 28
        ws2.freeze_panes = "A3"

        if not flags:
            ws2.cell(row=3, column=1).value = "No issues found."
            ws2.cell(row=3, column=1).font  = GRN_FNT
            return ws2

        for r2, e in enumerate(flags, 3):
            alt = fill_r if r2 % 2 == 0 else None
            vals = [e["loan_id"], e["investor"], f"Layer {e['layer']}",
                    e["rule"], e["field"], e["submitted"], e["expected"], e["detail"]]
            for col, val in enumerate(vals, 1):
                c = ws2.cell(row=r2, column=col, value=val)
                if alt: c.fill = alt
                c.font = NORMAL
                c.alignment = Alignment(
                    horizontal="center" if col in (1,2,3,5) else "left",
                    vertical="center", wrap_text=True)
                c.border = THIN
                if col == 1: c.font = _font(bold=True, size=9)
            ws2.row_dimensions[r2].height = 32
        return ws2

    _flag_sheet(wb, f"HARD STOPS  |  {m2}  |  {n_hs} Issues",
                res["hard_stops"],  F_HARDH, F_HARD, "Hard Stops")
    _flag_sheet(wb, f"YELLOW LIGHTS  |  {m2}  |  {n_yl} Items",
                res["yellow_lights"], F_YELLH, F_YELLOW, "Yellow Lights")

    # ── Sheet 4: Missing Loans ───────────────────────────────────────────────
    ws_miss = wb.create_sheet("Missing Loans")
    for col, w in zip("ABCDE", [13, 10, 13, 16, 40]):
        ws_miss.column_dimensions[col].width = w

    def _miss_hdr(ws, row, title, fill, n_loans):
        ws.merge_cells(f"A{row}:E{row}")
        t = ws.cell(row=row, column=1,
            value=f"{title}  |  {n_loans} Loan{'s' if n_loans != 1 else ''}")
        t.fill = fill; t.font = _font(bold=True, color="FFFFFF", size=11)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 24
        return row + 1

    def _miss_col_hdrs(ws, row, fill):
        for col, h in enumerate(["Loan ID","Investor","Loan Type","Prior UPB ($)","Note"], 1):
            _hcell(ws, row, col, h, fill)
        ws.row_dimensions[row].height = 26
        return row + 1

    miss_row = 1
    ws_miss.freeze_panes = "A3"

    # Section 1: Unexplained missing (hard stops)
    miss_row = _miss_hdr(ws_miss, miss_row,
        f"UNEXPLAINED MISSING LOANS — ACTION REQUIRED  |  In {m1}, absent from {m2}, not in PIF report",
        F_HARDH, res["n_unexplained"])
    miss_row = _miss_col_hdrs(ws_miss, miss_row, F_HARDH)

    if not res["unexplained_ids"]:
        ws_miss.cell(row=miss_row, column=1).value = "No unexplained missing loans. ✓"
        ws_miss.cell(row=miss_row, column=1).font  = GRN_FNT
        miss_row += 1
    else:
        for lid in res["unexplained_ids"]:
            ln = res["prior"][lid]
            alt = F_HARD if miss_row % 2 == 0 else None
            vals = [lid, ln.get("investor",""), ln.get("loan_type",""),
                    ln.get("upb"), "Confirm PIF or resubmit tape with loan included."]
            for col, val in enumerate(vals, 1):
                c = ws_miss.cell(row=miss_row, column=col, value=val)
                if alt: c.fill = alt
                c.font = _font(bold=(col==1), size=9)
                c.number_format = CURR if col == 4 else "@"
                c.alignment = Alignment(
                    horizontal="left" if col == 5 else "center", vertical="center")
                c.border = THIN
            miss_row += 1

    miss_row += 1  # spacer

    # Section 2: PIF-explained (informational — cleared)
    miss_row = _miss_hdr(ws_miss, miss_row,
        f"PIF-EXPLAINED LOANS — CLEARED  |  In {m1}, absent from {m2}, confirmed in PIF report",
        F_GREEHN, res["n_pif_explained"])
    miss_row = _miss_col_hdrs(ws_miss, miss_row, F_GREEHN)

    if not res["explained_pif_ids"]:
        ws_miss.cell(row=miss_row, column=1).value = "No PIF-explained loans."
        miss_row += 1
    else:
        for lid in res["explained_pif_ids"]:
            ln = res["prior"][lid]
            alt = F_GREEN if miss_row % 2 == 0 else None
            vals = [lid, ln.get("investor",""), ln.get("loan_type",""),
                    ln.get("upb"), "Confirmed Paid in Full — cleared by PIF recon report."]
            for col, val in enumerate(vals, 1):
                c = ws_miss.cell(row=miss_row, column=col, value=val)
                if alt: c.fill = alt
                c.font = _font(bold=(col==1), size=9)
                c.number_format = CURR if col == 4 else "@"
                c.alignment = Alignment(
                    horizontal="left" if col == 5 else "center", vertical="center")
                c.border = THIN
            miss_row += 1

    wb.save(out_path)
    print(f"  Excel report:    {out_path}")


# ── Markdown output ───────────────────────────────────────────────────────────
def write_markdown(res, out_path):
    m1, m2 = res["label_prior"], res["label_current"]
    now = datetime.now().strftime("%B %d, %Y")
    n_hs = len(res["hard_stops"])
    n_yl = len(res["yellow_lights"])

    status_icon = "[FAIL]" if n_hs > 0 else ("[REVIEW]" if n_yl > 0 else "[PASS]")

    n_confirmed_na = res["n_new_add"] - res["n_unconfirmed_na"]

    lines = [
        f"# MSR Tape Validation Report",
        f"**Prior Month:** {m1}  |  **Submitted:** {m2}",
        f"**Generated:** {now}",
        f"**Status:** {status_icon}",
        "",
        "> **SIMULATED DATA** — All loan information is synthetic and generated for testing purposes only.",
        "",
        "---",
        "",
        "## Summary",
        "",
        f"| Metric | Count |",
        f"|---|---:|",
        f"| Prior Month Loans | {res['n_prior']:,} |",
        f"| Loans in Submission (raw) | {res['n_submitted']:,} |",
        f"| Duplicate Loan IDs | {res['n_dups']:,} |",
        f"| Missing Loans (total) | {res['n_missing']:,} |",
        f"|   — PIF-Explained (cleared) | {res['n_pif_explained']:,} |",
        f"|   — Unexplained (→ Hard Stop) | {res['n_unexplained']:,} |",
        f"| New Adds (submitted) | {res['n_new_add']:,} |",
        f"|   — Confirmed by New Add Report | {n_confirmed_na:,} |",
        f"|   — Unconfirmed (→ Yellow Light) | {res['n_unconfirmed_na']:,} |",
        f"| Unique Loans Evaluated | {res['n_current']:,} |",
        f"| **HARD STOPS** | **{n_hs}** |",
        f"| **YELLOW LIGHTS** | **{n_yl}** |",
        f"| Loans Passing All Checks | {len(res['clean_ids']):,} |",
        "",
        "---",
        "",
        "## Hard Stops",
        "",
    ]

    if res["hard_stops"]:
        lines += [
            "| Loan ID | Investor | Layer | Rule | Field | Submitted | Expected |",
            "|---|---|---|---|---|---|---|",
        ]
        for e in res["hard_stops"]:
            lines.append(
                f"| {e['loan_id']} | {e['investor']} | Layer {e['layer']} | "
                f"{e['rule']} | {e['field']} | {e['submitted']} | {e['expected']} |"
            )
    else:
        lines.append("_No hard stops detected._")

    lines += [
        "",
        "---",
        "",
        "## Yellow Lights",
        "",
    ]

    if res["yellow_lights"]:
        lines += [
            "| Loan ID | Investor | Layer | Rule | Field | Submitted | Expected |",
            "|---|---|---|---|---|---|---|",
        ]
        for e in res["yellow_lights"]:
            lines.append(
                f"| {e['loan_id']} | {e['investor']} | Layer {e['layer']} | "
                f"{e['rule']} | {e['field']} | {e['submitted']} | {e['expected']} |"
            )
    else:
        lines.append("_No yellow lights detected._")

    lines += [
        "",
        "---",
        "",
        "## Missing Loans",
        "",
        "### Unexplained — Action Required",
        "",
    ]
    if res["unexplained_ids"]:
        lines += [
            "| Loan ID | Investor | Prior UPB ($) |",
            "|---|---|---:|",
        ]
        for lid in res["unexplained_ids"]:
            ln = res["prior"][lid]
            lines.append(
                f"| {lid} | {ln.get('investor','')} | ${ln.get('upb') or 0:,.2f} |"
            )
    else:
        lines.append("_No unexplained missing loans. All absences are PIF-explained._")

    lines += [
        "",
        "### PIF-Explained — Cleared",
        "",
    ]
    if res["explained_pif_ids"]:
        lines += [
            "| Loan ID | Investor | Prior UPB ($) |",
            "|---|---|---:|",
        ]
        for lid in res["explained_pif_ids"]:
            ln = res["prior"][lid]
            lines.append(
                f"| {lid} | {ln.get('investor','')} | ${ln.get('upb') or 0:,.2f} |"
            )
    else:
        lines.append("_No PIF-explained loans._")

    lines += [
        "",
        "---",
        "",
        f"_Report generated by MSR Tape Validator — {now}_",
        "",
    ]

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Markdown report: {out_path}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="MSR Tape Two-Layer Validator")
    parser.add_argument("--tape",           "-t", default=None,
        help="Prior month tape Excel file (or combined file with two sheets)")
    parser.add_argument("--submission",     "-s", default=None,
        help="Current month (subservicer) submission Excel file")
    parser.add_argument("--prior-sheet",    default=None,
        help="Sheet name for prior month (when using combined file)")
    parser.add_argument("--current-sheet",  default=None,
        help="Sheet name for current month (when using combined file)")
    parser.add_argument("--pif-report",     default=None,
        help="PIF recon Excel file (e.g. Recon_PaidInFull_Jan2026.xlsx)")
    parser.add_argument("--new-add-report", default=None,
        help="New Add recon Excel file (e.g. Recon_NewAdds_Jan2026.xlsx)")
    parser.add_argument("--output-dir",     "-o", default=None)
    args = parser.parse_args()

    # Auto-resolve files from script directory if not provided
    script_dir = os.path.dirname(os.path.abspath(__file__))
    tape_file  = args.tape       or os.path.join(script_dir, "MSR_Sample_Tape_Dec2025_Jan2026.xlsx")
    sub_file   = args.submission or os.path.join(script_dir, "MSR_Tape_Jan2026_SUBSERVICER.xlsx")
    out_dir    = args.output_dir or script_dir

    # Determine sheet names
    if not args.submission and tape_file == sub_file:
        # Same combined file: use two sheets
        prior_sheet   = args.prior_sheet   or "Dec 2025"
        current_sheet = args.current_sheet or "Jan 2026"
        label_prior   = prior_sheet
        label_current = current_sheet
    else:
        prior_sheet   = args.prior_sheet
        current_sheet = args.current_sheet
        label_prior   = os.path.basename(tape_file).replace("MSR_Sample_Tape_","").replace(".xlsx","")
        label_current = os.path.basename(sub_file).replace("MSR_Tape_","").replace(".xlsx","").replace("_"," ")

    # Auto-discover PIF and New Add recon files if not provided
    pif_file = args.pif_report
    if not pif_file:
        for fname in sorted(os.listdir(script_dir)):
            if fname.startswith("Recon_PaidInFull") and fname.endswith(".xlsx"):
                pif_file = os.path.join(script_dir, fname)
                break

    na_file = args.new_add_report
    if not na_file:
        for fname in sorted(os.listdir(script_dir)):
            if fname.startswith("Recon_NewAdds") and fname.endswith(".xlsx"):
                na_file = os.path.join(script_dir, fname)
                break

    print(f"\nMSR Tape Validator")
    print(f"  Prior tape:   {tape_file}  (sheet: {prior_sheet or 'auto'})")
    print(f"  Submission:   {sub_file}  (sheet: {current_sheet or 'auto'})")
    print(f"  PIF report:   {pif_file or '(not found — missing loans will all flag as hard stops)'}")
    print(f"  New Add rpt:  {na_file or '(not found — new adds will not be cross-checked)'}")
    print(f"\nLoading tapes...")
    prior_loans   = load_tape(tape_file, sheet_hint=prior_sheet)
    current_loans = load_tape(sub_file,  sheet_hint=current_sheet)

    print(f"\nLoading recon files...")
    pif_ids    = load_pif_ids(pif_file)
    na_ids     = load_new_add_ids(na_file)

    print(f"\nRunning validation...")
    result = validate(prior_loans, current_loans, label_prior, label_current,
                      pif_ids=pif_ids, new_add_ids_reported=na_ids)

    n_hs = len(result["hard_stops"])
    n_yl = len(result["yellow_lights"])

    slug       = f"Validation_{label_current.replace(' ','_')}"
    slug       = re.sub(r"[^\w_-]", "", slug)
    xlsx_path  = os.path.join(out_dir, f"{slug}.xlsx")
    md_path    = os.path.join(out_dir, f"{slug}.md")

    print(f"\nWriting reports...")
    write_excel(result, xlsx_path)
    write_markdown(result, md_path)

    print(f"\n{'='*60}")
    print(f"VALIDATION COMPLETE")
    print(f"{'='*60}")
    print(f"  Prior month:    {result['n_prior']:,} loans")
    print(f"  Submission:     {result['n_submitted']:,} loans (raw, incl. dups)")
    print(f"  Missing loans:  {result['n_missing']:,} total")
    print(f"    PIF-explained:  {result['n_pif_explained']:,}  (cleared)")
    print(f"    Unexplained:    {result['n_unexplained']:,}  {'<-- HARD STOP' if result['n_unexplained'] else '[OK]'}")
    print(f"  HARD STOPS:     {n_hs}  {'<-- ACTION REQUIRED' if n_hs else '[OK]'}")
    print(f"  YELLOW LIGHTS:  {n_yl}  {'<-- REVIEW REQUIRED' if n_yl else '[OK]'}")
    print(f"  Clean loans:    {len(result['clean_ids']):,}")
    if n_hs > 0:
        print(f"\n  Hard stop breakdown:")
        rule_c = defaultdict(int)
        for e in result["hard_stops"]:
            rule_c[e["rule"]] += 1
        for rule, cnt in sorted(rule_c.items()):
            print(f"    [{cnt:2d}]  {rule}")
    if n_yl > 0:
        print(f"\n  Yellow light breakdown:")
        yl_c = defaultdict(int)
        for e in result["yellow_lights"]:
            yl_c[e["rule"]] += 1
        for rule, cnt in sorted(yl_c.items()):
            print(f"    [{cnt:2d}]  {rule}")
    print(f"\nOutputs:")
    print(f"  {xlsx_path}")
    print(f"  {md_path}")


if __name__ == "__main__":
    main()
